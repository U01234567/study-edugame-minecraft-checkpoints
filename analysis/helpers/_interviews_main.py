from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from ._shared import INTERVIEW_MANIFEST_PATH, INTERVIEW_TRANSCRIPTS_DIR, clean

DEFAULT_INTERVIEW_CATEGORIES = {
    "low_prior_game_experience": "Low prior game experience; preferably from the required continue condition.",
    "high_prior_game_experience": "High prior game experience; preferably from the optional pauses condition.",
    "required_pause_chapter_change": "Required pause participant with a clear chapter-to-chapter change after the pause.",
    "optional_pause_chose_pause": "Optional pause participant who chose at least one 2-minute pause.",
    "many_disallowed_empty_disabled_clicks": "Participant with many disallowed, empty, or disabled clicks during instructions or gameplay.",
    "unusual_task_performance": "Participant with unusually high or low task performance.",
    "high_total_movement_distance": "Participant with high or low total movement distance.",
    "age_gender_coverage": "Participant selected to improve age and gender coverage across the interview sample.",
    "other_intuition": "Other / intuition.",
}


def read_manifest(path: Path = INTERVIEW_MANIFEST_PATH) -> dict[str, Any]:
    if not path.exists():
        return {"category_definitions": DEFAULT_INTERVIEW_CATEGORIES, "interviews": []}
    data = json.loads(path.read_text(encoding="utf-8"))
    data.setdefault("category_definitions", DEFAULT_INTERVIEW_CATEGORIES)
    data.setdefault("interviews", [])
    return data


def interview_no_from_filename(path: Path) -> int | None:
    match = re.search(r"interview[_\-\s]*0*(\d+)", path.stem, flags=re.IGNORECASE)
    if not match:
        return None
    return int(match.group(1))


def manifest_entry_for_file(manifest: dict[str, Any], path: Path) -> dict[str, Any]:
    interviews = manifest.get("interviews", {})
    interview_no = interview_no_from_filename(path)

    if isinstance(interviews, dict) and interview_no is not None:
        entry = interviews.get(str(interview_no)) or interviews.get(f"{interview_no:02d}")
        if isinstance(entry, dict):
            return {"interview_no": interview_no, **entry}

    if isinstance(interviews, list):
        for entry in interviews:
            if clean(entry.get("filename")).lower() == path.name.lower():
                return entry

    return {"interview_no": interview_no}


def read_transcript_xlsx(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    header_row = None
    for row in rows:
        headers = [clean(cell) for cell in row]
        if any(headers):
            header_row = headers
            break

    if not header_row:
        return []

    header_lookup = {
        header.lower(): index
        for index, header in enumerate(header_row)
        if header
    }
    speaker_index = header_lookup.get("speaker")
    transcript_index = header_lookup.get("transcript")

    turns: list[dict[str, str]] = []
    for row in rows:
        speaker = clean(row[speaker_index]) if speaker_index is not None and speaker_index < len(row) else ""
        transcript = clean(row[transcript_index]) if transcript_index is not None and transcript_index < len(row) else ""
        if speaker or transcript:
            turns.append({"speaker": speaker, "transcript": transcript})
    return turns


def participant_ids_from_turns(turns: list[dict[str, str]]) -> list[str]:
    ids: list[str] = []
    for turn in turns:
        speaker = clean(turn.get("speaker"))
        if speaker and speaker.lower() != "researcher" and speaker not in ids:
            ids.append(speaker)
    return ids


def age_group(age: object) -> str:
    try:
        value = int(float(clean(age)))
    except ValueError:
        return "Unknown age"
    if value < 18:
        return "<18"
    if value <= 24:
        return "18-24"
    if value <= 34:
        return "25-34"
    if value <= 44:
        return "35-44"
    return "45+"


def compact_participant_type(participant_id: str, lookup: dict[str, dict[str, Any]]) -> dict[str, str]:
    participant = lookup.get(participant_id, {})
    return {
        "participant_id": participant_id,
        "condition": clean(participant.get("condition")) or "Unknown condition",
        "gender": clean(participant.get("gender")) or "Unknown gender",
        "age_group": age_group(participant.get("age")),
    }


def load_interview_overview(
    transcripts_dir: Path = INTERVIEW_TRANSCRIPTS_DIR,
    participants: list[dict[str, Any]] | None = None,
    manifest_path: Path = INTERVIEW_MANIFEST_PATH,
) -> dict[str, Any]:
    manifest = read_manifest(manifest_path)
    categories = manifest.get("category_definitions", DEFAULT_INTERVIEW_CATEGORIES)
    participant_lookup = {
        clean(row.get("participant_id")): row
        for row in (participants or [])
        if clean(row.get("participant_id"))
    }

    if not transcripts_dir.exists():
        return {
            "available": False,
            "n_files": 0,
            "n_turns": 0,
            "unique_participant_ids": [],
            "category_rows": [],
            "type_rows": [],
            "transcripts": [],
            "notes": ["Interview transcript directory not found."],
        }

    transcript_paths = sorted(path for path in transcripts_dir.glob("*.xlsx") if not path.name.startswith("~$"))
    transcripts: list[dict[str, Any]] = []
    category_to_ids: dict[str, set[str]] = defaultdict(set)
    category_to_files: dict[str, set[str]] = defaultdict(set)
    category_to_transcript_ids: dict[str, list[str]] = defaultdict(list)
    type_counter: Counter[tuple[str, str, str, str]] = Counter()
    all_participant_ids: set[str] = set()
    total_turns = 0
    notes: list[str] = []

    for index, path in enumerate(transcript_paths, start=1):
        turns = read_transcript_xlsx(path)
        speaker_ids = participant_ids_from_turns(turns)
        all_participant_ids.update(speaker_ids)
        total_turns += len(turns)

        manifest_entry = manifest_entry_for_file(manifest, path)
        selection_categories = [
            clean(item)
            for item in manifest_entry.get("selection_categories", [])
            if clean(item)
        ]

        unknown_categories = [item for item in selection_categories if item not in categories]
        if unknown_categories:
            notes.append(f"{path.name}: unknown manifest categories: {', '.join(unknown_categories)}")

        participant_types = [compact_participant_type(pid, participant_lookup) for pid in speaker_ids]

        transcript_id = f"interview-{index:02d}"

        for category in selection_categories:
            category_to_files[category].add(path.name)
            if transcript_id not in category_to_transcript_ids[category]:
                category_to_transcript_ids[category].append(transcript_id)

        for ptype in participant_types:
            for category in selection_categories:
                category_to_ids[category].add(ptype["participant_id"])

            type_counter[
                (
                    ptype["condition"],
                    ptype["gender"],
                    ptype["age_group"],
                    ", ".join(selection_categories) or "No manifest category",
                )
            ] += 1

        transcripts.append({
            "transcript_id": transcript_id,
            "filename": path.name,
            "interview_no": manifest_entry.get("interview_no") or index,
            "title": " + ".join(speaker_ids) if speaker_ids else path.stem,
            "speaker_ids": speaker_ids,
            "speaker_count": len(speaker_ids),
            "n_turns": len(turns),
            "selection_categories": selection_categories,
            "category_labels": [categories.get(category, category) for category in selection_categories],
            "notes": clean(manifest_entry.get("notes")),
            "participant_types": participant_types,
            "turns": turns,
        })

    transcripts_by_id = {transcript["transcript_id"]: transcript for transcript in transcripts}

    def slot_payload(transcript_id: str | None) -> dict[str, Any]:
        if not transcript_id or transcript_id not in transcripts_by_id:
            return {
                "label": "",
                "transcript_id": "",
                "filename": "",
                "speaker_ids": [],
                "n_turns": 0,
            }

        transcript = transcripts_by_id[transcript_id]
        speaker_ids = transcript.get("speaker_ids") or []
        return {
            "label": " + ".join(speaker_ids) if speaker_ids else transcript.get("filename", ""),
            "transcript_id": transcript["transcript_id"],
            "filename": transcript.get("filename", ""),
            "speaker_ids": speaker_ids,
            "n_turns": transcript.get("n_turns", 0),
        }

    category_rows = []
    for category, definition in categories.items():
        ids = sorted(category_to_ids.get(category, set()))
        files = sorted(category_to_files.get(category, set()))
        transcript_ids = category_to_transcript_ids.get(category, [])
        assigned_slots = [slot_payload(transcript_id) for transcript_id in transcript_ids]
        printed_slots = [slot_payload(transcript_ids[index] if index < len(transcript_ids) else None) for index in range(3)]

        category_rows.append({
            "category": category,
            "definition": definition,
            "slots": assigned_slots,
            "slot_1": printed_slots[0],
            "slot_2": printed_slots[1],
            "slot_3": printed_slots[2],
            "overflow_count": max(0, len(transcript_ids) - 3),
            "n_participants": len(ids),
            "n_interviews": len(files),
            "participant_ids": ", ".join(ids),
            "files": ", ".join(files),
        })

    type_rows = [
        {
            "condition": condition,
            "gender": gender,
            "age_group": group,
            "selection_categories": selection_categories,
            "n_participants": count,
        }
        for (condition, gender, group, selection_categories), count in sorted(type_counter.items())
    ]

    return {
        "available": bool(transcript_paths),
        "n_files": len(transcript_paths),
        "n_turns": total_turns,
        "unique_participant_ids": sorted(all_participant_ids),
        "category_rows": category_rows,
        "type_rows": type_rows,
        "transcripts": transcripts,
        "notes": notes,
    }