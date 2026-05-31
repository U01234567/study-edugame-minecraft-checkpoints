from __future__ import annotations

import csv
import json
import re
import shutil
import subprocess
import tempfile
import uuid
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Any

from openpyxl import load_workbook

from ._delayed_response_filter import DELAYED_INCLUDED_COLUMN, annotate_survey_rows_with_delayed_inclusion, build_delayed_response_checklist
from ._logs_main import LOG_LINE_RE
from ._shared import clean, canonical_condition
from ._survey_io import detect_text_encoding, looks_like_qualtrics_label_row

DEFAULT_IDENTITY_FILE = Path.home() / ".minecraft-study" / "minecraft-study-logs-age-key.txt"
MAX_ARCHIVE_DEPTH = 5
COPY_CHUNK_SIZE = 1024 * 1024

MCID_IN_FILENAME_RE = re.compile(r"-([23456789ABCDEFGHJKMNPQRSTVWXYZ]{4}-[23456789ABCDEFGHJKMNPQRSTVWXYZ]{4})\.log$", re.IGNORECASE)
WINDOWS_LOCAL_PATH_RE = re.compile(r"(?<![A-Za-z0-9])(?:[A-Za-z]:[\\/][^\s|,;]*)")
LOCAL_UNIX_PATH_RE = re.compile(r"(^|[\s('\"=])/(Users|home|tmp|var|mnt|private|Applications|Volumes)/")
USER_HOME_PATH_RE = re.compile(r"(^|[\s('\"=])~/")
HTTP_URL_RE = re.compile(r"^https?://", re.IGNORECASE)
HTTP_URL_ANY_RE = re.compile(r"https?://[^\s|]+", re.IGNORECASE)
FILE_URI_RE = re.compile(r"(^|[\s('\"=])file:(/{2,})?", re.IGNORECASE)
SURVEY_COLUMNS_TO_REMOVE = {
    "IPAddress",
    "Finished",
    "RecordedDate",
    "RecipientLastName",
    "RecipientFirstName",
    "RecipientEmail",
    "ExternalReference",
    "LocationLatitude",
    "LocationLongitude",
    "DistributionChannel",
    "UserLanguage",
}
CONFIG_FILENAMES = [
    "collection_locations.json",
    "interview_manifest.json",
    "retention_rubrics.json",
]
RETENTION_SOURCE_NAMES = [
    "retention_scoring.tsv",
    "retention_scoring.csv",
    "retention_scores.tsv",
    "retention_scores.csv",
]


@dataclass
class ExtractionResult:
    copied: list[Path] = field(default_factory=list)
    skipped_existing: list[Path] = field(default_factory=list)
    ignored_members: list[str] = field(default_factory=list)
    decrypted_archives: list[Path] = field(default_factory=list)
    processed_archives: list[Path] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


@dataclass
class RuntimeContext:
    logs_dir: Path
    identity_file: Path
    work_dir: Path
    age_command_checked: bool = False
    age_command_available: bool | None = None


@dataclass
class PublishSummary:
    included_count: int = 0
    logs_written: int = 0
    transcripts_written: int = 0
    survey_rows_written: int = 0
    configs_written: int = 0
    retention_rows_written: int | None = None
    warnings: list[str] = field(default_factory=list)

    def lines(self) -> list[str]:
        rows = [
            f"Included MCIDs loaded: {self.included_count}",
            f"Publishable logs written: {self.logs_written}",
            f"Publishable transcripts written: {self.transcripts_written}",
            f"Publishable survey rows written: {self.survey_rows_written}",
            f"Config files written: {self.configs_written}",
        ]
        if self.retention_rows_written is not None:
            rows.append(f"Retention scoring rows written: {self.retention_rows_written}")
        rows.extend(f"WARNING: {warning}" for warning in self.warnings)
        return rows


def is_age_file(path: Path) -> bool:
    return path.name.lower().endswith(".age")


def is_zip_file(path: Path) -> bool:
    return path.suffix.lower() == ".zip"


def supported_archive_paths(logs_dir: Path) -> list[Path]:
    if not logs_dir.exists():
        return []
    return sorted(
        [path for path in logs_dir.iterdir() if path.is_file() and (is_age_file(path) or is_zip_file(path))],
        key=lambda item: (item.stat().st_mtime, item.name),
    )


def decrypted_zip_path(age_file: Path, output_dir: Path) -> Path:
    name = age_file.name
    while name.lower().endswith(".age"):
        name = name[:-4]
    if not name.lower().endswith(".zip"):
        name = f"{name}.zip"
    return output_dir / name


def require_age_command(context: RuntimeContext) -> None:
    if context.age_command_checked:
        if context.age_command_available:
            return
        raise RuntimeError("The age command was not found. Install it first with: winget install -e --id FiloSottile.age")

    context.age_command_checked = True
    context.age_command_available = shutil.which("age") is not None
    if not context.age_command_available:
        raise RuntimeError("The age command was not found. Install it first with: winget install -e --id FiloSottile.age")


def require_identity_file(identity_file: Path) -> None:
    if not identity_file.exists():
        raise FileNotFoundError(f"Identity file not found: {identity_file}")


def decrypt_age_file(age_file: Path, output_zip: Path, identity_file: Path) -> bool:
    if output_zip.exists() and output_zip.stat().st_mtime >= age_file.stat().st_mtime:
        return False

    output_zip.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(prefix=f"{output_zip.stem}-", suffix=".tmp", dir=output_zip.parent, delete=False) as temp_file:
        temp_path = Path(temp_file.name)

    try:
        subprocess.run(
            [
                "age",
                "--decrypt",
                "--identity",
                str(identity_file),
                "--output",
                str(temp_path),
                str(age_file),
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        temp_path.replace(output_zip)
        return True
    except subprocess.CalledProcessError as exc:
        temp_path.unlink(missing_ok=True)
        message = exc.stderr.strip() or exc.stdout.strip() or f"exit code {exc.returncode}"
        raise RuntimeError(f"age failed while decrypting {age_file.name}: {message}") from exc
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def safe_member_name(member_name: str) -> str:
    cleaned = member_name.replace("\\", "/").strip("/")
    return Path(cleaned).name.replace("\x00", "").strip()


def is_study_log_name(filename: str) -> bool:
    return filename.startswith("study-") and filename.lower().endswith(".log")


def copy_zip_member_to_file(archive: zipfile.ZipFile, member: zipfile.ZipInfo, target_path: Path) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with archive.open(member, "r") as source, target_path.open("wb") as destination:
        shutil.copyfileobj(source, destination, length=COPY_CHUNK_SIZE)


def copy_zip_member_to_temp(archive: zipfile.ZipFile, member: zipfile.ZipInfo, member_name: str, work_dir: Path) -> Path:
    temp_path = work_dir / f"nested-{uuid.uuid4().hex}-{member_name}"
    copy_zip_member_to_file(archive, member, temp_path)
    return temp_path


def process_age_file(age_file: Path, context: RuntimeContext, result: ExtractionResult, depth: int, keep_decrypted_zip: bool) -> None:
    require_identity_file(context.identity_file)
    require_age_command(context)
    output_dir = context.logs_dir if keep_decrypted_zip else context.work_dir
    output_zip = decrypted_zip_path(age_file, output_dir)
    decrypt_age_file(age_file, output_zip, context.identity_file)
    result.decrypted_archives.append(output_zip)
    process_zip_file(output_zip, context, result, depth=depth + 1)


def process_zip_file(zip_path: Path, context: RuntimeContext, result: ExtractionResult, depth: int) -> None:
    if depth > MAX_ARCHIVE_DEPTH:
        result.errors.append(f"Skipped {zip_path.name}: archive nesting exceeded depth {MAX_ARCHIVE_DEPTH}.")
        return
    if not zipfile.is_zipfile(zip_path):
        result.errors.append(f"Skipped {zip_path.name}: not a valid zip file.")
        return

    result.processed_archives.append(zip_path)
    try:
        with zipfile.ZipFile(zip_path, "r") as archive:
            for member in archive.infolist():
                if member.is_dir():
                    continue

                member_name = safe_member_name(member.filename)
                if not member_name:
                    result.ignored_members.append(member.filename)
                    continue

                if is_study_log_name(member_name):
                    target_path = context.logs_dir / member_name
                    if target_path.exists():
                        result.skipped_existing.append(target_path)
                        continue
                    copy_zip_member_to_file(archive, member, target_path)
                    result.copied.append(target_path)
                    continue

                if is_age_file(Path(member_name)) or is_zip_file(Path(member_name)):
                    try:
                        nested_path = copy_zip_member_to_temp(archive, member, member_name, context.work_dir)
                        process_input_file(nested_path, context, result, depth=depth + 1, keep_decrypted_zip=False)
                    except Exception as exc:
                        result.errors.append(f"Could not process nested archive {member.filename} in {zip_path.name}: {exc}")
                    continue

                result.ignored_members.append(member.filename)
    except zipfile.BadZipFile:
        result.errors.append(f"Skipped {zip_path.name}: bad zip file.")


def process_input_file(input_file: Path, context: RuntimeContext, result: ExtractionResult, depth: int = 0, keep_decrypted_zip: bool = True) -> None:
    if depth > MAX_ARCHIVE_DEPTH:
        result.errors.append(f"Skipped {input_file.name}: archive nesting exceeded depth {MAX_ARCHIVE_DEPTH}.")
        return
    if not input_file.exists():
        result.errors.append(f"Input file not found: {input_file}")
        return
    if is_age_file(input_file):
        process_age_file(input_file, context, result, depth=depth, keep_decrypted_zip=keep_decrypted_zip)
        return
    if is_zip_file(input_file):
        process_zip_file(input_file, context, result, depth=depth)
        return
    result.errors.append(f"Unsupported input file type: {input_file.name}")


def decrypt_all_log_archives(logs_dir: Path, identity_file: Path = DEFAULT_IDENTITY_FILE) -> ExtractionResult:
    logs_dir.mkdir(parents=True, exist_ok=True)
    archives = supported_archive_paths(logs_dir)
    result = ExtractionResult()
    if not archives:
        return result

    with tempfile.TemporaryDirectory(prefix="minecraft-study-analysis-") as temp_dir:
        context = RuntimeContext(logs_dir=logs_dir, identity_file=identity_file, work_dir=Path(temp_dir))
        for archive_path in archives:
            try:
                process_input_file(archive_path, context, result, keep_decrypted_zip=True)
            except Exception as exc:
                result.errors.append(f"Could not process {archive_path.name}: {exc}")
    return result


def require_path(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Expected {label} at {path}")


def read_delimited_rows(path: Path) -> tuple[list[list[str]], str]:
    encoding = detect_text_encoding(path)
    with path.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        return list(csv.reader(handle, delimiter=delimiter)), delimiter


def parse_log_body_parts(line: str) -> tuple[str, str, str, str, list[tuple[str, str]]] | None:
    match = LOG_LINE_RE.match(line)
    if not match:
        return None
    parts = [part.strip() for part in match.group("body").split("|")]
    if not parts or not parts[0]:
        return None
    fields: list[tuple[str, str]] = []
    for part in parts[1:]:
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        fields.append((clean(key), clean(value)))
    return match.group("date"), match.group("time"), match.group("body"), parts[0], fields


def fields_to_dict(fields: Iterable[tuple[str, str]]) -> dict[str, str]:
    return {key: value for key, value in fields}


def event_has_choice_agree(event: str, fields: list[tuple[str, str]]) -> bool:
    return event == "consent_choice" and fields_to_dict(fields).get("choice") == "agree_and_continue"


def is_questionnaire_button(event: str) -> bool:
    return event == "questionnaire_button_pressed"


def mcid_from_log(path: Path) -> str:
    try:
        for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
            parsed = parse_log_body_parts(line)
            if parsed is None:
                continue
            _date, _time, _raw_activity, _event, fields = parsed
            session_id = fields_to_dict(fields).get("session_id", "")
            if session_id:
                return session_id
    except OSError:
        pass
    match = MCID_IN_FILENAME_RE.search(path.name)
    return match.group(1).upper() if match else ""


def newest_by_mtime(paths: list[Path]) -> Path:
    return sorted(paths, key=lambda item: (item.stat().st_mtime, item.name))[-1]


def build_raw_log_index(raw_log_dir: Path, included_mcids: Iterable[str] | None = None) -> dict[str, Path]:
    require_path(raw_log_dir, "raw/logs directory")
    included_set = set(included_mcids or [])
    grouped: dict[str, list[Path]] = {}
    for path in sorted(raw_log_dir.glob("study-*.log")):
        mcid = mcid_from_log(path)
        if not mcid:
            continue
        if included_set and mcid not in included_set:
            continue
        grouped.setdefault(mcid, []).append(path)
    return {mcid: newest_by_mtime(paths) for mcid, paths in grouped.items()}


def raw_condition_from_log(path: Path) -> str:
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        parsed = parse_log_body_parts(line)
        if parsed is None:
            continue
        _date, _time, _raw_activity, event, fields = parsed
        if event != "experiment_condition_assigned":
            continue
        raw_condition = fields_to_dict(fields).get("condition", "")
        return canonical_condition(raw_condition) or raw_condition
    return ""


def build_condition_lookup(included_mcids: Iterable[str], log_index: dict[str, Path], summary: PublishSummary) -> dict[str, str]:
    condition_lookup: dict[str, str] = {}
    for mcid in included_mcids:
        raw_log_path = log_index.get(mcid)
        if raw_log_path is None:
            continue
        condition = raw_condition_from_log(raw_log_path)
        if condition:
            condition_lookup[mcid] = condition
        else:
            summary.warnings.append(f"No condition assignment found in raw log for {mcid}: {raw_log_path.name}")
    return condition_lookup


def text_without_http_urls(value: str) -> str:
    return HTTP_URL_ANY_RE.sub("", value)


def local_path_leak_snippet(value: str) -> str | None:
    scan_value = text_without_http_urls(clean(value))
    if not scan_value.strip():
        return None
    for pattern in (WINDOWS_LOCAL_PATH_RE, FILE_URI_RE, USER_HOME_PATH_RE, LOCAL_UNIX_PATH_RE):
        match = pattern.search(scan_value)
        if match:
            return match.group(0).strip()
    return None


def field_value_is_path_like(value: str) -> bool:
    value = clean(value).strip().strip("'\"")
    if not value:
        return False
    scan_value = text_without_http_urls(value)
    if not scan_value.strip():
        return False
    if local_path_leak_snippet(scan_value) is not None:
        return True
    if "\\" in scan_value:
        return True
    return False


def should_remove_log_field(raw_part: str) -> bool:
    if "=" not in raw_part:
        return False
    _raw_key, raw_value = raw_part.split("=", 1)
    return field_value_is_path_like(raw_value)


def sanitise_log_activity(raw_activity: str) -> str:
    parts = raw_activity.split(" | ")
    if len(parts) <= 1:
        return raw_activity
    kept_parts = [parts[0]]
    removed_any = False
    for part in parts[1:]:
        if should_remove_log_field(part):
            removed_any = True
            continue
        kept_parts.append(part)
    return " | ".join(kept_parts) if removed_any else raw_activity


def assert_publishable_text(path: Path) -> None:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            activity = row.get("Activity", "")
            leak = local_path_leak_snippet(activity)
            if leak is not None:
                raise ValueError(f"{path} row {row_number} still contains local path-like text: {leak!r}")
            for part in activity.split(" | ")[1:]:
                if "=" not in part:
                    continue
                key, value = part.split("=", 1)
                if field_value_is_path_like(value):
                    raise ValueError(f"{path} row {row_number} still contains local path value in field {clean(key)!r}")


def publishable_log_rows(raw_log_path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    started = False
    stopped = False
    for line in raw_log_path.read_text(encoding="utf-8", errors="replace").splitlines():
        parsed = parse_log_body_parts(line)
        if parsed is None:
            continue
        date, time, raw_activity, event, fields = parsed
        if not started:
            if event_has_choice_agree(event, fields):
                started = True
            continue
        activity = sanitise_log_activity(raw_activity)
        rows.append({"Date": date, "Time": time, "Activity": activity})
        if is_questionnaire_button(event):
            stopped = True
            break
    if not started:
        raise ValueError(f"No consent_choice | choice=agree_and_continue line found in {raw_log_path}")
    if not stopped:
        raise ValueError(f"No questionnaire_button_pressed line found after consent in {raw_log_path}")
    if not rows:
        raise ValueError(f"No publishable rows were produced from {raw_log_path}")
    return rows


def clear_generated_files(directory: Path, suffixes: tuple[str, ...]) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    for path in directory.iterdir():
        if path.is_file() and path.suffix.lower() in suffixes:
            path.unlink()


def write_log_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["Date", "Time", "Activity"], lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    assert_publishable_text(path)


def publish_logs(included_mcids: list[str], log_index: dict[str, Path], data_log_dir: Path, summary: PublishSummary) -> None:
    clear_generated_files(data_log_dir, (".csv",))
    missing = [mcid for mcid in included_mcids if mcid not in log_index]
    if missing:
        raise FileNotFoundError(f"Missing raw log(s) for included MCID(s): {', '.join(missing)}")
    for mcid in included_mcids:
        rows = publishable_log_rows(log_index[mcid])
        write_log_csv(data_log_dir / f"{mcid}.csv", rows)
        summary.logs_written += 1


def trim_trailing_empty(values: list[object]) -> list[str]:
    rendered = ["" if value is None else str(value) for value in values]
    while rendered and not clean(rendered[-1]):
        rendered.pop()
    return rendered


def read_xlsx_visible_rows(path: Path) -> list[list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = [trim_trailing_empty(list(row)) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    while rows and not any(clean(cell) for cell in rows[-1]):
        rows.pop()
    return rows


def find_header_row_with_speaker(rows: list[list[str]]) -> tuple[int, int]:
    for row_index, row in enumerate(rows[:30]):
        for column_index, cell in enumerate(row):
            if clean(cell).lower() == "speaker":
                return row_index, column_index
    raise ValueError("No 'Speaker' column found")


def transcript_speakers(rows: list[list[str]], header_index: int, speaker_column: int) -> set[str]:
    speakers: set[str] = set()
    for row in rows[header_index + 1:]:
        if speaker_column >= len(row):
            continue
        speaker = clean(row[speaker_column])
        if not speaker or speaker.lower() == "researcher":
            continue
        speakers.add(speaker)
    return speakers


def write_rows_csv(path: Path, rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    width = max((len(row) for row in rows), default=0)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        for row in rows:
            writer.writerow(row + [""] * (width - len(row)))


def publish_transcripts(raw_transcripts_dir: Path, data_transcripts_dir: Path, included_mcids: set[str], summary: PublishSummary) -> None:
    if not raw_transcripts_dir.exists():
        summary.warnings.append(f"No raw transcripts directory found at {raw_transcripts_dir}")
        return
    clear_generated_files(data_transcripts_dir, (".csv",))
    for path in sorted(raw_transcripts_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        rows = read_xlsx_visible_rows(path)
        if not rows:
            summary.warnings.append(f"Skipped empty transcript workbook: {path.name}")
            continue
        try:
            header_index, speaker_column = find_header_row_with_speaker(rows)
        except ValueError as exc:
            summary.warnings.append(f"Skipped {path.name}: {exc}")
            continue
        speakers = transcript_speakers(rows, header_index, speaker_column)
        if not speakers:
            summary.warnings.append(f"Skipped {path.name}: no participant speakers found")
            continue
        if not speakers.issubset(included_mcids):
            excluded = ", ".join(sorted(speakers - included_mcids))
            summary.warnings.append(f"Skipped {path.name}: contains non-included MCID(s): {excluded}")
            continue
        write_rows_csv(data_transcripts_dir / f"{path.stem}.csv", rows[header_index:])
        summary.transcripts_written += 1


def newest_tsv_file(directory: Path) -> Path:
    require_path(directory, "raw/survey directory")
    candidates = sorted(
        (path for path in directory.glob("*.tsv") if path.is_file()),
        key=lambda item: (item.stat().st_mtime, item.name),
    )
    if not candidates:
        raise FileNotFoundError(f"No .tsv files found in {directory}")
    return candidates[-1]


def looks_like_qualtrics_import_row(row: list[str]) -> bool:
    if not row:
        return False
    import_cells = sum(1 for cell in row if '"ImportId"' in clean(cell))
    return import_cells >= max(3, len(row) // 4)


def load_qualtrics_rows_without_metadata(path: Path) -> tuple[list[str], list[list[str]]]:
    encoding = detect_text_encoding(path)
    with path.open("r", encoding=encoding, newline="") as handle:
        raw_rows = list(csv.reader(handle, delimiter="\t"))
    if not raw_rows:
        raise ValueError(f"Survey export is empty: {path}")

    # Keep the original export header exactly as written by Qualtrics.
    # The publish step must only remove explicit rows and columns; it must not
    # rewrite names such as age, gender, SEEN, or ret_slot##_... .
    header = [clean(cell) for cell in raw_rows[0]]

    data_start_index = 1
    if len(raw_rows) > data_start_index and looks_like_qualtrics_label_row(raw_rows[data_start_index]):
        data_start_index += 1
    if len(raw_rows) > data_start_index and looks_like_qualtrics_import_row(raw_rows[data_start_index]):
        data_start_index += 1

    data_rows = [row for row in raw_rows[data_start_index:] if any(clean(cell) for cell in row)]
    return header, data_rows


def selected_column_indexes(header: list[str]) -> list[int]:
    remove_lookup = {name.lower() for name in SURVEY_COLUMNS_TO_REMOVE}
    return [index for index, column in enumerate(header) if clean(column).lower() not in remove_lookup]


def index_of_column(header: list[str], column_name: str) -> int:
    for index, column in enumerate(header):
        if clean(column).lower() == column_name.lower():
            return index
    raise ValueError(f"Survey export is missing required column: {column_name}")


def publish_survey(raw_survey_dir: Path, data_survey_path: Path, included_mcids: set[str], condition_lookup: dict[str, str], summary: PublishSummary) -> None:
    raw_survey_path = newest_tsv_file(raw_survey_dir)
    header, rows = load_qualtrics_rows_without_metadata(raw_survey_path)
    mcid_index = index_of_column(header, "MCID")
    excluded_publish_columns = {"condition", DELAYED_INCLUDED_COLUMN.lower()}
    keep_indexes = [
        index
        for index in selected_column_indexes(header)
        if clean(header[index]).lower() not in excluded_publish_columns
    ]
    output_header = [header[index] for index in keep_indexes] + ["condition"]
    output_rows: list[list[str]] = []
    delayed_source_rows: list[dict[str, str]] = []

    for row in rows:
        padded = row + [""] * max(0, len(header) - len(row))
        mcid = clean(padded[mcid_index])
        if mcid not in included_mcids:
            continue
        output_row = [padded[index] for index in keep_indexes]
        output_row.append(condition_lookup.get(mcid, ""))
        output_rows.append(output_row)
        delayed_source_rows.append({output_header[index]: value for index, value in enumerate(output_row)})

    delayed_checklist = build_delayed_response_checklist(delayed_source_rows)
    annotated_rows = annotate_survey_rows_with_delayed_inclusion(delayed_source_rows, delayed_checklist)

    data_survey_path.parent.mkdir(parents=True, exist_ok=True)
    with data_survey_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow(output_header + [DELAYED_INCLUDED_COLUMN])
        for output_row, annotated_row in zip(output_rows, annotated_rows):
            writer.writerow(output_row + [annotated_row.get(DELAYED_INCLUDED_COLUMN, "")])
    output_text = data_survey_path.read_text(encoding="utf-8")
    if "@" in output_text:
        raise ValueError(f"{data_survey_path} contains '@' after removing known personal-data columns")
    summary.survey_rows_written = len(output_rows)


def json_source_path(filename: str, *, raw_config_dir: Path, data_config_dir: Path, resources_dir: Path) -> Path | None:
    for path in (raw_config_dir / filename, data_config_dir / filename, resources_dir / filename):
        if path.exists():
            return path
    return None


def publish_config(*, raw_config_dir: Path, data_config_dir: Path, resources_dir: Path, summary: PublishSummary) -> None:
    data_config_dir.mkdir(parents=True, exist_ok=True)
    for filename in CONFIG_FILENAMES:
        source = json_source_path(filename, raw_config_dir=raw_config_dir, data_config_dir=data_config_dir, resources_dir=resources_dir)
        if source is None:
            summary.warnings.append(f"No source found for config file {filename}")
            continue
        payload = json.loads(source.read_text(encoding="utf-8"))
        target = data_config_dir / filename
        target.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        summary.configs_written += 1


def retention_source_path(*, raw_config_dir: Path, raw_dir: Path, data_dir: Path) -> Path | None:
    for directory in (raw_config_dir, raw_dir, data_dir):
        for filename in RETENTION_SOURCE_NAMES:
            path = directory / filename
            if path.exists():
                return path
    return None


def publish_retention_scoring(*, raw_config_dir: Path, raw_dir: Path, data_dir: Path, included_mcids: set[str], summary: PublishSummary) -> None:
    source = retention_source_path(raw_config_dir=raw_config_dir, raw_dir=raw_dir, data_dir=data_dir)
    target = data_dir / "retention_scoring.tsv"
    if source is None:
        summary.retention_rows_written = None
        summary.warnings.append("No retention scoring file found to publish; score_retention.py can create data/retention_scoring.tsv later.")
        return
    rows, delimiter = read_delimited_rows(source)
    if not rows:
        summary.retention_rows_written = 0
        return
    header = [clean(cell) for cell in rows[0]]
    lower_header = [cell.lower() for cell in header]
    mcid_column = None
    for candidate in ("participant_id", "mcid"):
        if candidate in lower_header:
            mcid_column = lower_header.index(candidate)
            break
    if mcid_column is None:
        summary.warnings.append(f"Skipped retention scoring file without participant_id/MCID column: {source}")
        return
    output_rows = [row for row in rows[1:] if mcid_column < len(row) and clean(row[mcid_column]) in included_mcids]
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow(header)
        writer.writerows(output_rows)
    summary.retention_rows_written = len(output_rows)
    if delimiter != "\t":
        summary.warnings.append(f"Converted retention scoring source {source.name} to TSV.")


def publish_data_for_included_mcids(
    included_mcids: list[str],
    *,
    raw_dir: Path,
    data_dir: Path,
    resources_dir: Path,
) -> PublishSummary:
    included_mcids = list(dict.fromkeys(included_mcids))
    included_set = set(included_mcids)
    summary = PublishSummary(included_count=len(included_mcids))
    require_path(raw_dir, "raw directory")
    data_dir.mkdir(parents=True, exist_ok=True)

    raw_log_dir = raw_dir / "logs"
    raw_survey_dir = raw_dir / "survey"
    raw_transcripts_dir = raw_dir / "transcripts"
    raw_config_dir = raw_dir / "config"
    data_log_dir = data_dir / "logs"
    data_transcripts_dir = data_dir / "transcripts"
    data_config_dir = data_dir / "config"
    data_survey_path = data_dir / "survey_export.tsv"

    raw_log_index = build_raw_log_index(raw_log_dir, included_mcids=included_mcids)
    condition_lookup = build_condition_lookup(included_mcids, raw_log_index, summary)
    publish_survey(raw_survey_dir, data_survey_path, included_set, condition_lookup, summary)
    publish_logs(included_mcids, raw_log_index, data_log_dir, summary)
    publish_transcripts(raw_transcripts_dir, data_transcripts_dir, included_set, summary)
    publish_config(raw_config_dir=raw_config_dir, data_config_dir=data_config_dir, resources_dir=resources_dir, summary=summary)
    publish_retention_scoring(raw_config_dir=raw_config_dir, raw_dir=raw_dir, data_dir=data_dir, included_mcids=included_set, summary=summary)
    return summary
