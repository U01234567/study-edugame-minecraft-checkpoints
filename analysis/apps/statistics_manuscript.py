from __future__ import annotations

# -----------------------------------------------------------------------------
# statistics_manuscript.py
# -----------------------------------------------------------------------------
# Table of contents for critical readers
#   1. Imports
#   2. Global paths and analysis constants
#   3. Small technical utilities: escaping, file reading, formatting, boxplots
#   4. RETENTION SCORING
#   5. DESCRIPTIVE STATISTICS
#      - Location
#      - Co-present participants
#      - Age
#      - Gender
#      - Creatures seen
#      - Perceived control
#      - Immediate retention
#      - Delayed retention
#      - Retention decay/change
#      - Intrinsic cognitive load
#      - Intrinsic cognitive load by chapter
#      - Extraneous cognitive load
#      - Extraneous cognitive load by chapter
#      - Germane cognitive load
#      - Engagement
#      - Engagement by chapter
#   6. ASSUMPTION CHECKS AND FINAL INFERENTIAL MODELS
#      - X: perceived control manipulation check
#      - H1 + Y: retention, delayed retention, and decay
#      - H2/H2a/H2b: cognitive-load mediation
#      - H3/H3a/H3b: engagement mediation
#      - H4: engagement and cognitive load
#      - Z: dropout analysis
#      - Interview-subsample representativeness
#   7. HTML report assembly
#
# This file intentionally repeats a lot of code. The goal is not elegance. The
# goal is that a non-programmer can open one output function and see which files
# are opened, which columns are used, how complete cases are formed, and how the
# displayed table/figure is produced.
#
# Current inferential-statistics policy: this stage-1 version keeps the
# preregistered confirmatory route visible first, then adds revised/sensitivity
# analyses underneath. This prevents improved implementation choices from being
# mistaken for preregistered choices.
# -----------------------------------------------------------------------------

import csv
import datetime as dt
import html
import json
import math
import re
import statistics
import sys
from itertools import combinations
from pathlib import Path
from typing import Any

import numpy as np

try:
    from scipy import stats as scipy_stats
except Exception:  # pragma: no cover - keeps the report usable without SciPy.
    scipy_stats = None

try:
    import pandas as pd
except Exception:  # pragma: no cover - only needed for the final mixed-effects model.
    pd = None

try:
    import statsmodels.formula.api as smf
except Exception:  # pragma: no cover - only needed for the final mixed-effects model.
    smf = None


# -----------------------------------------------------------------------------
# GLOBAL VARS WITH BASE FOLDERS AND SPECIFIC DATA FILES
# -----------------------------------------------------------------------------

THIS_FILE = Path(__file__).resolve()
if THIS_FILE.parent.name == "apps":
    ANALYSIS_DIR = THIS_FILE.parents[1]
else:
    ANALYSIS_DIR = THIS_FILE.parent

if str(ANALYSIS_DIR) not in sys.path:
    sys.path.insert(0, str(ANALYSIS_DIR))


try:
    from helpers._interviews_main import load_interview_overview
except Exception:  # pragma: no cover - keeps the report usable without openpyxl/interview helper imports.
    load_interview_overview = None

try:
    from helpers._logs_main import load_log_index
except Exception:  # pragma: no cover - interview representativeness can still use survey-only fallback rows.
    load_log_index = None

try:
    from helpers._main_overview import build_merged_dataset
except Exception:  # pragma: no cover - fallback rows are built inside this file.
    build_merged_dataset = None

try:
    from helpers._ret_main import attach_retention_scores, load_retention_scores
except Exception:  # pragma: no cover - fallback retention scores are read inside this file.
    attach_retention_scores = None
    load_retention_scores = None

try:
    from helpers._retention_coding import AMOUNT_GENAI, AMOUNT_HUMAN
except Exception:  # pragma: no cover - keeps the report usable when opened outside the repository.
    AMOUNT_GENAI = 2
    AMOUNT_HUMAN = 2

try:
    from helpers._shared import CREATURE_NAME_BY_ID, IGNORED_SEEN_EXTRAS, MAX_RETENTION_SLOTS
except Exception:  # pragma: no cover - keeps the report usable when opened outside the repository.
    CREATURE_NAME_BY_ID = {
        "abyss_deer": "Abyss deer",
        "amethyst_scarab": "Amethyst scarab",
        "axolotl_dragon": "Axolotl dragon",
        "cave_dweller": "Cave dweller",
        "ender_ape": "Ender ape",
        "flying_bunny": "Flying bunny",
        "glare": "Glare",
        "grand_grassling_father": "Grand grassling father",
        "ice_golem": "Ice golem",
        "killer_crab": "Killer crab",
        "lizard_knight": "Lizard knight",
        "mushroom_bup": "Mushroom bup",
        "orc": "Orc",
        "prototype_warden": "Prototype warden",
        "retro_tv_robot": "Retro TV robot",
        "scrambler_king": "Scrambler king",
        "walking_robot_guy": "Walking robot guy",
        "wardigo": "Wardigo",
    }
    IGNORED_SEEN_EXTRAS = {"cow", "chicken", "pig"}
    MAX_RETENTION_SLOTS = 18

DATA_DIR = ANALYSIS_DIR / "data"
OUTPUT_DIR = ANALYSIS_DIR / "output"
RESOURCES_DIR = ANALYSIS_DIR / "resources"
DATA_CONFIG_DIR = DATA_DIR / "config"
DATA_LOG_DIR = DATA_DIR / "logs"

SURVEY_EXPORT_PATH = DATA_DIR / "survey_export.tsv"
RETENTION_SCORES_MERGED_PATH = DATA_DIR / "retention_scores_merged.tsv"  # q_element-level audit/agreement file only
RETENTION_SCORES_FINAL_PATH = DATA_DIR / "retention_scores_final.tsv"    # participant-level score file used by stats_manu

RETENTION_FINAL_SCORE_VARIATION_PATHS = {
    "clean": RETENTION_SCORES_FINAL_PATH,
    "divide_by_18": DATA_DIR / "retention_scores_final-divide_by_18.tsv",
    "card_open_count_penalty": DATA_DIR / "retention_scores_final-card_open_count_penalty.tsv",
    "card_read_seconds_penalty": DATA_DIR / "retention_scores_final-card_read_seconds_penalty.tsv",
}

RETENTION_ANALYSIS_SCORES_PATH = RETENTION_SCORES_FINAL_PATH
COLLECTION_LOCATIONS_PATH = DATA_CONFIG_DIR / "collection_locations.json"
INTERVIEW_TRANSCRIPTS_DIR = DATA_DIR / "transcripts"
INTERVIEW_MANIFEST_PATH = DATA_CONFIG_DIR / "interview_manifest.json"
STATISTICS_MANUSCRIPT_OUTPUT_PATH = OUTPUT_DIR / "statistics_manuscript.html"

CONDITION_ORDER = ["Required continue", "Required pauses", "Optional pauses"]
DISPLAY_CONDITION = {
    "Required continue": "required continue",
    "Required pauses": "required pauses",
    "Optional pauses": "optional pauses",
    "Total": "total",
}
CONDITION_COLOURS = {
    "Required continue": "#2563eb",
    "Required pauses": "#f97316",
    "Optional pauses": "#16a34a",
    "Total": "#111827",
}
CONDITION_CODES = {
    "Required continue": {"required_pause_contrast": -1.0, "optional_pause_contrast": -0.5},
    "Required pauses": {"required_pause_contrast": 1.0, "optional_pause_contrast": -0.5},
    "Optional pauses": {"required_pause_contrast": 0.0, "optional_pause_contrast": 1.0},
}
CONTRAST_MULTIPLIERS = {
    "required_pause_contrast": 2.0,
    "optional_pause_contrast": 1.5,
}
CONTRAST_DISPLAY = {
    "required_pause_contrast": "Required pauses − required continue",
    "optional_pause_contrast": "Optional pauses − average system-controlled",
}
ALPHA = 0.05

RETENTION_COMPONENTS = [
    ("Q1_name", ["Q1_name"]),
    ("Q2_facts", ["Q2_fact1", "Q2_fact2", "Q2_fact3"]),
    ("Q3_looks", ["Q3_looks"]),
    ("Q4_location", ["Q4_chapter", "Q4_env"]),
]
VALID_RETENTION_ELEMENTS = {element for _component, elements in RETENTION_COMPONENTS for element in elements}


# -----------------------------------------------------------------------------
# SMALL TECHNICAL UTILITIES ONLY
# -----------------------------------------------------------------------------
# The calculations are deliberately repeated in the output functions. These
# utilities only keep the file readable: escaping HTML, reading delimited files,
# rendering repeated visual/table shells, and fitting a transparent OLS model.


def h(value: object) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def first_present(row: dict[str, Any] | None, names: list[str]) -> str:
    if row is None:
        return ""
    lower_lookup = {key.lower(): key for key in row.keys()}
    for name in names:
        if name in row and clean(row.get(name)):
            return clean(row.get(name))
        actual = lower_lookup.get(name.lower())
        if actual is not None and clean(row.get(actual)):
            return clean(row.get(actual))
    return ""


def parse_float(value: object) -> float | None:
    text = clean(value).replace(",", ".")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return number


def read_tsv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    for encoding in ("utf-8-sig", "utf-16", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                delimiter = "\t"
                if "\t" not in sample and "," in sample:
                    delimiter = ","
                return [dict(row) for row in csv.DictReader(handle, delimiter=delimiter)]
        except UnicodeError:
            continue
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle, delimiter="\t")]


def canonical_condition(value: object) -> str:
    text = clean(value)
    key = text.lower().replace("-", "_").replace(" ", "_")
    mapping = {
        "cond_continue": "Required continue",
        "continue": "Required continue",
        "required_continue": "Required continue",
        "required continue": "Required continue",
        "cond_pause": "Required pauses",
        "pause": "Required pauses",
        "required_pause": "Required pauses",
        "required_pauses": "Required pauses",
        "required pauses": "Required pauses",
        "cond_choice": "Optional pauses",
        "choice": "Optional pauses",
        "optional_pause": "Optional pauses",
        "optional_pauses": "Optional pauses",
        "optional pauses": "Optional pauses",
    }
    return mapping.get(key, mapping.get(text.lower(), text))


def mcid_from_row(row: dict[str, Any]) -> str:
    return first_present(row, ["MCID", "mcid", "participant_id", "Participant ID", "session_id"]).upper()


def delayed_flag(row: dict[str, Any]) -> bool:
    return first_present(row, ["DELAYED", "Delayed", "delayed"]).lower() in {"1", "true", "yes", "delayed"}


def delayed_included_flag(row: dict[str, Any]) -> bool:
    value = first_present(row, ["delayed_included", "DELAYED_INCLUDED", "DelayedIncluded"])
    if not value:
        return True
    return value.lower() in {"1", "true", "yes", "included"}


def is_truthy_seen_value(value: object) -> bool:
    text = clean(value).lower()
    return text in {"1", "true", "yes", "y", "x"}


def parse_seen_creature_ids(value: object) -> tuple[list[str], list[str]]:
    """Return unique valid creature IDs and invalid non-ignored IDs from SEEN."""
    valid: list[str] = []
    invalid: list[str] = []
    seen_valid: set[str] = set()
    raw_value = clean(value)
    if not raw_value or raw_value.upper() == "ERROR":
        return valid, invalid
    for part in raw_value.replace("|", ",").split(","):
        item = clean(part)
        if not item or item in IGNORED_SEEN_EXTRAS:
            continue
        if item in CREATURE_NAME_BY_ID:
            if item not in seen_valid:
                seen_valid.add(item)
                valid.append(item)
        else:
            invalid.append(item)
    return valid, invalid


def seen_creatures_from_row(row: dict[str, Any]) -> tuple[list[str], list[str], str]:
    """Read creature exposure from SEEN first, then fall back to seen_<creature_id> columns."""
    raw_seen = clean(row.get("SEEN"))
    if raw_seen and raw_seen.upper() != "ERROR":
        valid, invalid = parse_seen_creature_ids(raw_seen)
        return valid[:MAX_RETENTION_SLOTS], invalid, "SEEN"

    available_seen_columns = [creature_id for creature_id in CREATURE_NAME_BY_ID if f"seen_{creature_id}" in row]
    if available_seen_columns:
        valid = [creature_id for creature_id in available_seen_columns if is_truthy_seen_value(row.get(f"seen_{creature_id}"))]
        return valid[:MAX_RETENTION_SLOTS], [], "seen_* columns"

    return [], [], ""


RETENTION_FINAL_SCORE_COLUMN_ALIASES = {
    "Immediate": [
        "score_immediate",
        "ret_immediate_score",
        "immediate_retention",
        "retention_immediate",
    ],
    "Delayed": [
        "score_delayed",
        "ret_delayed_score",
        "delayed_retention",
        "retention_delayed",
    ],
}


def retention_score_file_label(path: Path | None = None) -> str:
    active_path = path or RETENTION_ANALYSIS_SCORES_PATH
    try:
        return "/" + str(active_path.resolve().relative_to(ANALYSIS_DIR.resolve())).replace("\\", "/")
    except ValueError:
        return str(active_path)


def retention_scores_final_path_for_mode(mode: str) -> Path:
    cleaned_mode = clean(mode)
    if cleaned_mode not in RETENTION_FINAL_SCORE_VARIATION_PATHS:
        available = ", ".join(RETENTION_FINAL_SCORE_VARIATION_PATHS)
        raise ValueError(f"Unknown retention final-score mode: {cleaned_mode}. Available modes: {available}")
    return RETENTION_FINAL_SCORE_VARIATION_PATHS[cleaned_mode]


def valid_participant_retention_score(value: object) -> float | None:
    """Parse participant-level retention from retention_scores_final*.tsv.

    These values are participant-level means, so unlike q_element-level final_score
    values they are not expected to be integers.
    """
    text = clean(value)
    if not text:
        return None
    number = parse_float(text)
    if number is None:
        return None
    if number < 0 or number > 2:
        return None
    return float(number)


def _has_any_column(header: set[str], candidates: list[str]) -> bool:
    lower_header = {column.lower() for column in header}
    return any(candidate.lower() in lower_header for candidate in candidates)


def participant_retention_scores_from_final(
    path: Path | None = None,
) -> tuple[dict[str, dict[str, float | None]], list[str], list[str]]:
    """Read participant-level immediate/delayed retention from retention_scores_final*.tsv.

    Default source is data/retention_scores_final.tsv. Mode variations can be used
    by passing one of the retention_scores_final-{mode}.tsv paths.
    """
    active_path = path or RETENTION_ANALYSIS_SCORES_PATH
    errors: list[str] = []
    warnings: list[str] = []

    final_rows = read_tsv(active_path)
    if not active_path.exists():
        return {}, [f"Missing participant-level retention score file: {active_path}"], []
    if not final_rows:
        return {}, [f"{retention_score_file_label(active_path)} is empty."], []

    header = set(final_rows[0].keys())
    if not _has_any_column(header, ["MCID"]):
        return {}, [f"{retention_score_file_label(active_path)} must contain an MCID column."], []

    missing_groups = [
        moment
        for moment, candidates in RETENTION_FINAL_SCORE_COLUMN_ALIASES.items()
        if not _has_any_column(header, candidates)
    ]
    if missing_groups:
        return {}, [
            f"{retention_score_file_label(active_path)} is missing participant-level score column(s) for: "
            + ", ".join(missing_groups)
            + ". Expected columns are usually score_immediate and score_delayed."
        ], []

    participant_scores: dict[str, dict[str, float | None]] = {}
    invalid_cells: list[str] = []
    duplicate_mcids: set[str] = set()

    for row_index, row in enumerate(final_rows, start=2):
        participant_id = clean(first_present(row, ["MCID"])).upper()
        if not participant_id:
            warnings.append(f"Skipped row {row_index} in {retention_score_file_label(active_path)} because MCID is empty.")
            continue

        if participant_id in participant_scores:
            duplicate_mcids.add(participant_id)
            continue

        immediate_raw = first_present(row, RETENTION_FINAL_SCORE_COLUMN_ALIASES["Immediate"])
        delayed_raw = first_present(row, RETENTION_FINAL_SCORE_COLUMN_ALIASES["Delayed"])

        immediate_score = valid_participant_retention_score(immediate_raw)
        delayed_score = valid_participant_retention_score(delayed_raw)

        if clean(immediate_raw) and immediate_score is None:
            invalid_cells.append(f"{participant_id}/Immediate row {row_index}")
        if clean(delayed_raw) and delayed_score is None:
            invalid_cells.append(f"{participant_id}/Delayed row {row_index}")

        participant_scores[participant_id] = {
            "Immediate": immediate_score,
            "Delayed": delayed_score,
        }

    if duplicate_mcids:
        warnings.append(
            f"{retention_score_file_label(active_path)} contains duplicate MCID row(s); the first row was kept for: "
            + ", ".join(sorted(duplicate_mcids)[:20])
        )

    if invalid_cells:
        preview = "; ".join(invalid_cells[:20])
        suffix = f"; plus {len(invalid_cells) - 20} more" if len(invalid_cells) > 20 else ""
        warnings.append(
            f"Found invalid participant-level retention score cell(s) in {retention_score_file_label(active_path)}; "
            f"these cells are treated as NA: {preview}{suffix}."
        )

    return participant_scores, errors, warnings


def parse_datetime(value: object) -> dt.datetime | None:
    text = clean(value).replace("T", " ").replace("Z", "")
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S:%f",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ):
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            pass
    try:
        return dt.datetime.fromisoformat(text)
    except ValueError:
        return None


SLOT_LABELS = [
    "09:40 - 10:40",
    "10:40 - 11:40",
    "11:40 - 12:40",
    "12:40 - 13:40",
    "13:40 - 14:40",
    "14:40 - 15:40",
    "15:40 - 16:40",
    "Outside 09:40 - 16:40",
]


def collection_slot_for_start(start: dt.datetime | None) -> tuple[str, int]:
    """Return the data-collection slot used by the merged-summary route.

    Sessions were scheduled on an :40 anchor.  Grouping by calendar hour would
    split one scheduled session across two bins whenever participants started on
    different sides of the clock hour.
    """
    if start is None:
        return "Unknown start time", 999

    time_of_day = start.time().replace(second=0, microsecond=0)
    anchor = dt.datetime.combine(dt.date(2000, 1, 1), dt.time(9, 40))
    for index in range(7):
        begin = (anchor + dt.timedelta(hours=index)).time()
        end = (anchor + dt.timedelta(hours=index + 1)).time()
        if begin <= time_of_day < end:
            return SLOT_LABELS[index], index

    return SLOT_LABELS[-1], 7


def canonical_collection_location(value: object) -> str:
    """Return the manuscript's three-level location code."""
    text = clean(value)
    if text in {"CreaSp", "Creative Space"}:
        return "CreaSp"
    if text in {"LivingR", "Living Room"}:
        return "LivingR"
    if text in {"Remote", "At home"}:
        return "Remote"
    return ""


def load_collection_location_map() -> tuple[dict[str, str], list[str]]:
    """Read data/config/collection_locations.json without depending on helpers."""
    try:
        with open(COLLECTION_LOCATIONS_PATH, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except Exception as exc:
        return {}, [f"Could not read {COLLECTION_LOCATIONS_PATH}: {exc}"]

    raw_locations = payload.get("locations_by_date", payload) if isinstance(payload, dict) else {}
    if not isinstance(raw_locations, dict):
        return {}, [f"{COLLECTION_LOCATIONS_PATH} must be a JSON object or contain a locations_by_date object."]
    return {clean(date_key): clean(value) for date_key, value in raw_locations.items()}, []


def event_name_from_log_activity(value: object) -> str:
    return clean(value).split("|", 1)[0].strip()


def log_start_from_csv(path: Path) -> dt.datetime | None:
    """Return study_session_started time from a publishable log CSV, falling back to the first timestamp."""
    first_timestamp: dt.datetime | None = None
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                parsed = parse_datetime(f"{clean(row.get('Date'))} {clean(row.get('Time'))}")
                if parsed is None:
                    continue
                if first_timestamp is None:
                    first_timestamp = parsed
                if event_name_from_log_activity(row.get("Activity")) == "study_session_started":
                    return parsed
    except OSError:
        return None
    return first_timestamp


def log_starts_by_mcid() -> tuple[dict[str, dt.datetime], list[str]]:
    """Return the log-start map used for collection context, without importing sum_merged helpers."""
    if not DATA_LOG_DIR.exists():
        return {}, [f"Log folder not found: {DATA_LOG_DIR}; collection context uses survey start date only."]

    starts: dict[str, dt.datetime] = {}
    for path in sorted(DATA_LOG_DIR.glob("*.csv")):
        parsed = log_start_from_csv(path)
        if parsed is not None:
            starts[path.stem.upper()] = parsed
    return starts, []


def build_collection_context_rows(
    immediate_by_mcid: dict[str, dict[str, str]],
    raw_locations: dict[str, str],
    log_start_by_mcid: dict[str, dt.datetime],
    *,
    require_valid_condition: bool,
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    """Build manuscript-coded location and co-presence rows.

    The calculation mirrors the merged-summary timing source and slot boundaries,
    but keeps the manuscript's labels: CreaSp, LivingR, and Remote. Remote rows
    are coded as 0 co-present participants because that is the value entered in
    the covariate-adjusted models; the location factor keeps remote participation
    distinguishable from lab participants who happened to be alone.
    """
    provisional_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []

    for participant_id, row in sorted(immediate_by_mcid.items()):
        raw_condition = first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"])
        condition = canonical_condition(raw_condition)
        if require_valid_condition and condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": raw_condition, "reason": "Missing or invalid condition"})
            continue

        is_remote = first_present(row, ["REMOTE", "remote", "Remote"]).lower() in {"1", "true", "yes"}
        survey_start = parse_datetime(first_present(row, ["startDate", "StartDate", "Start Date", "Start date"]))
        start = log_start_by_mcid.get(participant_id) or survey_start

        if is_remote:
            source_date = start.date().isoformat() if start else ""
            provisional_rows.append({
                "MCID": participant_id,
                "condition": condition,
                "location": "Remote",
                "date": source_date,
                "slot": "Remote",
                "slot_order": None,
                "slot_key": None,
            })
            continue

        if start is None:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "No log or survey start date available for collection-context lookup"})
            continue

        source_date = start.date().isoformat()
        location = canonical_collection_location(raw_locations.get(source_date))
        if location not in {"CreaSp", "LivingR"}:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": f"No valid collection_locations.json entry for {source_date}"})
            continue

        slot_label, slot_order = collection_slot_for_start(start)
        slot_key = (source_date, slot_label, location)
        provisional_rows.append({
            "MCID": participant_id,
            "condition": condition,
            "location": location,
            "date": source_date,
            "slot": slot_label,
            "slot_order": slot_order,
            "slot_key": slot_key,
        })

    slot_counts: dict[tuple[str, str, str], int] = {}
    for row in provisional_rows:
        slot_key = row.get("slot_key")
        if slot_key is not None:
            slot_counts[slot_key] = slot_counts.get(slot_key, 0) + 1

    complete_rows: list[dict[str, Any]] = []
    for row in provisional_rows:
        value = 0.0 if row.get("location") == "Remote" else float(max(0, slot_counts.get(row.get("slot_key"), 1) - 1))
        out = dict(row)
        out["value"] = value
        out["co_present_participants"] = f"{value:.0f}"
        complete_rows.append(out)

    return complete_rows, excluded


def normalise_gender(value: object) -> str:
    text = clean(value).lower()
    if not text:
        return ""
    if text in {"1", "male", "man", "m"}:
        return "Male"
    if text in {"2", "female", "woman", "f"}:
        return "Female"
    if text in {"3", "other", "non-binary", "nonbinary", "prefer to self-describe"}:
        return "Other"
    return "Other"


def mean_sd_min_max_html(values: list[float]) -> str:
    if not values:
        return "—"
    mean = statistics.fmean(values)
    sd = statistics.stdev(values) if len(values) > 1 else 0.0
    return (
        '<table class="micro-stat-table"><tbody>'
        f'<tr><th>Mean</th><td>{mean:.2f}</td></tr>'
        f'<tr><th><em>SD</em></th><td>{sd:.2f}</td></tr>'
        f'<tr><th>Min</th><td>{min(values):.2f}</td></tr>'
        f'<tr><th>Max</th><td>{max(values):.2f}</td></tr>'
        '</tbody></table>'
    )


def n_mean_sd_min_max_html(values: list[float]) -> str:
    clean_values = [float(value) for value in values if value is not None and math.isfinite(float(value))]
    if not clean_values:
        return (
            '<table class="micro-stat-table"><tbody>'
            '<tr><th><em>n</em></th><td>0</td></tr>'
            '<tr><th>Mean</th><td>—</td></tr>'
            '<tr><th><em>SD</em></th><td>—</td></tr>'
            '<tr><th>Min</th><td>—</td></tr>'
            '<tr><th>Max</th><td>—</td></tr>'
            '</tbody></table>'
        )
    mean = statistics.fmean(clean_values)
    sd = statistics.stdev(clean_values) if len(clean_values) > 1 else 0.0
    return (
        '<table class="micro-stat-table"><tbody>'
        f'<tr><th><em>n</em></th><td>{len(clean_values)}</td></tr>'
        f'<tr><th>Mean</th><td>{mean:.2f}</td></tr>'
        f'<tr><th><em>SD</em></th><td>{sd:.2f}</td></tr>'
        f'<tr><th>Min</th><td>{min(clean_values):.2f}</td></tr>'
        f'<tr><th>Max</th><td>{max(clean_values):.2f}</td></tr>'
        '</tbody></table>'
    )


def quantile(sorted_values: list[float], probability: float) -> float | None:
    if not sorted_values:
        return None
    index = (len(sorted_values) - 1) * probability
    lower = math.floor(index)
    upper = math.ceil(index)
    if lower == upper:
        return sorted_values[lower]
    return sorted_values[lower] + (sorted_values[upper] - sorted_values[lower]) * (index - lower)


def boxplot_svg(figure_id: str, rows: list[dict[str, Any]], value_key: str, title: str, *, min_value: float | None = None, max_value: float | None = None) -> str:
    numeric_values = [float(row[value_key]) for row in rows if parse_float(row.get(value_key)) is not None]
    if not numeric_values:
        return f'<div class="chart-box"><h3>{h(title)}</h3><p class="small">No numeric values available for a boxplot.</p></div>'
    plot_min = min_value if min_value is not None else min(numeric_values)
    plot_max = max_value if max_value is not None else max(numeric_values)
    if plot_min == plot_max:
        plot_min -= 0.5
        plot_max += 0.5
    width = 880
    height = 390
    margin_left = 70
    margin_top = 34
    inner_width = 760
    inner_height = 250
    slot_width = inner_width / max(1, len(CONDITION_ORDER))

    def scale_y(value: float) -> float:
        return margin_top + inner_height - ((value - plot_min) / max(0.001, plot_max - plot_min)) * inner_height

    y_ticks = []
    tick_count = 5
    for tick_index in range(tick_count):
        value = plot_min + (plot_max - plot_min) * tick_index / (tick_count - 1)
        y = scale_y(value)
        y_ticks.append(f'<g><line x1="{margin_left - 4}" x2="{margin_left + inner_width}" y1="{y:.2f}" y2="{y:.2f}" stroke="#d9e0e4"></line><text x="{margin_left - 8}" y="{y + 4:.2f}" text-anchor="end" font-size="11">{value:.2f}</text></g>')

    boxes = []
    for condition_index, condition in enumerate(CONDITION_ORDER):
        values = sorted(float(row[value_key]) for row in rows if row.get("condition") == condition and parse_float(row.get(value_key)) is not None)
        if not values:
            continue
        q1 = quantile(values, 0.25)
        median = quantile(values, 0.50)
        q3 = quantile(values, 0.75)
        assert q1 is not None and median is not None and q3 is not None
        value_min = min(values)
        value_max = max(values)
        x = margin_left + condition_index * slot_width + slot_width / 2
        colour = CONDITION_COLOURS.get(condition, "#111827")
        y_min = scale_y(value_min)
        y_q1 = scale_y(q1)
        y_med = scale_y(median)
        y_q3 = scale_y(q3)
        y_max = scale_y(value_max)
        top = min(y_q1, y_q3)
        box_height = max(1, abs(y_q3 - y_q1))
        details = f"{condition}: n={len(values)}, min={value_min:.2f}, Q1={q1:.2f}, median={median:.2f}, Q3={q3:.2f}, max={value_max:.2f}"
        boxes.append(
            f'<g class="clickable-box" onclick="showBoxplotDetails(\'{h(figure_id)}\', \'{h(details)}\')">'
            f'<line x1="{x:.2f}" x2="{x:.2f}" y1="{y_min:.2f}" y2="{y_max:.2f}" stroke="{colour}" stroke-width="2"></line>'
            f'<line x1="{x - 12:.2f}" x2="{x + 12:.2f}" y1="{y_min:.2f}" y2="{y_min:.2f}" stroke="{colour}" stroke-width="2"></line>'
            f'<line x1="{x - 12:.2f}" x2="{x + 12:.2f}" y1="{y_max:.2f}" y2="{y_max:.2f}" stroke="{colour}" stroke-width="2"></line>'
            f'<rect x="{x - 18:.2f}" y="{top:.2f}" width="36" height="{box_height:.2f}" fill="white" stroke="{colour}" stroke-width="2"><title>{h(details)}</title></rect>'
            f'<line x1="{x - 18:.2f}" x2="{x + 18:.2f}" y1="{y_med:.2f}" y2="{y_med:.2f}" stroke="{colour}" stroke-width="3"></line>'
            f'<text x="{x:.2f}" y="{max(12, y_max - 6):.2f}" text-anchor="middle" font-size="10">n={len(values)}</text>'
            f'<text x="{x:.2f}" y="{min(height - 75, y_med - 6):.2f}" text-anchor="middle" font-size="10">{median:.2f}</text>'
            f'<text x="{x:.2f}" y="{height - 42}" text-anchor="middle" font-size="12">{h(DISPLAY_CONDITION[condition])}</text>'
            '</g>'
        )
    return (
        f'<div class="chart-box"><h3>{h(title)}</h3>'
        '<p class="small">Clickable standalone SVG: centre line = median; boxes = Q1–Q3; whiskers = min–max; labels show <em>n</em> and median.</p>'
        f'<svg class="standalone-figure" viewBox="0 0 {width} {height}" role="img">'
        f'<rect x="0" y="0" width="{width}" height="{height}" fill="white"></rect>'
        + "".join(y_ticks)
        + f'<line x1="{margin_left}" y1="{margin_top}" x2="{margin_left}" y2="{margin_top + inner_height}" stroke="#5f6c73"></line>'
        + f'<line x1="{margin_left}" y1="{margin_top + inner_height}" x2="{margin_left + inner_width}" y2="{margin_top + inner_height}" stroke="#5f6c73"></line>'
        + "".join(boxes)
        + '</svg>'
        f'<p class="small boxplot-detail" id="{h(figure_id)}-detail">Click a box to see its quartiles.</p>'
        '</div>'
    )


def categorical_bars(figure_id: str, rows: list[dict[str, Any]], value_key: str, title: str, categories: list[str]) -> str:
    if not rows:
        return f'<div class="chart-box"><h3>{h(title)}</h3><p class="small">No categorical rows available.</p></div>'
    width = 880
    height = 340
    left = 60
    top = 45
    bar_height = 20
    gap = 11
    max_count = 1
    counts: dict[tuple[str, str], int] = {}
    for condition in CONDITION_ORDER:
        for category in categories:
            count = sum(1 for row in rows if row.get("condition") == condition and row.get(value_key) == category)
            counts[(condition, category)] = count
            max_count = max(max_count, count)
    bars = []
    y = top
    for condition in CONDITION_ORDER:
        bars.append(f'<text x="{left}" y="{y + 14}" font-size="12" font-weight="700">{h(DISPLAY_CONDITION[condition])}</text>')
        y += 24
        for category in categories:
            count = counts[(condition, category)]
            bar_width = 680 * count / max_count if max_count else 0
            colour = CONDITION_COLOURS.get(condition, "#111827")
            detail = f"{condition} | {category}: n={count}"
            bars.append(
                f'<g class="clickable-box" onclick="showBoxplotDetails(\'{h(figure_id)}\', \'{h(detail)}\')">'
                f'<text x="{left + 20}" y="{y + 14}" font-size="11">{h(category)}</text>'
                f'<rect x="{left + 150}" y="{y}" width="{bar_width:.2f}" height="{bar_height}" fill="{colour}" opacity="0.82"><title>{h(detail)}</title></rect>'
                f'<text x="{left + 160 + bar_width:.2f}" y="{y + 14}" font-size="11">{count}</text>'
                '</g>'
            )
            y += bar_height + gap
        y += 8
    return (
        f'<div class="chart-box"><h3>{h(title)}</h3>'
        '<p class="small">Categorical values do not have a meaningful boxplot, so this table uses a clickable count figure instead.</p>'
        f'<svg class="standalone-figure" viewBox="0 0 {width} {height}" role="img">'
        f'<rect x="0" y="0" width="{width}" height="{height}" fill="white"></rect>'
        + "".join(bars) + '</svg>'
        f'<p class="small boxplot-detail" id="{h(figure_id)}-detail">Click a bar to see its count.</p>'
        '</div>'
    )


def status_messages(errors: list[str], warnings: list[str], ok_text: str) -> str:
    parts: list[str] = []
    if errors:
        parts.append('<div class="status status-red"><strong>ERROR.</strong> ' + h(" ".join(errors)) + '</div>')
    if warnings:
        parts.append('<div class="status status-orange"><strong>WARNING.</strong> ' + h(" ".join(warnings)) + '</div>')
    if not errors and not warnings:
        parts.append('<div class="status status-green"><strong>OK.</strong> ' + h(ok_text) + '</div>')
    return "".join(parts)


def excluded_details(excluded: list[dict[str, str]]) -> str:
    rows = "".join(
        f'<tr><td>{h(item.get("MCID"))}</td><td>{h(item.get("condition"))}</td><td>{h(item.get("reason"))}</td></tr>'
        for item in excluded
    )
    if not rows:
        rows = '<tr><td colspan="3">No complete-case exclusions for this output.</td></tr>'
    return (
        '<details class="compact-details"><summary>Show excluded complete-case MCIDs</summary>'
        '<div class="table-wrap"><table><thead><tr><th>MCID</th><th>Condition</th><th>Reason</th></tr></thead><tbody>'
        + rows + '</tbody></table></div></details>'
    )


def first_three_details(rows: list[dict[str, Any]], value_columns: list[str]) -> str:
    all_rows = []
    for condition in CONDITION_ORDER:
        scoped = [row for row in rows if row.get("condition") == condition]
        for row in scoped[:3]:
            cells = [f'<td>{h(row.get("MCID"))}</td>', f'<td>{h(DISPLAY_CONDITION[condition])}</td>']
            for column in value_columns:
                cells.append(f'<td>{h(row.get(column))}</td>')
            all_rows.append('<tr>' + "".join(cells) + '</tr>')
    if not all_rows:
        all_rows.append(f'<tr><td colspan="{2 + len(value_columns)}">No complete rows to preview.</td></tr>')
    header = '<th>MCID</th><th>Condition</th>' + "".join(f'<th>{h(column)}</th>' for column in value_columns)
    return (
        '<details class="compact-details"><summary>Show first three rows per condition used for this table</summary>'
        '<div class="table-wrap"><table><thead><tr>' + header + '</tr></thead><tbody>'
        + "".join(all_rows) + '</tbody></table></div></details>'
    )


def table_shell(title: str, in_body: bool, description: str, table_html: str, figure_html: str, status_html: str, details_html: str) -> str:
    section_class = "body-output" if in_body else "appendix-output"
    where = "Manuscript body" if in_body else "Appendix"
    return (
        f'<section id="{h(slugify(title))}" class="card {section_class}">'
        f'<h2>{h(title)} <span>{h(where)}</span></h2>'
        f'<p>{h(description)}</p>'
        f'<div class="table-wrap">{table_html}</div>'
        f'{figure_html}'
        f'{status_html}'
        f'{details_html}'
        '</section>'
    )


# =============================================================================
# RETENTION SCORING AGREEMENT
# =============================================================================


def _natural_source_key(label: str) -> list[Any]:
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", clean(label))]


def _configured_source_labels(kind: str, amount: int) -> list[str]:
    return [f"{kind}{index}" for index in range(1, max(0, int(amount)) + 1)]


def _display_source_label(label: str) -> str:
    text = clean(label)
    match = re.match(r"^(genai|grader)(\d+)(?:_(\d+))?$", text, flags=re.IGNORECASE)
    if not match:
        if text.lower() == "final_score":
            return "Final score"
        return text or "Unknown source"
    source_type, number, duplicate = match.groups()
    base = "GenAI" if source_type.lower() == "genai" else "Human grader"
    suffix = f" / {duplicate}" if duplicate else ""
    return f"{base} #{number}{suffix}"


def _source_labels_from_rows(rows: list[dict[str, str]], kind: str, amount: int) -> list[str]:
    header = set(rows[0].keys()) if rows else set()
    configured = _configured_source_labels(kind, amount)
    labels = [label for label in configured if f"{label}_score" in header]
    detected: set[str] = set(labels)
    for column in header:
        match = re.match(rf"^({kind}\d+(?:_\d+)?)_score$", column)
        if match:
            detected.add(match.group(1))
    if kind == "genai" and not detected and "genai_score" in header:
        detected.add("genai")
    return sorted(detected, key=_natural_source_key)


def _valid_int_score(value: object) -> int | None:
    number = parse_float(value)
    if number is None or not float(number).is_integer():
        return None
    score = int(number)
    return score if score in (0, 1, 2) else None


def _source_score(row: dict[str, str], label: str) -> int | None:
    if label == "final_score":
        return _valid_int_score(row.get("final_score"))
    if label.startswith("grader") and clean(row.get(f"{label}_status")) not in {"", "graded"}:
        return None
    if label == "genai":
        return _valid_int_score(row.get("genai_score"))
    return _valid_int_score(row.get(f"{label}_score"))


def quadratic_weighted_kappa(pairs: list[tuple[int, int]], categories: list[int] | None = None) -> float | None:
    if not pairs:
        return None
    cats = categories or [0, 1, 2]
    index = {cat: i for i, cat in enumerate(cats)}
    k = len(cats)
    matrix = [[0.0 for _ in cats] for _ in cats]
    for left, right in pairs:
        if left in index and right in index:
            matrix[index[left]][index[right]] += 1.0
    n = sum(sum(row) for row in matrix)
    if n == 0:
        return None
    row_totals = [sum(row) for row in matrix]
    col_totals = [sum(matrix[i][j] for i in range(k)) for j in range(k)]
    observed = 0.0
    expected = 0.0
    for i in range(k):
        for j in range(k):
            weight = ((i - j) ** 2) / ((k - 1) ** 2)
            observed += weight * matrix[i][j]
            expected += weight * (row_totals[i] * col_totals[j] / n)
    if expected == 0:
        return 1.0 if observed == 0 else None
    return 1.0 - observed / expected


def ordinal_krippendorff_alpha(ratings_by_unit: list[list[int]], categories: list[int] | None = None) -> float | None:
    cats = categories or [0, 1, 2]
    index = {cat: i for i, cat in enumerate(cats)}
    k = len(cats)
    coincidence = [[0.0 for _ in cats] for _ in cats]
    valid_units = 0
    for raw_unit in ratings_by_unit:
        values = [score for score in raw_unit if score in index]
        m = len(values)
        if m < 2:
            continue
        valid_units += 1
        counts = [0 for _ in cats]
        for score in values:
            counts[index[score]] += 1
        for i in range(k):
            for j in range(k):
                if i == j:
                    coincidence[i][j] += counts[i] * (counts[j] - 1) / (m - 1)
                else:
                    coincidence[i][j] += counts[i] * counts[j] / (m - 1)
    n = sum(sum(row) for row in coincidence)
    if valid_units == 0 or n <= 1:
        return None
    marginals = [sum(row) for row in coincidence]
    observed = 0.0
    expected = 0.0
    for i in range(k):
        for j in range(k):
            delta = ((i - j) ** 2) / ((k - 1) ** 2)
            observed += coincidence[i][j] * delta
            expected += marginals[i] * marginals[j] * delta / (n - 1)
    if expected == 0:
        return 1.0 if observed == 0 else None
    return 1.0 - observed / expected


def _fmt_stat(value: float | None, digits: int = 3) -> str:
    if value is None:
        return "—"
    return f"{value:.{digits}f}"


def _fmt_percent(numerator: int, denominator: int) -> str:
    if denominator <= 0:
        return "—"
    return f"{100 * numerator / denominator:.1f}%"


def _agreement_summary_row(label: str, rows: list[dict[str, str]], source_labels: list[str]) -> dict[str, Any]:
    ratings_by_unit: list[list[int]] = []
    complete_units: list[list[int]] = []
    for row in rows:
        ratings = [_source_score(row, source_label) for source_label in source_labels]
        valid_ratings = [score for score in ratings if score is not None]
        if len(valid_ratings) >= 2:
            ratings_by_unit.append(valid_ratings)
        if len(ratings) >= 2 and all(score is not None for score in ratings):
            complete_units.append([int(score) for score in ratings if score is not None])

    pair_details: list[str] = []
    kappas: list[float] = []
    pair_count = 0
    for left, right in combinations(source_labels, 2):
        pairs = []
        for row in rows:
            left_score = _source_score(row, left)
            right_score = _source_score(row, right)
            if left_score is not None and right_score is not None:
                pairs.append((left_score, right_score))
        pair_count += len(pairs)
        kappa = quadratic_weighted_kappa(pairs)
        if kappa is not None:
            kappas.append(kappa)
        pair_details.append(f"{_display_source_label(left)} × {_display_source_label(right)}: {_fmt_stat(kappa)} (n={len(pairs)})")

    exact = sum(1 for ratings in complete_units if len(set(ratings)) == 1)
    alpha = ordinal_krippendorff_alpha(ratings_by_unit)
    mean_kappa = statistics.fmean(kappas) if kappas else None
    return {
        "group": label,
        "ordinal_alpha": _fmt_stat(alpha),
        "quadratic_weighted_kappa": _fmt_stat(mean_kappa),
        "percent_exact_agreement": _fmt_percent(exact, len(complete_units)),
        "pairwise_quadratic_weighted_kappa": "; ".join(pair_details) if pair_details else "—",
        "n_units": str(len(ratings_by_unit)),
        "n_pairwise": str(pair_count),
        "sources": ", ".join(_display_source_label(source_label) for source_label in source_labels) or "—",
    }


def _agreement_with_final_rows(rows: list[dict[str, str]], source_labels: list[str]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for source_label in source_labels:
        pairs: list[tuple[int, int]] = []
        ratings_by_unit: list[list[int]] = []
        for row in rows:
            source_score = _source_score(row, source_label)
            final_score = _source_score(row, "final_score")
            if source_score is None or final_score is None:
                continue
            pairs.append((source_score, final_score))
            ratings_by_unit.append([source_score, final_score])
        exact = sum(1 for left, right in pairs if left == right)
        kappa = quadratic_weighted_kappa(pairs)
        alpha = ordinal_krippendorff_alpha(ratings_by_unit)
        display = _display_source_label(source_label)
        output.append({
            "group": f"Agreement each with final score · {display}",
            "ordinal_alpha": _fmt_stat(alpha),
            "quadratic_weighted_kappa": _fmt_stat(kappa),
            "percent_exact_agreement": _fmt_percent(exact, len(pairs)),
            "pairwise_quadratic_weighted_kappa": f"{display} × Final score: {_fmt_stat(kappa)} (n={len(pairs)})",
            "n_units": str(len(ratings_by_unit)),
            "n_pairwise": str(len(pairs)),
            "sources": f"{display}, Final score",
        })
    return output


def retention_scoring() -> str:
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    rows = read_tsv(RETENTION_SCORES_MERGED_PATH)
    if not RETENTION_SCORES_MERGED_PATH.exists():
        errors.append(f"Missing merged retention scoring file: {RETENTION_SCORES_MERGED_PATH}")
    if rows and "final_score" not in rows[0]:
        errors.append("retention_scores_merged.tsv is missing the final_score column needed for final-score agreement checks.")

    genai_labels = _source_labels_from_rows(rows, "genai", AMOUNT_GENAI)
    human_labels = _source_labels_from_rows(rows, "grader", AMOUNT_HUMAN)
    all_labels = genai_labels + human_labels
    if rows and not genai_labels:
        warnings.append("No GenAI score columns were detected in retention_scores_merged.tsv.")
    if rows and not human_labels:
        warnings.append("No human-grader score columns were detected in retention_scores_merged.tsv.")

    summary_rows: list[dict[str, Any]] = []
    if all_labels:
        summary_rows.append(_agreement_summary_row("Agreement all graders", rows, all_labels))
    if genai_labels:
        summary_rows.append(_agreement_summary_row("Agreement all GenAI graders", rows, genai_labels))
    if human_labels:
        summary_rows.append(_agreement_summary_row("Agreement all human graders", rows, human_labels))
    summary_rows.extend(_agreement_with_final_rows(rows, all_labels))

    body_rows = []
    for row in summary_rows:
        body_rows.append(
            "<tr>"
            f"<th>{h(row.get('group'))}</th>"
            f"<td>{h(row.get('ordinal_alpha'))}</td>"
            f"<td>{h(row.get('quadratic_weighted_kappa'))}</td>"
            f"<td>{h(row.get('percent_exact_agreement'))}</td>"
            f"<td>{h(row.get('pairwise_quadratic_weighted_kappa'))}</td>"
            f"<td>{h(row.get('n_units'))}</td>"
            f"<td>{h(row.get('n_pairwise'))}</td>"
            "</tr>"
        )
    if not body_rows:
        body_rows.append('<tr><td colspan="7">No retention-scoring agreement rows could be calculated.</td></tr>')

    table_html = (
        '<table><thead><tr>'
        '<th>Agreement set</th><th>Ordinal Krippendorff α</th><th>Quadratic weighted Cohen κ</th>'
        '<th>Percent exact agreement</th><th>Pairwise quadratic weighted Cohen κ</th><th>n units</th><th>n pairwise</th>'
        '</tr></thead><tbody>' + "".join(body_rows) + '</tbody></table>'
    )
    detail_rows = "".join(
        f"<tr><td>{h(row.get('group'))}</td><td>{h(row.get('sources'))}</td></tr>"
        for row in summary_rows
    ) or '<tr><td colspan="2">No sources detected.</td></tr>'
    details_html = (
        '<details class="compact-details"><summary>Show scorer/source labels used for agreement rows</summary>'
        '<div class="table-wrap"><table><thead><tr><th>Agreement set</th><th>Sources</th></tr></thead><tbody>'
        + detail_rows + '</tbody></table></div></details>'
    )
    return table_shell(
        "RETENTION SCORING",
        IN_BODY,
        "Scoring-agreement diagnostics are calculated from retention_scores_merged.tsv. The retention descriptives below use retention_scores_final.tsv instead.",
        table_html,
        "",
        status_messages(errors, warnings, "Retention scoring agreement diagnostics were calculated."),
        details_html,
    )


# =============================================================================
# DESCRIPTIVE STATISTICS
# =============================================================================


def descriptive_location() -> str:
    """Open survey_export.tsv, data/logs/*.csv, and collection_locations.json; classify each complete participant as CreaSp, LivingR, or Remote; then display counts by condition."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not survey_rows and not errors:
        errors.append("survey_export.tsv was empty or could not be read.")

    raw_locations, location_errors = load_collection_location_map()
    errors.extend(location_errors)
    log_start_by_mcid, log_warnings = log_starts_by_mcid()
    warnings.extend(log_warnings)

    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if not participant_id or delayed_flag(row):
            continue
        if participant_id not in immediate_by_mcid:
            immediate_by_mcid[participant_id] = row

    complete_rows, excluded = build_collection_context_rows(
        immediate_by_mcid,
        raw_locations,
        log_start_by_mcid,
        require_valid_condition=True,
    )

    categories = ["CreaSp", "LivingR", "Remote"]
    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        counts = {category: sum(1 for row in scoped if row["location"] == category) for category in categories}
        total = len(scoped)
        count_cell = "<br>".join(f"{category}: {counts[category]} ({(100 * counts[category] / total):.1f}%)" if total else f"{category}: 0" for category in categories)
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{total}</td><td>{count_cell}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Location counts</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    if excluded:
        warnings.append(f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays {len(complete_rows)} after complete-case exclusion.")
    return table_shell(
        "Descriptives: location",
        IN_BODY,
        "Location is operationalised as CreaSp, LivingR, or Remote. Remote comes from REMOTE=1; lab rows come from data/config/collection_locations.json by collection date. When logs are available, the collection date is taken from the study_session_started event; otherwise the survey start date is used.",
        table_html,
        categorical_bars("location-bars", complete_rows, "location", "Location distribution", categories),
        status_messages(errors, warnings, "Location values were available and valid for all displayed participants."),
        excluded_details(excluded) + first_three_details(complete_rows, ["location", "date", "slot"]),
    )


def descriptive_co_present() -> str:
    """Open survey_export.tsv, data/logs/*.csv, and collection_locations.json; compute how many other included participants were in the same date/session-slot/location slot; then display numeric descriptives by condition."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")
    if not survey_rows and not errors:
        errors.append("survey_export.tsv was empty or could not be read.")
    raw_locations, location_errors = load_collection_location_map()
    errors.extend(location_errors)
    log_start_by_mcid, log_warnings = log_starts_by_mcid()
    warnings.extend(log_warnings)

    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if not participant_id or delayed_flag(row):
            continue
        immediate_by_mcid.setdefault(participant_id, row)

    complete_rows, excluded = build_collection_context_rows(
        immediate_by_mcid,
        raw_locations,
        log_start_by_mcid,
        require_valid_condition=True,
    )

    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        values = [float(row["value"]) for row in scoped]
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{len(values)}</td><td>{mean_sd_min_max_html(values)}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Co-present participants</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    if excluded:
        warnings.append(f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays {len(complete_rows)} after complete-case exclusion.")
    return table_shell(
        "Descriptives: co-present participants",
        IN_BODY,
        "Co-present participants is operationalised as the number of other included participants in the same lab date/session-slot/location slot; Remote is coded 0. Session slots use the same 09:40-anchored schedule as sum_merged.",
        table_html,
        boxplot_svg("copresent-boxplot", complete_rows, "value", "Co-present participant boxplot", min_value=0),
        status_messages(errors, warnings, "Co-present participant values were calculated for all displayed participants."),
        excluded_details(excluded) + first_three_details(complete_rows, ["co_present_participants", "location", "date", "slot"]),
    )


def descriptive_age() -> str:
    """Open survey_export.tsv; take MCID, condition, and age only; require valid age for complete cases; then display numeric descriptives by condition."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")
    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)
    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    for participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        age = parse_float(first_present(row, ["age", "Age"]))
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition"})
            continue
        if age is None:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing age"})
            continue
        if age < 0 or age > 120:
            errors.append(f"Out-of-range age for {participant_id}: {age}.")
            excluded.append({"MCID": participant_id, "condition": condition, "reason": f"Out-of-range age: {age}"})
            continue
        complete_rows.append({"MCID": participant_id, "condition": condition, "value": age, "age": f"{age:.0f}" if float(age).is_integer() else f"{age:.2f}"})
    if excluded:
        warnings.append(f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays {len(complete_rows)} after complete-case exclusion.")
    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        values = [float(row["value"]) for row in scoped]
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{len(values)}</td><td>{mean_sd_min_max_html(values)}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Age</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    return table_shell("Descriptives: age", IN_BODY, "Age is read from the immediate survey row and must be present and between 0 and 120.", table_html, boxplot_svg("age-boxplot", complete_rows, "value", "Age boxplot"), status_messages(errors, warnings, "Age values were valid for all displayed participants."), excluded_details(excluded) + first_three_details(complete_rows, ["age"]))


def descriptive_gender() -> str:
    """Open survey_export.tsv; take MCID, condition, and gender only; require a non-empty gender value; then display counts by condition."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")
    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)
    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    raw_other_values: set[str] = set()
    for participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        raw_gender = first_present(row, ["gender", "Gender"])
        gender = normalise_gender(raw_gender)
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition"})
            continue
        if not gender:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing gender"})
            continue
        if gender == "Other" and clean(raw_gender).lower() not in {"3", "other", "non-binary", "nonbinary", "prefer to self-describe"}:
            raw_other_values.add(raw_gender)
        complete_rows.append({"MCID": participant_id, "condition": condition, "gender": gender, "raw_gender": raw_gender})
    if raw_other_values:
        warnings.append("Some non-empty gender values were grouped as Other: " + ", ".join(sorted(raw_other_values)) + ".")
    if excluded:
        warnings.append(f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays {len(complete_rows)} after complete-case exclusion.")
    categories = ["Male", "Female", "Other"]
    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        total = len(scoped)
        cell = "<br>".join(f"{category}: {sum(1 for row in scoped if row['gender'] == category)} ({(100 * sum(1 for row in scoped if row['gender'] == category) / total):.1f}%)" if total else f"{category}: 0" for category in categories)
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{total}</td><td>{cell}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Gender counts</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    return table_shell("Descriptives: gender", IN_BODY, "Gender is read from the immediate survey row and grouped as Male, Female, or Other.", table_html, categorical_bars("gender-bars", complete_rows, "gender", "Gender distribution", categories), status_messages(errors, warnings, "Gender values were available for all displayed participants."), excluded_details(excluded) + first_three_details(complete_rows, ["gender", "raw_gender"]))


def descriptive_creatures_seen() -> str:
    """Open survey_export.tsv; count unique valid creature IDs in SEEN, falling back to seen_* columns; then display numeric descriptives by condition."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")
    if not survey_rows and not errors:
        errors.append("survey_export.tsv was empty or could not be read.")

    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)

    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    invalid_seen_by_mcid: list[str] = []
    for participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition"})
            continue
        seen_ids, invalid_ids, source = seen_creatures_from_row(row)
        if not source:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing SEEN and seen_* creature exposure fields"})
            continue
        if invalid_ids:
            invalid_seen_by_mcid.append(f"{participant_id}: " + ", ".join(invalid_ids[:5]))
        value = len(seen_ids)
        if value < 0 or value > MAX_RETENTION_SLOTS:
            errors.append(f"Creatures-seen count outside 0-{MAX_RETENTION_SLOTS} for {participant_id}.")
            excluded.append({"MCID": participant_id, "condition": condition, "reason": f"Out-of-range creatures-seen count: {value}"})
            continue
        complete_rows.append({
            "MCID": participant_id,
            "condition": condition,
            "value": float(value),
            "creatures_seen": str(value),
            "seen_source": source,
            "seen_ids": ", ".join(seen_ids),
        })

    if invalid_seen_by_mcid:
        warnings.append("Ignored invalid/non-configured SEEN entries for " + str(len(invalid_seen_by_mcid)) + " participant(s). Examples: " + "; ".join(invalid_seen_by_mcid[:5]) + ".")
    if excluded:
        warnings.append(f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays {len(complete_rows)} after complete-case exclusion.")

    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        values = [float(row["value"]) for row in scoped]
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{len(values)}</td><td>{mean_sd_min_max_html(values)}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Creatures seen</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    return table_shell(
        "Descriptives: creatures seen",
        IN_BODY,
        "Creature exposure is the number of unique valid configured creature IDs in the immediate survey row. The SEEN field is used first because it drives retention slots; seen_* columns are used only as a fallback.",
        table_html,
        boxplot_svg("creatures-seen-boxplot", complete_rows, "value", "Creatures-seen boxplot", min_value=0, max_value=MAX_RETENTION_SLOTS),
        status_messages(errors, warnings, "Creatures-seen values were available for all displayed participants."),
        excluded_details(excluded) + first_three_details(complete_rows, ["creatures_seen", "seen_source", "seen_ids"]),
    )


def descriptive_perceived_control() -> str:
    """Open survey_export.tsv; take MCID, condition, ctrl_scores_1, and ctrl_scores_2 only; require both 1-7 values; average the two items; then display numeric descriptives by condition."""
    IN_BODY = False
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")
    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)
    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    for participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        item1 = parse_float(row.get("ctrl_scores_1"))
        item2 = parse_float(row.get("ctrl_scores_2"))
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition"})
            continue
        if item1 is None or item2 is None:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing perceived-control item(s)"})
            continue
        if item1 < 1 or item1 > 7 or item2 < 1 or item2 > 7:
            errors.append(f"Perceived-control value outside 1-7 for {participant_id}.")
            excluded.append({"MCID": participant_id, "condition": condition, "reason": f"Out-of-range item(s): {item1}, {item2}"})
            continue
        value = (item1 + item2) / 2
        complete_rows.append({"MCID": participant_id, "condition": condition, "value": value, "ctrl_scores_1": item1, "ctrl_scores_2": item2, "perceived_control": f"{value:.2f}"})
    if excluded:
        warnings.append(f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays {len(complete_rows)} after complete-case exclusion.")
    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        values = [float(row["value"]) for row in scoped]
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{len(values)}</td><td>{mean_sd_min_max_html(values)}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Perceived control</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    return table_shell("Descriptives: perceived control", IN_BODY, "Manipulation-check score: mean of ctrl_scores_1 and ctrl_scores_2. Both 1-7 items are required for complete-case inclusion.", table_html, boxplot_svg("control-boxplot", complete_rows, "value", "Perceived-control boxplot", min_value=1, max_value=7), status_messages(errors, warnings, "All displayed perceived-control scores were based on two valid items."), excluded_details(excluded) + first_three_details(complete_rows, ["perceived_control", "ctrl_scores_1", "ctrl_scores_2"]))


def descriptive_retention(moment_label: str) -> str:
    """Display participant-level retention descriptives from retention_scores_final.tsv."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")

    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)

    condition_by_mcid: dict[str, str] = {}
    for participant_id, row in immediate_by_mcid.items():
        condition_by_mcid[participant_id] = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))

    score_source = retention_score_file_label()
    participant_scores, score_errors, score_warnings = participant_retention_scores_from_final()
    errors.extend(score_errors)
    warnings.extend(score_warnings)

    moment_key = "Immediate" if moment_label.lower().startswith("immediate") else "Delayed"
    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []

    for participant_id, scores in sorted(participant_scores.items()):
        condition = condition_by_mcid.get(participant_id, "")
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition in survey_export.tsv"})
            continue
        value = scores.get(moment_key)
        if value is None:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": f"No complete {moment_key.lower()} participant-level retention value in {score_source}"})
            continue
        if value < 0 or value > 2:
            errors.append(f"{moment_key} retention outside 0-2 for {participant_id}.")
            excluded.append({"MCID": participant_id, "condition": condition, "reason": f"Out-of-range {moment_key} retention: {value}"})
            continue
        complete_rows.append({
            "MCID": participant_id,
            "condition": condition,
            "value": value,
            "retention_score": f"{value:.2f}",
            "source_column": score_source,
        })

    if excluded:
        warnings.append(
            f"{score_source} yielded {len(participant_scores)} participant row(s), but this table displays "
            f"{len(complete_rows)} complete {moment_label.lower()} retention score(s)."
        )

    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        values = [float(row["value"]) for row in scoped]
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{len(values)}</td><td>{mean_sd_min_max_html(values)}</td></tr>')
    table_html = f'<table><thead><tr><th>Group</th><th>n</th><th>{h(moment_label)} retention score</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    return table_shell(
        f"Descriptives: {moment_label.lower()} retention",
        IN_BODY,
        f"{moment_label} participant-level retention is read directly from {score_source}. Scores remain on the 0-2 retention scale.",
        table_html,
        boxplot_svg(f"{moment_label.lower()}-retention-boxplot", complete_rows, "value", f"{moment_label} retention boxplot", min_value=0, max_value=2),
        status_messages(errors, warnings, f"All displayed {moment_label.lower()} retention scores came from {score_source}."),
        excluded_details(excluded) + first_three_details(complete_rows, ["retention_score", "source_column"]),
    )


def descriptive_retention_decay() -> str:
    """Display participant-level retention change/decay descriptives from paired immediate and delayed participant-level final-score-file values."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")

    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)

    condition_by_mcid: dict[str, str] = {}
    for participant_id, row in immediate_by_mcid.items():
        condition_by_mcid[participant_id] = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))

    score_source = retention_score_file_label()
    participant_scores, score_errors, score_warnings = participant_retention_scores_from_final()
    errors.extend(score_errors)
    warnings.extend(score_warnings)

    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    for participant_id, scores in sorted(participant_scores.items()):
        condition = condition_by_mcid.get(participant_id, "")
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition in survey_export.tsv"})
            continue
        immediate = scores.get("Immediate")
        delayed = scores.get("Delayed")
        if immediate is None or delayed is None:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": f"Requires both complete immediate and delayed participant-level retention values in {score_source}"})
            continue
        if immediate < 0 or immediate > 2 or delayed < 0 or delayed > 2:
            errors.append(f"Immediate or delayed retention outside 0-2 for {participant_id}.")
            excluded.append({"MCID": participant_id, "condition": condition, "reason": f"Out-of-range immediate/delayed retention: {immediate}, {delayed}"})
            continue
        value = float(delayed) - float(immediate)
        complete_rows.append({
            "MCID": participant_id,
            "condition": condition,
            "value": value,
            "retention_decay": f"{value:.2f}",
            "immediate_retention": f"{float(immediate):.2f}",
            "delayed_retention": f"{float(delayed):.2f}",
            "source_column": f"delayed - immediate; {score_source}",
        })

    if excluded:
        warnings.append(
            f"{score_source} yielded {len(participant_scores)} participant row(s), but this table displays "
            f"{len(complete_rows)} paired immediate-delayed retention-change score(s)."
        )

    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        values = [float(row["value"]) for row in scoped]
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{len(values)}</td><td>{mean_sd_min_max_html(values)}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Retention decay/change: delayed − immediate</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    return table_shell(
        "Descriptives: retention decay",
        IN_BODY,
        f"Participant-level retention decay/change is calculated as delayed retention minus immediate retention after both occasion scores have been read from {score_source}. Negative values indicate lower delayed than immediate retention; positive values indicate improvement.",
        table_html,
        boxplot_svg("retention-decay-boxplot", complete_rows, "value", "Retention decay/change boxplot", min_value=-2, max_value=2),
        status_messages(errors, warnings, "All displayed retention-decay values are paired participant-level delayed-minus-immediate differences."),
        excluded_details(excluded) + first_three_details(complete_rows, ["retention_decay", "immediate_retention", "delayed_retention", "source_column"]),
    )


def _mean_valid_scale_items(row: dict[str, str], columns: list[str], *, minimum: float, maximum: float) -> tuple[float | None, list[str], list[str]]:
    """Mirror sum_merged's scale handling: invalid/missing values are omitted from the mean, not forced to zero."""
    values: list[float] = []
    missing: list[str] = []
    out_of_range: list[str] = []
    for column in columns:
        value = parse_float(row.get(column))
        if value is None:
            missing.append(column)
        elif value < minimum or value > maximum:
            out_of_range.append(f"{column}={value}")
        else:
            values.append(value)
    if not values:
        return None, missing, out_of_range
    return statistics.fmean(values), missing, out_of_range


ScaleItemSpec = tuple[str, float, float, bool]


def _format_cronbach_alpha(value: float | None) -> str:
    if value is None or not math.isfinite(value):
        return "—"
    return f"{value:.3f}".replace("0.", ".").replace("-0.", "-.")


def _score_scale_item(row: dict[str, str], spec: ScaleItemSpec) -> float | None:
    column, minimum, maximum, reverse = spec
    value = parse_float(row.get(column))
    if value is None or value < minimum or value > maximum:
        return None
    if reverse:
        return minimum + maximum - value
    return value


def _cronbach_alpha_from_vectors(vectors: list[list[float]]) -> float | None:
    """Calculate Cronbach's alpha over complete item/component vectors."""
    if len(vectors) < 2:
        return None
    item_count = len(vectors[0]) if vectors else 0
    if item_count < 2:
        return None
    if any(len(vector) != item_count for vector in vectors):
        return None

    item_variances = [statistics.variance([vector[index] for vector in vectors]) for index in range(item_count)]
    total_scores = [sum(vector) for vector in vectors]
    total_variance = statistics.variance(total_scores)
    if total_variance <= 0:
        return None
    return item_count / (item_count - 1) * (1 - sum(item_variances) / total_variance)


def _scale_vectors_from_item_specs(
    immediate_by_mcid: dict[str, dict[str, str]],
    item_specs: list[ScaleItemSpec],
) -> tuple[list[list[float]], int]:
    """Return complete participant vectors for item-level alpha and the eligible n."""
    vectors: list[list[float]] = []
    eligible_n = 0
    for _participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        if condition not in CONDITION_ORDER:
            continue
        eligible_n += 1
        values: list[float] = []
        for spec in item_specs:
            value = _score_scale_item(row, spec)
            if value is None:
                values = []
                break
            values.append(value)
        if values:
            vectors.append(values)
    return vectors, eligible_n


def _scale_vectors_from_component_specs(
    immediate_by_mcid: dict[str, dict[str, str]],
    component_specs: list[list[ScaleItemSpec]],
) -> tuple[list[list[float]], int]:
    """Return complete participant vectors for alpha over intermediate component means."""
    vectors: list[list[float]] = []
    eligible_n = 0
    for _participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        if condition not in CONDITION_ORDER:
            continue
        eligible_n += 1
        component_values: list[float] = []
        complete = True
        for specs in component_specs:
            values: list[float] = []
            for spec in specs:
                value = _score_scale_item(row, spec)
                if value is None:
                    complete = False
                    break
                values.append(value)
            if not complete or not values:
                complete = False
                break
            component_values.append(statistics.fmean(values))
        if complete:
            vectors.append(component_values)
    return vectors, eligible_n


def descriptive_chapter_scale(
    *,
    title: str,
    description: str,
    item_indices: tuple[int, ...],
    column_prefix: str,
    minimum: float,
    maximum: float,
    status_ok: str,
) -> str:
    """Display per-chapter participant-level scale descriptives by condition.

    This follows summarise_merged's participant-level logic: calculate each participant's
    chapter score as the mean of the valid item responses for that chapter, then summarise
    those participant scores by condition. It therefore does not collapse across chapters
    before the chapter descriptives are produced.
    """
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")

    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)

    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    for participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition"})
            continue

        result_row: dict[str, Any] = {"MCID": participant_id, "condition": condition}
        issue_parts: list[str] = []
        valid_chapter_count = 0
        for chapter in (1, 2, 3):
            columns = [f"{column_prefix}_ch{chapter}_scores_{index}" for index in item_indices]
            value, missing, out_of_range = _mean_valid_scale_items(row, columns, minimum=minimum, maximum=maximum)
            result_row[f"ch{chapter}"] = value
            result_row[f"ch{chapter}_score"] = "" if value is None else f"{value:.2f}"
            if value is not None:
                valid_chapter_count += 1
            if missing:
                issue_parts.append(f"Ch{chapter} missing: " + ", ".join(missing))
            if out_of_range:
                errors.append(f"{title} value outside {minimum:g}-{maximum:g} for {participant_id}.")
                issue_parts.append(f"Ch{chapter} out of range: " + ", ".join(out_of_range))

        if valid_chapter_count == 0:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "; ".join(issue_parts) or "No valid chapter score"})
            continue
        if issue_parts:
            warnings.append(f"{participant_id}: " + "; ".join(issue_parts))
        complete_rows.append(result_row)

    if excluded:
        warnings.append(
            f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays "
            f"{len(complete_rows)} participant row(s) with at least one valid chapter score."
        )

    table_rows: list[str] = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        chapter_cells = []
        for chapter in (1, 2, 3):
            values = [float(row[f"ch{chapter}"]) for row in scoped if row.get(f"ch{chapter}") is not None]
            chapter_cells.append(f"<td>{n_mean_sd_min_max_html(values)}</td>")
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th>' + "".join(chapter_cells) + "</tr>")

    table_html = (
        '<table><thead><tr><th>Group</th><th>Ch1</th><th>Ch2</th><th>Ch3</th></tr></thead><tbody>'
        + "".join(table_rows)
        + '</tbody></table>'
    )
    return table_shell(
        title,
        IN_BODY,
        description,
        table_html,
        "",
        status_messages(errors, warnings, status_ok),
        excluded_details(excluded) + first_three_details(complete_rows, ["ch1_score", "ch2_score", "ch3_score"]),
    )


def descriptive_intrinsic_load_by_chapter() -> str:
    return descriptive_chapter_scale(
        title="Descriptives: intrinsic cognitive load by chapter",
        description="Intrinsic cognitive load by chapter is calculated like sum_merged: for each immediate participant row and chapter, valid cl_ch*_scores_1 through cl_ch*_scores_3 values are averaged first; those participant-level chapter means are then summarised by condition. Scores range from 0 to 10.",
        item_indices=(1, 2, 3),
        column_prefix="cl",
        minimum=0,
        maximum=10,
        status_ok="All displayed intrinsic chapter scores were calculated from valid 0-10 chapter-level cognitive-load items.",
    )


def descriptive_extraneous_load_by_chapter() -> str:
    return descriptive_chapter_scale(
        title="Descriptives: extraneous cognitive load by chapter",
        description="Chapter-level extraneous cognitive load follows sum_merged's per-chapter table and uses the environment-related extraneous cognitive-load items only: for each immediate participant row and chapter, valid cl_ch*_scores_4 through cl_ch*_scores_7 values are averaged first; those participant-level chapter means are then summarised by condition. Instruction- and interaction-related extraneous load are overall-game items and therefore do not have chapter-specific scores. Scores range from 0 to 10.",
        item_indices=(4, 5, 6, 7),
        column_prefix="cl",
        minimum=0,
        maximum=10,
        status_ok="All displayed extraneous chapter scores were calculated from valid 0-10 environment-related chapter-level cognitive-load items.",
    )


def descriptive_engagement_by_chapter() -> str:
    return descriptive_chapter_scale(
        title="Descriptives: engagement by chapter",
        description="Engagement by chapter is calculated like sum_merged: for each immediate participant row and chapter, valid eng_ch*_scores_1 through eng_ch*_scores_5 values are averaged first; those participant-level chapter means are then summarised by condition. Scores range from 1 to 7.",
        item_indices=(1, 2, 3, 4, 5),
        column_prefix="eng",
        minimum=1,
        maximum=7,
        status_ok="All displayed engagement chapter scores were calculated from valid 1-7 chapter-level engagement items.",
    )


def internal_consistency_cognitive_load_engagement() -> str:
    """Report Cronbach's alpha only for Cognitive Load and Engagement scales.

    Full-construct alpha follows the same aggregation logic as the final score:
    when a score is based on intermediate chapter/component means, alpha is
    calculated over those intermediate means rather than over all raw items equally.
    """
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")

    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)

    table_rows: list[str] = []

    def cl(column: str, reverse: bool = False) -> ScaleItemSpec:
        return (column, 0.0, 10.0, reverse)

    def eng(column: str, reverse: bool = False) -> ScaleItemSpec:
        return (column, 1.0, 7.0, reverse)

    def add_item_alpha(construct: str, scope: str, item_specs: list[ScaleItemSpec], basis: str) -> None:
        vectors, eligible_n = _scale_vectors_from_item_specs(immediate_by_mcid, item_specs)
        alpha = _cronbach_alpha_from_vectors(vectors)
        excluded_n = max(0, eligible_n - len(vectors))
        table_rows.append(
            f'<tr><th>{h(construct)}</th><td>{h(scope)}</td><td>{len(vectors)}</td><td>{len(item_specs)}</td>'
            f'<td>{_format_cronbach_alpha(alpha)}</td><td>{excluded_n}</td><td>{h(basis)}</td></tr>'
        )

    def add_component_alpha(construct: str, scope: str, component_specs: list[list[ScaleItemSpec]], basis: str) -> None:
        vectors, eligible_n = _scale_vectors_from_component_specs(immediate_by_mcid, component_specs)
        alpha = _cronbach_alpha_from_vectors(vectors)
        excluded_n = max(0, eligible_n - len(vectors))
        table_rows.append(
            f'<tr><th>{h(construct)}</th><td>{h(scope)}</td><td>{len(vectors)}</td><td>{len(component_specs)}</td>'
            f'<td>{_format_cronbach_alpha(alpha)}</td><td>{excluded_n}</td><td>{h(basis)}</td></tr>'
        )

    def add_not_applicable(construct: str, scope: str, reason: str) -> None:
        table_rows.append(
            f'<tr><th>{h(construct)}</th><td>{h(scope)}</td><td>—</td><td>—</td><td>—</td><td>—</td><td>{h(reason)}</td></tr>'
        )

    intrinsic_chapter_specs = [
        [cl(f"cl_ch{chapter}_scores_{index}") for index in (1, 2, 3)]
        for chapter in (1, 2, 3)
    ]
    for chapter, specs in zip((1, 2, 3), intrinsic_chapter_specs, strict=True):
        add_item_alpha("Intrinsic cognitive load", f"Chapter {chapter}", specs, "Three chapter-specific intrinsic-load items.")
    add_not_applicable("Intrinsic cognitive load", "Game overall", "No game-overall intrinsic-load items were administered.")
    add_component_alpha("Intrinsic cognitive load", "Full construct", intrinsic_chapter_specs, "Three chapter means, matching the final equal-chapter aggregation.")

    extraneous_environment_chapter_specs = [
        [cl(f"cl_ch{chapter}_scores_{index}") for index in (4, 5, 6, 7)]
        for chapter in (1, 2, 3)
    ]
    for chapter, specs in zip((1, 2, 3), extraneous_environment_chapter_specs, strict=True):
        add_item_alpha("Extraneous cognitive load", f"Chapter {chapter}", specs, "Four environment-related extraneous-load items.")
    extraneous_instruction_specs = [cl(f"cl_overall_scores_{index}", reverse=True) for index in (1, 2, 3)]
    extraneous_interaction_specs = [cl(f"cl_overall_scores_{index}", reverse=True) for index in (4, 5, 6, 7)]
    add_item_alpha("Extraneous cognitive load", "Game overall", extraneous_instruction_specs + extraneous_interaction_specs, "Seven overall instruction- and interaction-related items after reverse-coding.")
    extraneous_environment_specs = [spec for chapter_specs in extraneous_environment_chapter_specs for spec in chapter_specs]
    add_component_alpha("Extraneous cognitive load", "Full construct", [extraneous_environment_specs, extraneous_instruction_specs, extraneous_interaction_specs], "Environment, instruction, and interaction means, matching the final equal-component aggregation.")

    for chapter in (1, 2, 3):
        add_not_applicable("Germane cognitive load", f"Chapter {chapter}", "No chapter-specific germane-load items were administered.")
    germane_overall_specs = [cl(f"cl_overall_scores_{index}") for index in (8, 9, 10, 11)]
    add_item_alpha("Germane cognitive load", "Game overall", germane_overall_specs, "Four overall germane-load items.")
    add_item_alpha("Germane cognitive load", "Full construct", germane_overall_specs, "Same four overall germane-load items; no intermediate aggregation was used.")

    engagement_chapter_specs = [
        [eng(f"eng_ch{chapter}_scores_{index}") for index in (1, 2, 3, 4, 5)]
        for chapter in (1, 2, 3)
    ]
    for chapter, specs in zip((1, 2, 3), engagement_chapter_specs, strict=True):
        add_item_alpha("Engagement", f"Chapter {chapter}", specs, "Five chapter-specific engagement items.")
    engagement_overall_specs = [
        eng("eng_overall_scores_1", reverse=True),
        eng("eng_overall_scores_2", reverse=True),
        eng("eng_overall_scores_3"),
        eng("eng_overall_scores_4"),
    ]
    add_item_alpha("Engagement", "Game overall", engagement_overall_specs, "Four overall engagement items after reverse-coding frustration and confusion.")
    engagement_all_chapter_specs = [spec for chapter_specs in engagement_chapter_specs for spec in chapter_specs]
    add_component_alpha("Engagement", "Full construct", [engagement_all_chapter_specs, engagement_overall_specs], "Chapter-specific engagement mean and overall engagement mean, matching the final equal-component aggregation.")

    table_html = (
        '<table><thead><tr><th>Construct</th><th>Scope</th><th>complete n</th><th>input units k</th>'
        '<th>Cronbach&apos;s α</th><th>excluded n</th><th>Input used</th></tr></thead><tbody>'
        + "".join(table_rows)
        + '</tbody></table>'
    )
    if not immediate_by_mcid:
        warnings.append("No immediate survey rows were available for internal-consistency estimates.")
    return table_shell(
        "Internal consistency: cognitive load and engagement",
        IN_BODY,
        "Cronbach's alpha is reported only for Cognitive Load and Engagement. Estimates are calculated after reverse-coding, and full-construct estimates use the same intermediate chapter/component means as the final scale scores rather than weighting all raw items equally.",
        table_html,
        "",
        status_messages(errors, warnings, "Internal-consistency estimates were generated for Cognitive Load and Engagement only."),
        "",
    )


def descriptive_intrinsic_load() -> str:
    """Open survey_export.tsv; take condition and cl_ch[1-3]_scores_[1-3]; require all nine 0-10 items; average them; then display numeric descriptives by condition."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")
    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)
    needed = [f"cl_ch{chapter}_scores_{index}" for chapter in (1, 2, 3) for index in (1, 2, 3)]
    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    for participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition"})
            continue
        values = []
        missing = []
        out_of_range = []
        for column in needed:
            value = parse_float(row.get(column))
            if value is None:
                missing.append(column)
            elif value < 0 or value > 10:
                out_of_range.append(f"{column}={value}")
            else:
                values.append(value)
        if missing or out_of_range:
            if out_of_range:
                errors.append(f"Intrinsic-load value outside 0-10 for {participant_id}.")
            excluded.append({"MCID": participant_id, "condition": condition, "reason": ("Missing: " + ", ".join(missing) if missing else "") + ("; Out of range: " + ", ".join(out_of_range) if out_of_range else "")})
            continue
        value = statistics.fmean(values)
        complete_rows.append({"MCID": participant_id, "condition": condition, "value": value, "intrinsic_load": f"{value:.2f}", "item_count": str(len(values))})
    if excluded:
        warnings.append(f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays {len(complete_rows)} after complete-case exclusion.")
    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        values = [float(row["value"]) for row in scoped]
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{len(values)}</td><td>{mean_sd_min_max_html(values)}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Intrinsic cognitive load</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    return table_shell("Descriptives: intrinsic cognitive load", IN_BODY, "Intrinsic cognitive load is the mean of nine required 0-10 chapter-level items.", table_html, boxplot_svg("intrinsic-boxplot", complete_rows, "value", "Intrinsic-load boxplot", min_value=0, max_value=10), status_messages(errors, warnings, "All displayed intrinsic-load scores use nine valid 0-10 items."), excluded_details(excluded) + first_three_details(complete_rows, ["intrinsic_load", "item_count"]))


def descriptive_extraneous_load() -> str:
    """Open survey_export.tsv; take condition plus environment-, instruction-, and interaction-extraneous CL items; require all nineteen 0-10 items; reverse-code instruction/interaction items; average the three subscale means; then display numeric descriptives by condition."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")
    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)
    env_cols = [f"cl_ch{chapter}_scores_{index}" for chapter in (1, 2, 3) for index in (4, 5, 6, 7)]
    instruction_cols = [f"cl_overall_scores_{index}" for index in (1, 2, 3)]
    interaction_cols = [f"cl_overall_scores_{index}" for index in (4, 5, 6, 7)]
    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    for participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition"})
            continue
        missing: list[str] = []
        out_of_range: list[str] = []
        env_values: list[float] = []
        instruction_values: list[float] = []
        interaction_values: list[float] = []
        for column in env_cols:
            value = parse_float(row.get(column))
            if value is None: missing.append(column)
            elif value < 0 or value > 10: out_of_range.append(f"{column}={value}")
            else: env_values.append(value)
        for column in instruction_cols:
            value = parse_float(row.get(column))
            if value is None: missing.append(column)
            elif value < 0 or value > 10: out_of_range.append(f"{column}={value}")
            else: instruction_values.append(10 - value)
        for column in interaction_cols:
            value = parse_float(row.get(column))
            if value is None: missing.append(column)
            elif value < 0 or value > 10: out_of_range.append(f"{column}={value}")
            else: interaction_values.append(10 - value)
        if missing or out_of_range:
            if out_of_range:
                errors.append(f"Extraneous-load value outside 0-10 for {participant_id}.")
            excluded.append({"MCID": participant_id, "condition": condition, "reason": ("Missing: " + ", ".join(missing) if missing else "") + ("; Out of range: " + ", ".join(out_of_range) if out_of_range else "")})
            continue
        value = statistics.fmean([statistics.fmean(env_values), statistics.fmean(instruction_values), statistics.fmean(interaction_values)])
        complete_rows.append({"MCID": participant_id, "condition": condition, "value": value, "extraneous_load": f"{value:.2f}", "environment_mean": f"{statistics.fmean(env_values):.2f}", "instruction_mean_reversed": f"{statistics.fmean(instruction_values):.2f}", "interaction_mean_reversed": f"{statistics.fmean(interaction_values):.2f}"})
    if excluded:
        warnings.append(f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays {len(complete_rows)} after complete-case exclusion.")
    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        values = [float(row["value"]) for row in scoped]
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{len(values)}</td><td>{mean_sd_min_max_html(values)}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Extraneous cognitive load</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    return table_shell("Descriptives: extraneous cognitive load", IN_BODY, "Extraneous load is the mean of environment-, reverse-coded instruction-, and reverse-coded interaction-extraneous subscale means. All nineteen 0-10 items are required.", table_html, boxplot_svg("extraneous-boxplot", complete_rows, "value", "Extraneous-load boxplot", min_value=0, max_value=10), status_messages(errors, warnings, "All displayed extraneous-load scores use all required valid 0-10 items."), excluded_details(excluded) + first_three_details(complete_rows, ["extraneous_load", "environment_mean", "instruction_mean_reversed", "interaction_mean_reversed"]))


def descriptive_germane_load() -> str:
    """Open survey_export.tsv; take condition and cl_overall_scores_8 through cl_overall_scores_11; require all four 0-10 items; average them; then display numeric descriptives by condition."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")
    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)
    needed = [f"cl_overall_scores_{index}" for index in (8, 9, 10, 11)]
    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    for participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition"})
            continue
        values: list[float] = []
        missing: list[str] = []
        out_of_range: list[str] = []
        for column in needed:
            value = parse_float(row.get(column))
            if value is None:
                missing.append(column)
            elif value < 0 or value > 10:
                out_of_range.append(f"{column}={value}")
            else:
                values.append(value)
        if missing or out_of_range:
            if out_of_range:
                errors.append(f"Germane-load value outside 0-10 for {participant_id}.")
            excluded.append({"MCID": participant_id, "condition": condition, "reason": ("Missing: " + ", ".join(missing) if missing else "") + ("; Out of range: " + ", ".join(out_of_range) if out_of_range else "")})
            continue
        value = statistics.fmean(values)
        complete_rows.append({"MCID": participant_id, "condition": condition, "value": value, "germane_load": f"{value:.2f}", "item_count": str(len(values))})
    if excluded:
        warnings.append(f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays {len(complete_rows)} after complete-case exclusion.")
    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        values = [float(row["value"]) for row in scoped]
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{len(values)}</td><td>{mean_sd_min_max_html(values)}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Germane cognitive load</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    return table_shell("Descriptives: germane cognitive load", IN_BODY, "Germane cognitive load is the mean of cl_overall_scores_8 through cl_overall_scores_11. All four 0-10 items are required.", table_html, boxplot_svg("germane-boxplot", complete_rows, "value", "Germane-load boxplot", min_value=0, max_value=10), status_messages(errors, warnings, "All displayed germane-load scores use four valid 0-10 items."), excluded_details(excluded) + first_three_details(complete_rows, ["germane_load", "item_count"]))


def descriptive_engagement() -> str:
    """Open survey_export.tsv; take condition, all engagement chapter items, and all engagement overall items; require all nineteen 1-7 items; reverse-code frustration/confusion; average chapter and overall engagement; then display numeric descriptives by condition."""
    IN_BODY = True
    errors: list[str] = []
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    if not SURVEY_EXPORT_PATH.exists():
        errors.append(f"Missing survey file: {SURVEY_EXPORT_PATH}")
    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if participant_id and not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)
    chapter_cols = [f"eng_ch{chapter}_scores_{index}" for chapter in (1, 2, 3) for index in (1, 2, 3, 4, 5)]
    overall_cols = ["eng_overall_scores_1", "eng_overall_scores_2", "eng_overall_scores_3", "eng_overall_scores_4"]
    complete_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    for participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        if condition not in CONDITION_ORDER:
            excluded.append({"MCID": participant_id, "condition": condition, "reason": "Missing or invalid condition"})
            continue
        missing: list[str] = []
        out_of_range: list[str] = []
        chapter_values: list[float] = []
        overall_values: list[float] = []
        for column in chapter_cols:
            value = parse_float(row.get(column))
            if value is None: missing.append(column)
            elif value < 1 or value > 7: out_of_range.append(f"{column}={value}")
            else: chapter_values.append(value)
        for column in overall_cols:
            value = parse_float(row.get(column))
            if value is None: missing.append(column)
            elif value < 1 or value > 7: out_of_range.append(f"{column}={value}")
            else:
                if column in {"eng_overall_scores_1", "eng_overall_scores_2"}:
                    overall_values.append(8 - value)
                else:
                    overall_values.append(value)
        if missing or out_of_range:
            if out_of_range:
                errors.append(f"Engagement value outside 1-7 for {participant_id}.")
            excluded.append({"MCID": participant_id, "condition": condition, "reason": ("Missing: " + ", ".join(missing) if missing else "") + ("; Out of range: " + ", ".join(out_of_range) if out_of_range else "")})
            continue
        chapter_mean = statistics.fmean(chapter_values)
        overall_mean = statistics.fmean(overall_values)
        value = statistics.fmean([chapter_mean, overall_mean])
        complete_rows.append({"MCID": participant_id, "condition": condition, "value": value, "engagement": f"{value:.2f}", "chapter_mean": f"{chapter_mean:.2f}", "overall_mean_reversed": f"{overall_mean:.2f}"})
    if excluded:
        warnings.append(f"The survey contained {len(immediate_by_mcid)} immediate participant row(s), but this table displays {len(complete_rows)} after complete-case exclusion.")
    table_rows = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = complete_rows if condition == "Total" else [row for row in complete_rows if row["condition"] == condition]
        values = [float(row["value"]) for row in scoped]
        table_rows.append(f'<tr><th>{h(DISPLAY_CONDITION[condition])}</th><td>{len(values)}</td><td>{mean_sd_min_max_html(values)}</td></tr>')
    table_html = '<table><thead><tr><th>Group</th><th>n</th><th>Engagement</th></tr></thead><tbody>' + "".join(table_rows) + '</tbody></table>'
    return table_shell("Descriptives: engagement", IN_BODY, "Engagement is the mean of the chapter-engagement mean and the reverse-coded overall-engagement mean. All nineteen 1-7 items are required.", table_html, boxplot_svg("engagement-boxplot", complete_rows, "value", "Engagement boxplot", min_value=1, max_value=7), status_messages(errors, warnings, "All displayed engagement scores use all required valid 1-7 items."), excluded_details(excluded) + first_three_details(complete_rows, ["engagement", "chapter_mean", "overall_mean_reversed"]))



# -----------------------------------------------------------------------------
# ASSUMPTION-CHECK HELPERS
# -----------------------------------------------------------------------------


def slugify(value: object) -> str:
    text = re.sub(r"[^a-zA-Z0-9]+", "-", clean(value).lower()).strip("-")
    return text or "section"


def list_html(items: list[str]) -> str:
    return "<ul>" + "".join(f"<li>{h(item)}</li>" for item in items) + "</ul>"


def p_text(value: float | None) -> str:
    if value is None or not math.isfinite(value):
        return "—"
    if value < 0.001:
        return "< .001"
    return f"{value:.3f}".replace("0.", ".")


def interpretation_badge(ok: bool | None, ok_text: str, warning_text: str) -> str:
    if ok is None:
        return '<span class="badge badge-neutral">not available</span>'
    if ok:
        return f'<span class="badge badge-good">{h(ok_text)}</span>'
    return f'<span class="badge badge-warning">{h(warning_text)}</span>'


def make_complete_cases(rows: list[dict[str, Any]], required_columns: list[str], label: str) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    complete: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    for row in rows:
        missing: list[str] = []
        for column in required_columns:
            value = row.get(column)
            if column in {"condition", "location", "gender", "completed_delayed", "completed_delayed_retention", "interview_status", "MCID"}:
                if not clean(value):
                    missing.append(column)
            elif parse_float(value) is None:
                missing.append(column)
        if missing:
            excluded.append({
                "MCID": clean(row.get("MCID")),
                "condition": clean(row.get("condition")),
                "reason": f"{label}: missing/invalid " + ", ".join(missing),
            })
        else:
            complete.append(row)
    return complete, excluded


def condition_count_table(rows: list[dict[str, Any]], title: str, extra_columns: list[str] | None = None) -> str:
    extra_columns = extra_columns or []
    header = '<tr><th>Group</th><th>n</th>' + ''.join(f'<th>{h(column)}</th>' for column in extra_columns) + '</tr>'
    body_rows: list[str] = []
    for condition in CONDITION_ORDER + ["Total"]:
        scoped = rows if condition == "Total" else [row for row in rows if row.get("condition") == condition]
        cells = [f'<th>{h(DISPLAY_CONDITION[condition])}</th>', f'<td>{len(scoped)}</td>']
        for column in extra_columns:
            values = [row.get(column) for row in scoped if clean(row.get(column))]
            if not values:
                cells.append('<td>—</td>')
            elif all(parse_float(value) is not None for value in values):
                nums = [float(parse_float(value)) for value in values if parse_float(value) is not None]
                cells.append(f'<td>{mean_sd_min_max_html(nums)}</td>')
            else:
                counts: dict[str, int] = {}
                for value in values:
                    counts[clean(value)] = counts.get(clean(value), 0) + 1
                cells.append('<td>' + '<br>'.join(f'{h(k)}: {v}' for k, v in sorted(counts.items())) + '</td>')
        body_rows.append('<tr>' + ''.join(cells) + '</tr>')
    return f'<h4>{h(title)}</h4><div class="table-wrap"><table><thead>{header}</thead><tbody>' + ''.join(body_rows) + '</tbody></table></div>'


def design_matrix_for_assumptions(rows: list[dict[str, Any]], numeric_predictors: list[str], categorical_predictors: list[str]) -> tuple[np.ndarray | None, list[str], list[str]]:
    if not rows:
        return None, [], ["No complete rows available for a diagnostic design matrix."]
    columns: list[list[float]] = [[1.0 for _ in rows]]
    names = ["Intercept"]
    notes: list[str] = []
    for predictor in numeric_predictors:
        values: list[float] = []
        for row in rows:
            value = parse_float(row.get(predictor))
            if value is None:
                notes.append(f"Numeric predictor {predictor} was missing after complete-case filtering.")
                return None, names, notes
            values.append(float(value))
        columns.append(values)
        names.append(predictor)
    for predictor in categorical_predictors:
        levels = sorted({clean(row.get(predictor)) for row in rows if clean(row.get(predictor))})
        if len(levels) < 2:
            notes.append(f"Categorical predictor {predictor} has fewer than two levels in complete cases and cannot be diagnosed as a covariate.")
            continue
        if predictor == "location":
            intended = [level for level in ["CreaSp", "LivingR", "Remote"] if level in levels]
            remaining = [level for level in levels if level not in intended]
            levels = intended + remaining
            reference = "CreaSp" if "CreaSp" in levels else levels[0]
        elif predictor == "gender":
            reference = "Male" if "Male" in levels else levels[0]
        else:
            reference = levels[0]
        compared = ", ".join(level for level in levels if level != reference) or "none"
        notes.append(f"{predictor} diagnostic coding: reference={reference}; compared levels={compared}.")
        for level in levels:
            if level == reference:
                continue
            columns.append([1.0 if clean(row.get(predictor)) == level else 0.0 for row in rows])
            names.append(f"{predictor}={level}")
    return np.array(columns, dtype=float).T, names, notes


def vif_table_from_matrix(x_matrix: np.ndarray, names: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if x_matrix.shape[1] <= 2:
        return rows
    for index in range(1, x_matrix.shape[1]):
        y = x_matrix[:, index]
        others = np.delete(x_matrix, index, axis=1)
        if np.std(y) == 0 or others.shape[1] < 1:
            vif = None
        else:
            beta = np.linalg.pinv(others.T @ others) @ others.T @ y
            predicted = others @ beta
            ss_total = float(np.sum((y - np.mean(y)) ** 2))
            if ss_total <= 0:
                vif = None
            else:
                r2 = max(0.0, min(0.999999, float(1 - np.sum((y - predicted) ** 2) / ss_total)))
                vif = 1.0 / (1.0 - r2)
        rows.append({"Predictor": names[index], "VIF": vif})
    return rows


def simple_scatter_svg(figure_id: str, x_values: list[float], y_values: list[float], title: str, x_label: str, y_label: str, *, diagonal: bool = False) -> str:
    if len(x_values) != len(y_values) or not x_values:
        return f'<div class="chart-box"><h3>{h(title)}</h3><p class="small">No points available.</p></div>'
    width, height = 760, 330
    left, top, inner_w, inner_h = 62, 30, 640, 230
    xmin, xmax = min(x_values), max(x_values)
    ymin, ymax = min(y_values), max(y_values)
    if xmin == xmax:
        xmin -= 0.5
        xmax += 0.5
    if ymin == ymax:
        ymin -= 0.5
        ymax += 0.5
    pad_x = (xmax - xmin) * 0.05
    pad_y = (ymax - ymin) * 0.08
    xmin -= pad_x
    xmax += pad_x
    ymin -= pad_y
    ymax += pad_y

    def sx(xv: float) -> float:
        return left + ((xv - xmin) / max(0.000001, xmax - xmin)) * inner_w

    def sy(yv: float) -> float:
        return top + inner_h - ((yv - ymin) / max(0.000001, ymax - ymin)) * inner_h

    points = ''.join(
        f'<circle cx="{sx(x):.2f}" cy="{sy(y):.2f}" r="3" fill="#1d4ed8" opacity="0.70"><title>{h(x_label)}={x:.3f}, {h(y_label)}={y:.3f}</title></circle>'
        for x, y in zip(x_values, y_values)
    )
    diag = ''
    if diagonal:
        lo = max(xmin, ymin)
        hi = min(xmax, ymax)
        diag = f'<line x1="{sx(lo):.2f}" y1="{sy(lo):.2f}" x2="{sx(hi):.2f}" y2="{sy(hi):.2f}" stroke="#92400e" stroke-width="2" stroke-dasharray="5 4"></line>'
    return (
        f'<div class="chart-box"><h3>{h(title)}</h3>'
        f'<svg class="standalone-figure" viewBox="0 0 {width} {height}" role="img">'
        f'<rect x="0" y="0" width="{width}" height="{height}" fill="white"></rect>'
        f'<line x1="{left}" y1="{top + inner_h}" x2="{left + inner_w}" y2="{top + inner_h}" stroke="#5f6c73"></line>'
        f'<line x1="{left}" y1="{top}" x2="{left}" y2="{top + inner_h}" stroke="#5f6c73"></line>'
        f'<text x="{left + inner_w/2}" y="{height - 26}" text-anchor="middle" font-size="12">{h(x_label)}</text>'
        f'<text x="16" y="{top + inner_h/2}" transform="rotate(-90 16 {top + inner_h/2})" text-anchor="middle" font-size="12">{h(y_label)}</text>'
        f'{diag}{points}</svg></div>'
    )


def diagnostic_plots_svg(figure_id: str, fitted: list[float], residuals: list[float], title: str) -> str:
    residual_plot = simple_scatter_svg(figure_id + "-resid", fitted, residuals, title + ": residuals vs fitted", "Fitted values", "Residuals")
    if len(residuals) < 3:
        return residual_plot
    sorted_resid = sorted(float(x) for x in residuals)
    normal = statistics.NormalDist()
    theoretical = [normal.inv_cdf((i - 0.5) / len(sorted_resid)) for i in range(1, len(sorted_resid) + 1)]
    qq_plot = simple_scatter_svg(figure_id + "-qq", theoretical, sorted_resid, title + ": Q-Q plot", "Theoretical normal quantiles", "Observed residual quantiles", diagonal=True)
    return residual_plot + qq_plot


def diagnostic_tests_table(tests: list[dict[str, Any]]) -> str:
    if not tests:
        return '<p class="small">No diagnostic tests available.</p>'
    body = "".join(
        f'<tr><td>{h(row["Assumption"])}</td><td>{h(row["How to read"])}</td><td>{h(row["Diagnostic"])}</td><td>{row["Flag"]}</td></tr>'
        for row in tests
    )
    return '<div class="table-wrap"><table><thead><tr><th>Assumption</th><th>How to read it</th><th>Diagnostic output</th><th>Flag</th></tr></thead><tbody>' + body + '</tbody></table></div>'


def fit_diagnostic_residuals(rows: list[dict[str, Any]], outcome: str, numeric_predictors: list[str], categorical_predictors: list[str]) -> dict[str, Any]:
    complete, excluded = make_complete_cases(rows, [outcome, *numeric_predictors, *categorical_predictors], f"diagnostic {outcome}")
    if len(complete) < 3:
        return {"n": len(complete), "excluded": excluded, "status": "Too few complete cases for diagnostics.", "tests": [], "plots": "", "notes": []}
    x_matrix, names, notes = design_matrix_for_assumptions(complete, numeric_predictors, categorical_predictors)
    if x_matrix is None or x_matrix.shape[0] <= x_matrix.shape[1]:
        return {"n": len(complete), "excluded": excluded, "status": "Too few complete cases relative to diagnostic parameters.", "tests": [], "plots": "", "notes": notes}
    y_values = np.array([float(parse_float(row.get(outcome))) for row in complete], dtype=float)
    beta = np.linalg.pinv(x_matrix.T @ x_matrix) @ x_matrix.T @ y_values
    fitted = x_matrix @ beta
    residuals = y_values - fitted
    n, p = x_matrix.shape
    residual_sd = float(np.std(residuals, ddof=max(1, min(p, n - 1)))) if n > 1 else 0.0
    standardized = residuals / residual_sd if residual_sd > 0 else residuals * 0
    xtx_inv = np.linalg.pinv(x_matrix.T @ x_matrix)
    leverage = np.sum((x_matrix @ xtx_inv) * x_matrix, axis=1)
    cooks = (standardized ** 2 / max(1, p)) * (leverage / np.clip(1 - leverage, 1e-8, None))

    shapiro_p = None
    if scipy_stats is not None and 3 <= n <= 5000:
        try:
            shapiro_p = float(scipy_stats.shapiro(residuals).pvalue)
        except Exception:
            shapiro_p = None
    bp_p = None
    if scipy_stats is not None and p > 1 and n > p + 1:
        z = np.column_stack([np.ones(n), x_matrix[:, 1:]])
        sq = residuals ** 2
        fitted_sq = z @ (np.linalg.pinv(z.T @ z) @ z.T @ sq)
        ss_total = float(np.sum((sq - np.mean(sq)) ** 2))
        if ss_total > 0:
            r2 = float(1 - np.sum((sq - fitted_sq) ** 2) / ss_total)
            lm_stat = max(0.0, n * r2)
            bp_p = float(scipy_stats.chi2.sf(lm_stat, z.shape[1] - 1))
    dw = None
    if n > 1 and float(np.sum(residuals ** 2)) > 0:
        dw = float(np.sum(np.diff(residuals) ** 2) / np.sum(residuals ** 2))

    severe_resid = int(np.sum(np.abs(standardized) > 3))
    high_leverage = int(np.sum(leverage > (2 * p / n))) if n else 0
    high_cook = int(np.sum(cooks > 1))
    vif_rows = vif_table_from_matrix(x_matrix, names)
    max_vif = None
    if vif_rows and any(row["VIF"] is not None for row in vif_rows):
        max_vif = max(row["VIF"] for row in vif_rows if row["VIF"] is not None)
    tests = [
        {"Assumption": "Residual normality", "How to read": "Prefer the Q-Q plot. Shapiro p ≥ .05 means no strong evidence against normal residuals.", "Diagnostic": f"Shapiro p = {p_text(shapiro_p)}", "Flag": interpretation_badge(None if shapiro_p is None else shapiro_p >= .05, "no clear issue", "inspect/possible issue")},
        {"Assumption": "Homoscedasticity", "How to read": "Residual-vs-fitted plot should look like a random cloud. Breusch-Pagan p ≥ .05 means no strong evidence of unequal residual variance.", "Diagnostic": f"Breusch-Pagan p = {p_text(bp_p)}", "Flag": interpretation_badge(None if bp_p is None else bp_p >= .05, "no clear issue", "inspect/possible issue")},
        {"Assumption": "Independence", "How to read": "Independence is mainly a design question. Durbin-Watson near 2 is ideal; values between about 1 and 3 are usually acceptable.", "Diagnostic": "Durbin-Watson = " + ("—" if dw is None else f"{dw:.2f}"), "Flag": interpretation_badge(None if dw is None else 1 <= dw <= 3, "no clear issue", "inspect/possible issue")},
        {"Assumption": "Outliers / influence", "How to read": "Inspect |standardized residual| > 3, high leverage, and Cook's D > 1 before deciding whether a point is a data error.", "Diagnostic": f"|std. residual|>3: {severe_resid}; high leverage: {high_leverage}; Cook's D>1: {high_cook}", "Flag": interpretation_badge(severe_resid == 0 and high_cook == 0, "no severe cases", "inspect cases")},
    ]
    if max_vif is not None:
        tests.append({"Assumption": "Multicollinearity", "How to read": "VIF values around 1 are ideal; >5 is a warning; >10 is serious.", "Diagnostic": f"Max VIF = {max_vif:.2f}", "Flag": interpretation_badge(max_vif <= 5, "no clear issue", "inspect/possible issue")})
    plots = diagnostic_plots_svg(f"diag-{slugify(outcome)}-{n}-{len(numeric_predictors)}-{len(categorical_predictors)}", fitted.tolist(), residuals.tolist(), f"Diagnostics for {outcome}")
    return {"n": n, "excluded": excluded, "status": "Diagnostic residuals calculated. No final model coefficients were reported.", "tests": tests, "plots": plots, "notes": notes, "vif_rows": vif_rows}


def diagnostics_block(title: str, diagnostic: dict[str, Any]) -> str:
    notes = diagnostic.get("notes", [])
    note_html = list_html(notes) if notes else '<p class="small">No coding notes.</p>'
    vif_html = ""
    vif_rows = diagnostic.get("vif_rows", [])
    if vif_rows:
        body_rows: list[str] = []
        for row in vif_rows:
            vif_value = row.get("VIF")
            vif_display = "—" if vif_value is None else f"{vif_value:.2f}"
            body_rows.append(f'<tr><td>{h(row.get("Predictor"))}</td><td>{vif_display}</td></tr>')
        vif_html = '<details class="compact-details"><summary>Show VIF details</summary><div class="table-wrap"><table><thead><tr><th>Predictor</th><th>VIF</th></tr></thead><tbody>' + ''.join(body_rows) + '</tbody></table></div></details>'
    return (
        f'<details class="compact-details" open><summary>{h(title)} · n={h(diagnostic.get("n"))}</summary>'
        f'<p class="small">{h(diagnostic.get("status"))}</p>'
        f'{diagnostic_tests_table(diagnostic.get("tests", []))}'
        f'{vif_html}'
        f'<h4>Diagnostic coding notes</h4>{note_html}'
        f'{diagnostic.get("plots", "")}'
        '</details>'
    )


def assumption_section_shell(title: str, hypothesis: str, requirements: list[str], assumptions: list[dict[str, str]], data_html: str, diagnostics_html: str, final_models_html: str, status_html: str, details_html: str) -> str:
    assumption_rows = ''.join(
        f'<tr><td>{h(row["Assumption"])}</td><td>{h(row["How to test"])}</td><td>{h(row["How to read"])}</td></tr>'
        for row in assumptions
    )
    return (
        f'<section id="{h(slugify(title))}" class="card body-output hypothesis-section">'
        f'<h2>{h(title)} <span>Assumptions + final models</span></h2>'
        f'<p>{h(hypothesis)}</p>'
        '<h3>Requirements</h3>' + list_html(requirements) +
        '<h3>Assumptions: tests and how to read them</h3>' +
        '<div class="table-wrap"><table><thead><tr><th>Assumption</th><th>How to test</th><th>How to read</th></tr></thead><tbody>' + assumption_rows + '</tbody></table></div>' +
        '<h3>Complete-case data prepared for assumption checks</h3>' + data_html +
        '<h3>Assumption-check output</h3>' + diagnostics_html +
        '<h3>Final statistical models</h3>' + final_models_html +
        status_html + details_html +
        '</section>'
    )


def combined_exclusion_details(title: str, exclusions: list[dict[str, str]]) -> str:
    rows = ''.join(f'<tr><td>{h(row.get("MCID"))}</td><td>{h(row.get("condition"))}</td><td>{h(row.get("reason"))}</td></tr>' for row in exclusions)
    if not rows:
        rows = '<tr><td colspan="3">No complete-case exclusions for these diagnostics.</td></tr>'
    return f'<details class="compact-details"><summary>{h(title)}</summary><div class="table-wrap"><table><thead><tr><th>MCID</th><th>Condition</th><th>Reason</th></tr></thead><tbody>{rows}</tbody></table></div></details>'


# -----------------------------------------------------------------------------
# FINAL INFERENTIAL MODEL HELPERS
# -----------------------------------------------------------------------------


def fmt(value: float | None, digits: int = 3) -> str:
    if value is None or not math.isfinite(value):
        return "—"
    if abs(value) < 0.0005:
        return "0.000"
    return f"{value:.{digits}f}"


def ci_text(low: float | None, high: float | None, digits: int = 3) -> str:
    if low is None or high is None or not math.isfinite(low) or not math.isfinite(high):
        return "—"
    return f"[{fmt(low, digits)}, {fmt(high, digits)}]"


def html_note(text: str) -> str:
    return f'<p class="small">{h(text)}</p>'


def model_status_note(text: str, level: str = "green") -> str:
    return f'<div class="status status-{h(level)}"><strong>Model note.</strong> {h(text)}</div>'


def fit_ols_arrays(rows: list[dict[str, Any]], outcome: str, numeric_predictors: list[str], categorical_predictors: list[str]) -> dict[str, Any]:
    required = [outcome, *numeric_predictors, *categorical_predictors]
    complete, excluded = make_complete_cases(rows, required, f"final OLS {outcome}")
    if len(complete) < 3:
        return {"ok": False, "n": len(complete), "excluded": excluded, "error": "Too few complete cases.", "rows": complete}
    x_matrix, names, notes = design_matrix_for_assumptions(complete, numeric_predictors, categorical_predictors)
    if x_matrix is None:
        return {"ok": False, "n": len(complete), "excluded": excluded, "error": "Could not build design matrix.", "notes": notes, "rows": complete}
    n, p = x_matrix.shape
    if n <= p:
        return {"ok": False, "n": n, "excluded": excluded, "error": f"Too few complete cases relative to parameters (n={n}, parameters={p}).", "notes": notes, "rows": complete}
    y_values = np.array([float(parse_float(row.get(outcome))) for row in complete], dtype=float)
    xtx_inv = np.linalg.pinv(x_matrix.T @ x_matrix)
    beta = xtx_inv @ x_matrix.T @ y_values
    fitted = x_matrix @ beta
    residuals = y_values - fitted
    df_resid = max(1, n - p)
    sse = float(np.sum(residuals ** 2))
    sst = float(np.sum((y_values - np.mean(y_values)) ** 2))
    mse = sse / df_resid
    classical_cov = xtx_inv * mse
    leverage = np.sum((x_matrix @ xtx_inv) * x_matrix, axis=1)
    hc3_scale = (residuals / np.clip(1 - leverage, 1e-8, None)) ** 2
    robust_cov = xtx_inv @ (x_matrix.T @ (x_matrix * hc3_scale[:, None])) @ xtx_inv
    se = np.sqrt(np.maximum(np.diag(robust_cov), 0))
    tcrit = scipy_stats.t.ppf(0.975, df_resid) if scipy_stats is not None else 1.96
    coeff_rows: list[dict[str, Any]] = []
    for idx, name in enumerate(names):
        t_value = float(beta[idx] / se[idx]) if se[idx] > 0 else None
        p_value = float(2 * scipy_stats.t.sf(abs(t_value), df_resid)) if scipy_stats is not None and t_value is not None else None
        coeff_rows.append({
            "Term": name,
            "b": float(beta[idx]),
            "SE_HC3": float(se[idx]),
            "t": t_value,
            "p": p_value,
            "CI_low": float(beta[idx] - tcrit * se[idx]),
            "CI_high": float(beta[idx] + tcrit * se[idx]),
        })
    r2 = float(1 - sse / sst) if sst > 0 else None
    adj_r2 = float(1 - (1 - r2) * (n - 1) / df_resid) if r2 is not None and df_resid > 0 else None
    f_stat = None
    f_p = None
    if p > 1 and sst > 0:
        ssr = sst - sse
        f_stat = float((ssr / (p - 1)) / mse) if mse > 0 else None
        f_p = float(scipy_stats.f.sf(f_stat, p - 1, df_resid)) if scipy_stats is not None and f_stat is not None else None
    return {
        "ok": True,
        "n": n,
        "p": p,
        "df_resid": df_resid,
        "names": names,
        "beta": beta,
        "cov": robust_cov,
        "coeff_rows": coeff_rows,
        "r2": r2,
        "adj_r2": adj_r2,
        "f_stat": f_stat,
        "f_p": f_p,
        "excluded": excluded,
        "notes": notes,
        "rows": complete,
        "outcome": outcome,
    }


def coefficient_table_html(coeff_rows: list[dict[str, Any]], *, ci_label: str = "95% CI") -> str:
    if not coeff_rows:
        return '<p class="small">No coefficient rows available.</p>'
    body: list[str] = []
    for row in coeff_rows:
        p_value = row.get("p")
        tr_class = ' class="significant-row"' if p_value is not None and math.isfinite(p_value) and p_value < ALPHA else ""
        body.append(
            f'<tr{tr_class}><td>{h(row.get("Term"))}</td><td>{fmt(row.get("b"))}</td><td>{fmt(row.get("SE_HC3", row.get("SE")))}</td>'
            f'<td>{fmt(row.get("t", row.get("z")))}</td><td>{p_text(row.get("p"))}</td><td>{ci_text(row.get("CI_low"), row.get("CI_high"))}</td></tr>'
        )
    return '<div class="table-wrap"><table class="model-table"><thead><tr><th>Term</th><th>b</th><th>SE</th><th>t/z</th><th>p</th><th>' + h(ci_label) + '</th></tr></thead><tbody>' + ''.join(body) + '</tbody></table></div>'


def scaled_effect_from_ols(model: dict[str, Any], term: str, multiplier: float) -> dict[str, Any] | None:
    if not model.get("ok") or term not in model.get("names", []):
        return None
    idx = model["names"].index(term)
    beta = model["beta"]
    cov = model["cov"]
    estimate = float(beta[idx] * multiplier)
    se = float(math.sqrt(max(0.0, cov[idx, idx])) * abs(multiplier))
    df = model.get("df_resid", 1)
    t_value = estimate / se if se > 0 else None
    p_value = float(2 * scipy_stats.t.sf(abs(t_value), df)) if scipy_stats is not None and t_value is not None else None
    tcrit = scipy_stats.t.ppf(0.975, df) if scipy_stats is not None else 1.96
    return {"estimate": estimate, "se": se, "stat": t_value, "p": p_value, "low": estimate - tcrit * se, "high": estimate + tcrit * se}


def holm_adjust(p_values: list[float | None]) -> list[float | None]:
    """Return Holm-adjusted p-values in the original order."""
    indexed = [(index, p) for index, p in enumerate(p_values) if p is not None and math.isfinite(p)]
    adjusted: list[float | None] = [None for _ in p_values]
    if not indexed:
        return adjusted
    indexed.sort(key=lambda item: item[1])
    m = len(indexed)
    running_max = 0.0
    for rank, (original_index, p_value) in enumerate(indexed, start=1):
        raw_adjusted = min(1.0, (m - rank + 1) * p_value)
        running_max = max(running_max, raw_adjusted)
        adjusted[original_index] = running_max
    return adjusted


def categorical_if_available(rows: list[dict[str, Any]], variable: str, outcome: str | None = None) -> list[str]:
    """Include a categorical predictor only when at least two levels are observable."""
    scoped = rows
    if outcome is not None:
        scoped = [row for row in rows if parse_float(row.get(outcome)) is not None]
    levels = {clean(row.get(variable)) for row in scoped if clean(row.get(variable))}
    return [variable] if len(levels) >= 2 else []


def h1_confirmatory_models_html(
    rows: list[dict[str, Any]],
    *,
    heading: str = "Confirmatory preregistered H1 models: separate HC3 linear models",
    note: str = "These are placed before the integrated retention/decay LMM because the preregistration specified separate immediate and delayed retention models with HC3 standard errors. The LMM below is an added sensitivity/extension, not a replacement for this table.",
    model_label_prefix: str = "Confirmatory H1",
) -> str:
    """Preregistered H1 route: separate HC3 linear models for immediate and delayed retention."""
    immediate_form = categorical_if_available(rows, "retention_form_order", "ret_immediate_score")
    delayed_form = categorical_if_available(rows, "retention_form_order", "ret_delayed_score")
    immediate_note = " + retention_form_order" if immediate_form else " (retention_form_order unavailable/constant; omitted)"
    delayed_note = " + retention_form_order" if delayed_form else " (retention_form_order unavailable/constant; omitted)"
    return (
        f"<h4>{h(heading)}</h4>"
        f'<p class="small">{h(note)}</p>'
        + ols_model_html(
            f"{model_label_prefix} base model: immediate retention",
            "immediate_retention ~ C1 + C2" + immediate_note,
            fit_ols_arrays(rows, "ret_immediate_score", ["required_pause_contrast", "optional_pause_contrast"], immediate_form),
        )
        + ols_model_html(
            f"{model_label_prefix} base model: delayed retention",
            "delayed_retention ~ C1 + C2" + delayed_note,
            fit_ols_arrays(rows, "ret_delayed_score", ["required_pause_contrast", "optional_pause_contrast"], delayed_form),
        )
        + f"<h4>{h(model_label_prefix)} covariate-adjusted sensitivity models</h4>"
        + ols_model_html(
            f"{model_label_prefix} sensitivity model: immediate retention + covariates",
            "immediate_retention ~ C1 + C2 + retention_form_order + location + co_present_participants + age + gender",
            fit_ols_arrays(rows, "ret_immediate_score", ["required_pause_contrast", "optional_pause_contrast", "co_present_participants", "age"], [*immediate_form, "location", "gender"]),
        )
        + ols_model_html(
            f"{model_label_prefix} sensitivity model: delayed retention + covariates",
            "delayed_retention ~ C1 + C2 + retention_form_order + location + co_present_participants + age + gender",
            fit_ols_arrays(rows, "ret_delayed_score", ["required_pause_contrast", "optional_pause_contrast", "co_present_participants", "age"], [*delayed_form, "location", "gender"]),
        )
    )


def h1_retention_score_variation_models_html() -> str:
    """Exploratory H1-only reruns for alternative participant-level retention score files."""
    blocks: list[str] = [
        '<h4>Exploratory H1 reruns with alternative retention-score calculations</h4>'
        '<p class="small">These models are exploratory sensitivity checks. They use modified participant-level retention-score files and do not replace the main preregistered retention outcome. Only the H1 immediate and delayed retention models are rerun here.</p>'
    ]

    for mode, path in RETENTION_FINAL_SCORE_VARIATION_PATHS.items():
        if mode == "clean":
            continue

        source = retention_score_file_label(path)

        if not path.exists():
            blocks.append(
                f'<details class="compact-details"><summary>{h(mode)}</summary>'
                + model_status_note(f"Skipped: {source} does not exist yet.", "orange")
                + '</details>'
            )
            continue

        variant_rows, variant_warnings = build_rows_for_inferential_models(retention_scores_path=path)
        variant_errors: list[str] = [] if variant_rows else [f"No participant-level rows could be built from {source}."]

        blocks.append(
            f'<details class="compact-details" open><summary>{h(mode)} · source: {h(source)}</summary>'
            + status_messages(
                variant_errors,
                variant_warnings,
                f"Exploratory H1 rows were built from {source}.",
            )
            + h1_confirmatory_models_html(
                variant_rows,
                heading=f"Exploratory H1 models: {mode}",
                note=f"Participant-level retention scores are read from {source}. This is an exploratory modified-score analysis only.",
                model_label_prefix=f"Exploratory H1 ({mode})",
            )
            + '</details>'
        )

    return "".join(blocks)


def h2a_holm_table_html(rows: list[dict[str, Any]], *, include_covariates: bool) -> str:
    """Direct-effect cognitive-load family for H2a with raw and Holm-adjusted p-values."""
    outcomes = [
        ("cl_intrinsic", "Intrinsic cognitive load", "secondary / no directional prediction"),
        ("cl_extraneous", "Extraneous cognitive load", "focal: required pauses expected lower than continue"),
        ("cl_germane", "Germane cognitive load", "focal: required pauses expected higher than continue"),
    ]
    cov_num = ["co_present_participants", "age"] if include_covariates else []
    cov_cat = ["location", "gender"] if include_covariates else []
    raw_rows: list[dict[str, Any]] = []
    exclusions: list[dict[str, str]] = []
    for outcome, label, role in outcomes:
        model = fit_ols_arrays(rows, outcome, ["required_pause_contrast", "optional_pause_contrast", *cov_num], cov_cat)
        exclusions.extend(model.get("excluded", []))
        for contrast, contrast_label in CONTRAST_DISPLAY.items():
            effect = scaled_effect_from_ols(model, contrast, CONTRAST_MULTIPLIERS[contrast]) if model.get("ok") else None
            raw_rows.append({
                "outcome": label,
                "role": role,
                "contrast": contrast_label,
                "n": model.get("n"),
                "estimate": None if effect is None else effect["estimate"],
                "se": None if effect is None else effect["se"],
                "p": None if effect is None else effect["p"],
                "low": None if effect is None else effect["low"],
                "high": None if effect is None else effect["high"],
            })
    adjusted = holm_adjust([row["p"] for row in raw_rows])
    body = []
    for row, p_holm in zip(raw_rows, adjusted):
        cls = ' class="significant-row"' if p_holm is not None and p_holm < ALPHA else ""
        body.append(
            f'<tr{cls}><td>{h(row["outcome"])}</td><td>{h(row["role"])}</td><td>{h(row["contrast"])}</td>'
            f'<td>{h(row["n"])}</td><td>{fmt(row["estimate"])}</td><td>{fmt(row["se"])}</td>'
            f'<td>{p_text(row["p"])}</td><td>{p_text(p_holm)}</td><td>{ci_text(row["low"], row["high"])}</td></tr>'
        )
    title = "H2a direct cognitive-load models with Holm correction" + (" · covariate-adjusted sensitivity" if include_covariates else " · base")
    return (
        f'<details class="compact-details" open><summary>{h(title)}</summary>'
        '<p class="small">Family-wise correction is applied across the displayed H2a cognitive-load contrast tests. Interpret the focal required-pause rows for extraneous and germane load first.</p>'
        '<div class="table-wrap"><table class="model-table"><thead><tr><th>Outcome</th><th>Role</th><th>Contrast</th><th>n</th><th>Estimate</th><th>SE</th><th>raw p</th><th>Holm p</th><th>95% CI</th></tr></thead><tbody>'
        + ''.join(body) + '</tbody></table></div>'
        + combined_exclusion_details("Show H2a model exclusions", exclusions)
        + '</details>'
    )



def model_fit_summary_cells(model: dict[str, Any]) -> str:
    """Compact OLS fit cells for report summary tables."""
    if not model.get("ok"):
        return f'<td>{h(model.get("n", "—"))}</td><td colspan="4">{h(model.get("error", "Model could not be fitted."))}</td>'
    return (
        f'<td>{h(model.get("n"))}</td>'
        f'<td>{fmt(model.get("r2"))}</td>'
        f'<td>{fmt(model.get("adj_r2"))}</td>'
        f'<td>{fmt(model.get("f_stat"))}</td>'
        f'<td>{p_text(model.get("f_p"))}</td>'
    )


def h3a_path_table_html(rows: list[dict[str, Any]], *, include_covariates: bool) -> str:
    """Report H3a explicitly: checkpoint design -> engagement.

    H3a is estimated once as the mediator model. Therefore the C1 a-path is shared by
    model labels [1] and [3], and the C2 a-path is shared by [2] and [4].
    """
    cov_num = ["co_present_participants", "age"] if include_covariates else []
    cov_cat = ["location", "gender"] if include_covariates else []
    model = fit_ols_arrays(rows, "engagement", ["required_pause_contrast", "optional_pause_contrast", *cov_num], cov_cat)
    title = "Final H3a model: checkpoint design → engagement" + (" · covariate-adjusted sensitivity" if include_covariates else " · base")
    if not model.get("ok"):
        return f'<details class="compact-details" open><summary>{h(title)}</summary>{model_status_note(model.get("error", "Model could not be fitted."), "red")}{combined_exclusion_details("Show H3a model exclusions", model.get("excluded", []))}</details>'

    hypothesis_labels = {
        "required_pause_contrast": "[1] / [3]",
        "optional_pause_contrast": "[2] / [4]",
    }
    body: list[str] = []
    for contrast in ["required_pause_contrast", "optional_pause_contrast"]:
        effect = scaled_effect_from_ols(model, contrast, CONTRAST_MULTIPLIERS[contrast])
        if effect is None:
            body.append(f'<tr><td>{h(hypothesis_labels[contrast])}</td><td>{h(CONTRAST_DISPLAY[contrast])}</td><td colspan="5">Unavailable; contrast term missing.</td></tr>')
            continue
        cls = ' class="significant-row"' if effect.get("p") is not None and effect["p"] < ALPHA else ""
        body.append(
            f'<tr{cls}><td>{h(hypothesis_labels[contrast])}</td><td>{h(CONTRAST_DISPLAY[contrast])}</td>'
            f'<td>{fmt(effect["estimate"])}</td><td>{fmt(effect["se"])}</td><td>{fmt(effect["stat"])}</td>'
            f'<td>{p_text(effect["p"])}</td><td>{ci_text(effect["low"], effect["high"])}</td></tr>'
        )

    cov_note = " + covariates" if include_covariates else ""
    return (
        f'<details class="compact-details" open><summary>{h(title)} · n={model.get("n")}</summary>'
        f'<p><strong>Formula:</strong> <code>engagement ~ C1 + C2{h(cov_note)}</code></p>'
        '<p class="small">H3a is the mediator model. The C1 a-path is reported as [1] / [3] because it is the same design → engagement path for the immediate- and delayed-retention mediation labels; the C2 a-path is reported as [2] / [4] for the same reason. Estimates are rescaled to the planned contrast scale.</p>'
        '<div class="table-wrap"><table class="model-table"><thead><tr><th>Model label(s)</th><th>Contrast</th><th>b</th><th>SE</th><th>t</th><th>p</th><th>95% robust CI</th></tr></thead><tbody>'
        + ''.join(body) + '</tbody></table></div>'
        + combined_exclusion_details("Show H3a model exclusions", model.get("excluded", []))
        + '</details>'
    )


def h3_component_model_fit_html(rows: list[dict[str, Any]], *, include_covariates: bool) -> str:
    """Summarise regression model fit for the H3 component equations."""
    cov_num = ["co_present_participants", "age"] if include_covariates else []
    cov_cat = ["location", "gender"] if include_covariates else []
    cov_formula = " + covariates" if include_covariates else ""
    title = "H3 component model fit" + (" · covariate-adjusted sensitivity" if include_covariates else " · base")
    component_models = [
        (
            "Mediator model for H3a ([1]/[3] and [2]/[4])",
            f"engagement ~ C1 + C2{cov_formula}",
            fit_ols_arrays(rows, "engagement", ["required_pause_contrast", "optional_pause_contrast", *cov_num], cov_cat),
        ),
        (
            "Immediate outcome model for direct/H3b paths ([1] and [2])",
            f"immediate retention ~ C1 + C2 + engagement{cov_formula}",
            fit_ols_arrays(rows, "ret_immediate_score", ["required_pause_contrast", "optional_pause_contrast", "engagement", *cov_num], cov_cat),
        ),
        (
            "Immediate total-effect model ([1] and [2])",
            f"immediate retention ~ C1 + C2{cov_formula}",
            fit_ols_arrays(rows, "ret_immediate_score", ["required_pause_contrast", "optional_pause_contrast", *cov_num], cov_cat),
        ),
        (
            "Delayed outcome model for direct/H3b paths ([3] and [4])",
            f"delayed retention ~ C1 + C2 + engagement{cov_formula}",
            fit_ols_arrays(rows, "ret_delayed_score", ["required_pause_contrast", "optional_pause_contrast", "engagement", *cov_num], cov_cat),
        ),
        (
            "Delayed total-effect model ([3] and [4])",
            f"delayed retention ~ C1 + C2{cov_formula}",
            fit_ols_arrays(rows, "ret_delayed_score", ["required_pause_contrast", "optional_pause_contrast", *cov_num], cov_cat),
        ),
    ]
    body = ''.join(
        f'<tr><td>{h(label)}</td><td><code>{h(formula)}</code></td>{model_fit_summary_cells(model)}</tr>'
        for label, formula, model in component_models
    )
    exclusions: list[dict[str, str]] = []
    for _label, _formula, model in component_models:
        exclusions.extend(model.get("excluded", []))
    return (
        f'<details class="compact-details" open><summary>{h(title)}</summary>'
        '<p class="small">These are regression-fit summaries for the component equations used to read H3a, H3b, direct effects, and total effects. Standard errors in the coefficient tables are HC3 robust SEs; the R²/F summaries are the corresponding OLS fit summaries.</p>'
        '<div class="table-wrap"><table class="model-table"><thead><tr><th>Component</th><th>Formula</th><th>n</th><th>R²</th><th>Adjusted R²</th><th>F</th><th>p</th></tr></thead><tbody>'
        + body + '</tbody></table></div>'
        + combined_exclusion_details("Show H3 component model-fit exclusions", exclusions)
        + '</details>'
    )
def ols_model_html(title: str, formula_label: str, model: dict[str, Any], *, include_planned_contrasts: bool = True) -> str:
    if not model.get("ok"):
        return f'<details class="compact-details" open><summary>{h(title)}</summary>{model_status_note(model.get("error", "Model could not be fitted."), "red")}</details>'
    r2_text = fmt(model.get("r2"))
    adj_r2_text = fmt(model.get("adj_r2"))
    f_text = fmt(model.get("f_stat"))
    fp_text = p_text(model.get("f_p"))
    notes = ''.join(f'<li>{h(note)}</li>' for note in model.get("notes", [])) or '<li>No special coding notes.</li>'
    contrast_html = ""
    if include_planned_contrasts:
        rows: list[str] = []
        for term, label in CONTRAST_DISPLAY.items():
            effect = scaled_effect_from_ols(model, term, CONTRAST_MULTIPLIERS[term])
            if effect is None:
                continue
            cls = ' class="significant-row"' if effect.get("p") is not None and effect["p"] < ALPHA else ""
            rows.append(f'<tr{cls}><td>{h(label)}</td><td>{fmt(effect["estimate"])}</td><td>{fmt(effect["se"])}</td><td>{fmt(effect["stat"])}</td><td>{p_text(effect["p"])}</td><td>{ci_text(effect["low"], effect["high"])}</td></tr>')
        if rows:
            contrast_html = '<h4>Planned contrast estimates on the outcome scale</h4><p class="small">The regression coefficients are coded-unit coefficients. This table rescales C1 by 2 and C2 by 1.5 so the estimates correspond to the planned group comparisons.</p><div class="table-wrap"><table class="model-table"><thead><tr><th>Contrast</th><th>Estimate</th><th>SE</th><th>t</th><th>p</th><th>95% CI</th></tr></thead><tbody>' + ''.join(rows) + '</tbody></table></div>'
    return (
        f'<details class="compact-details" open><summary>{h(title)} · n={model.get("n")}</summary>'
        f'<p><strong>Formula:</strong> <code>{h(formula_label)}</code></p>'
        f'<p><strong>Model fit:</strong> R² = {r2_text}, adjusted R² = {adj_r2_text}, F = {f_text}, p = {fp_text}. Standard errors are HC3 robust SEs.</p>'
        + coefficient_table_html(model.get("coeff_rows", []), ci_label="95% robust CI")
        + contrast_html
        + '<h4>Final-model coding notes</h4><ul>' + notes + '</ul>'
        + '</details>'
    )


def final_models_wrapper(*blocks: str) -> str:
    return '<div class="final-models">' + ''.join(blocks) + '</div>'


def rows_to_dataframe(rows: list[dict[str, Any]]):
    if pd is None:
        return None
    return pd.DataFrame(rows)


def fit_mixedlm_html(title: str, rows: list[dict[str, Any]], formula: str) -> str:
    if pd is None or smf is None:
        return f'<details class="compact-details" open><summary>{h(title)}</summary>{model_status_note("pandas/statsmodels are unavailable, so the Linear Mixed-Effects Model could not be fitted in this environment.", "red")}</details>'
    if not rows:
        return f'<details class="compact-details" open><summary>{h(title)}</summary>{model_status_note("No complete long-format rows available for this model.", "red")}</details>'
    frame = pd.DataFrame(rows).copy()
    try:
        for column in ["retention_score", "time", "required_pause_contrast", "optional_pause_contrast", "co_present_participants", "age"]:
            if column in frame.columns:
                frame[column] = pd.to_numeric(frame[column], errors="coerce")
        model = smf.mixedlm(formula, data=frame, groups=frame["MCID"])
        result = model.fit(reml=False, method="lbfgs", maxiter=500, disp=False)
    except Exception as first_error:
        try:
            result = model.fit(reml=False, method="powell", maxiter=1000, disp=False)
        except Exception as second_error:
            return f'<details class="compact-details" open><summary>{h(title)}</summary>{model_status_note("MixedLM failed to converge/fit: " + str(second_error), "red")}</details>'
    params = result.params
    conf = result.conf_int()
    coeff_rows: list[dict[str, Any]] = []
    for name in result.fe_params.index:
        p_value = float(result.pvalues.get(name, float("nan"))) if hasattr(result, "pvalues") else None
        coeff_rows.append({
            "Term": name,
            "b": float(result.fe_params[name]),
            "SE": float(result.bse_fe[name]),
            "z": float(result.tvalues[name]),
            "p": p_value if math.isfinite(p_value) else None,
            "CI_low": float(conf.loc[name, 0]),
            "CI_high": float(conf.loc[name, 1]),
        })
    effects = mixedlm_derived_effects_table(result)
    return (
        f'<details class="compact-details" open><summary>{h(title)} · observations={len(frame)}, participants={frame["MCID"].nunique()}</summary>'
        f'<p><strong>Formula:</strong> <code>{h(formula)} + random intercept for participant</code></p>'
        f'<p><strong>Fit:</strong> ML fit; log-likelihood = {fmt(float(result.llf))}; converged = {h(getattr(result, "converged", "unknown"))}.</p>'
        + coefficient_table_html(coeff_rows, ci_label="95% Wald CI")
        + effects
        + '</details>'
    )


def mixedlm_derived_effects_table(result: Any) -> str:
    names = list(result.fe_params.index)
    params = result.fe_params
    cov = result.cov_params()

    def derived(label: str, pieces: list[tuple[str, float]]) -> str:
        estimate = 0.0
        variance = 0.0
        missing = [name for name, _weight in pieces if name not in names]
        if missing:
            return f'<tr><td>{h(label)}</td><td colspan="5">Unavailable; missing term(s): {h(", ".join(missing))}</td></tr>'
        for name, weight in pieces:
            estimate += weight * float(params[name])
        for name1, weight1 in pieces:
            for name2, weight2 in pieces:
                variance += weight1 * weight2 * float(cov.loc[name1, name2])
        se = math.sqrt(max(0.0, variance))
        z_value = estimate / se if se > 0 else None
        p_value = float(2 * scipy_stats.norm.sf(abs(z_value))) if scipy_stats is not None and z_value is not None else None
        low = estimate - 1.96 * se
        high = estimate + 1.96 * se
        cls = ' class="significant-row"' if p_value is not None and p_value < ALPHA else ""
        return f'<tr{cls}><td>{h(label)}</td><td>{fmt(estimate)}</td><td>{fmt(se)}</td><td>{fmt(z_value)}</td><td>{p_text(p_value)}</td><td>{ci_text(low, high)}</td></tr>'

    rows = [
        derived("Immediate: required pauses − required continue", [("required_pause_contrast", 2.0)]),
        derived("Immediate: optional pauses − system-controlled average", [("optional_pause_contrast", 1.5)]),
        derived("Average/contrast-zero retention decay", [("time", 1.0)]),
        derived("Difference in decay: required pauses − required continue", [("time:required_pause_contrast", 2.0)]),
        derived("Difference in decay: optional pauses − system-controlled average", [("time:optional_pause_contrast", 1.5)]),
        derived("Delayed: required pauses − required continue", [("required_pause_contrast", 2.0), ("time:required_pause_contrast", 2.0)]),
        derived("Delayed: optional pauses − system-controlled average", [("optional_pause_contrast", 1.5), ("time:optional_pause_contrast", 1.5)]),
    ]
    return '<h4>Derived retention effects on the 0–2 retention scale</h4><p class="small">Immediate effects are the design contrasts at time = 0. Delayed effects combine each design contrast with its time interaction. Decay is delayed minus immediate because time is coded 0 = immediate and 1 = delayed.</p><div class="table-wrap"><table class="model-table"><thead><tr><th>Effect</th><th>Estimate</th><th>SE</th><th>z</th><th>p</th><th>95% CI</th></tr></thead><tbody>' + ''.join(rows) + '</tbody></table></div>'


def bootstrap_indices(n: int, iterations: int, seed: int) -> list[np.ndarray]:
    rng = np.random.default_rng(seed)
    return [rng.integers(0, n, size=n) for _ in range(iterations)]


def mediation_model_html(title: str, rows: list[dict[str, Any]], mediators: list[str], outcome: str, *, include_covariates: bool, iterations: int = 1000, seed: int = 20260622) -> str:
    cov_num = ["co_present_participants", "age"] if include_covariates else []
    cov_cat = ["location", "gender"] if include_covariates else []
    needed = ["required_pause_contrast", "optional_pause_contrast", *mediators, outcome, *cov_num, *cov_cat]
    complete, excluded = make_complete_cases(rows, needed, f"{title} mediation")
    if len(complete) < 10:
        return f'<details class="compact-details" open><summary>{h(title)}</summary>{model_status_note("Too few complete cases for mediation.", "red")}{combined_exclusion_details("Show mediation exclusions", excluded)}</details>'

    cov_note = " + covariates" if include_covariates else ""
    a_models = {mediator: fit_ols_arrays(complete, mediator, ["required_pause_contrast", "optional_pause_contrast", *cov_num], cov_cat) for mediator in mediators}
    y_model = fit_ols_arrays(complete, outcome, ["required_pause_contrast", "optional_pause_contrast", *mediators, *cov_num], cov_cat)
    total_model = fit_ols_arrays(complete, outcome, ["required_pause_contrast", "optional_pause_contrast", *cov_num], cov_cat)
    if not y_model.get("ok") or not total_model.get("ok") or any(not model.get("ok") for model in a_models.values()):
        return f'<details class="compact-details" open><summary>{h(title)}</summary>{model_status_note("One or more component regressions could not be fitted.", "red")}{combined_exclusion_details("Show mediation exclusions", excluded)}</details>'

    direct_rows: list[str] = []
    for contrast in ["required_pause_contrast", "optional_pause_contrast"]:
        multiplier = CONTRAST_MULTIPLIERS[contrast]
        label = CONTRAST_DISPLAY[contrast]
        total = scaled_effect_from_ols(total_model, contrast, multiplier)
        direct = scaled_effect_from_ols(y_model, contrast, multiplier)
        if total is not None and direct is not None:
            total_sig = total.get("p") is not None and math.isfinite(total["p"]) and total["p"] < ALPHA
            direct_sig = direct.get("p") is not None and math.isfinite(direct["p"]) and direct["p"] < ALPHA
            cls = ' class="significant-row"' if total_sig or direct_sig else ""
            direct_rows.append(
                f'<tr{cls}><td>{h(label)}</td>'
                f'<td>{fmt(total["estimate"])}</td><td>{fmt(total["se"])}</td><td>{fmt(total["stat"])}</td><td>{p_text(total["p"])}</td><td>{ci_text(total["low"], total["high"])}</td>'
                f'<td>{fmt(direct["estimate"])}</td><td>{fmt(direct["se"])}</td><td>{fmt(direct["stat"])}</td><td>{p_text(direct["p"])}</td><td>{ci_text(direct["low"], direct["high"])}</td></tr>'
            )

    point_indirect: dict[tuple[str, str], float] = {}
    for contrast in ["required_pause_contrast", "optional_pause_contrast"]:
        for mediator in mediators:
            a = a_models[mediator]["beta"][a_models[mediator]["names"].index(contrast)]
            b = y_model["beta"][y_model["names"].index(mediator)]
            point_indirect[(contrast, mediator)] = float(CONTRAST_MULTIPLIERS[contrast] * a * b)

    boot_values: dict[tuple[str, str], list[float]] = {key: [] for key in point_indirect}
    index_sets = bootstrap_indices(len(complete), iterations, seed)
    for indices in index_sets:
        sample = [complete[int(i)] for i in indices]
        try:
            sample_a = {mediator: fit_ols_arrays(sample, mediator, ["required_pause_contrast", "optional_pause_contrast", *cov_num], cov_cat) for mediator in mediators}
            sample_y = fit_ols_arrays(sample, outcome, ["required_pause_contrast", "optional_pause_contrast", *mediators, *cov_num], cov_cat)
            if not sample_y.get("ok") or any(not model.get("ok") for model in sample_a.values()):
                continue
            for contrast in ["required_pause_contrast", "optional_pause_contrast"]:
                for mediator in mediators:
                    a = sample_a[mediator]["beta"][sample_a[mediator]["names"].index(contrast)]
                    b = sample_y["beta"][sample_y["names"].index(mediator)]
                    boot_values[(contrast, mediator)].append(float(CONTRAST_MULTIPLIERS[contrast] * a * b))
        except Exception:
            continue

    indirect_rows: list[str] = []
    for contrast in ["required_pause_contrast", "optional_pause_contrast"]:
        for mediator in mediators:
            vals = boot_values[(contrast, mediator)]
            low = high = None
            if len(vals) >= 50:
                low, high = np.percentile(vals, [2.5, 97.5]).tolist()
            supported = low is not None and high is not None and (low > 0 or high < 0)
            cls = ' class="significant-row"' if supported else ""
            indirect_rows.append(f'<tr{cls}><td>{h(CONTRAST_DISPLAY[contrast])}</td><td>{h(mediator)}</td><td>{fmt(point_indirect[(contrast, mediator)])}</td><td>{ci_text(low, high)}</td><td>{len(vals)}</td></tr>')

    b_rows: list[str] = []
    for mediator in mediators:
        if mediator in y_model["names"]:
            idx = y_model["names"].index(mediator)
            row = y_model["coeff_rows"][idx]
            b_rows.append(f'<tr><td>{h(mediator)}</td><td>{fmt(row["b"])}</td><td>{fmt(row["SE_HC3"])}</td><td>{fmt(row["t"])}</td><td>{p_text(row["p"])}</td><td>{ci_text(row["CI_low"], row["CI_high"])}</td></tr>')

    return (
        f'<details class="compact-details" open><summary>{h(title)} · n={len(complete)} · bootstrap iterations requested={iterations}</summary>'
        f'<p><strong>Component models:</strong> mediator(s) ~ C1 + C2{h(cov_note)}; {h(outcome)} ~ C1 + C2 + mediator(s){h(cov_note)}. Total and direct effects are contrast-rescaled regression coefficients with HC3 robust SEs and CIs. Indirect effects are contrast-rescaled products a×b.</p>'
        '<h4>Total and direct planned contrast effects</h4><div class="table-wrap"><table class="model-table"><thead><tr><th>Contrast</th><th>Total b</th><th>Total SE</th><th>Total t</th><th>Total p</th><th>Total 95% robust CI</th><th>Direct b</th><th>Direct SE</th><th>Direct t</th><th>Direct p</th><th>Direct 95% robust CI</th></tr></thead><tbody>' + ''.join(direct_rows) + '</tbody></table></div>'
        '<h4>Indirect effects with percentile bootstrap CIs</h4><div class="table-wrap"><table class="model-table"><thead><tr><th>Contrast</th><th>Mediator</th><th>Indirect effect</th><th>95% bootstrap CI</th><th>Successful bootstraps</th></tr></thead><tbody>' + ''.join(indirect_rows) + '</tbody></table></div>'
        '<h4>b-paths from mediator(s) to outcome</h4><div class="table-wrap"><table class="model-table"><thead><tr><th>Mediator</th><th>b</th><th>SE</th><th>t</th><th>p</th><th>95% robust CI</th></tr></thead><tbody>' + ''.join(b_rows) + '</tbody></table></div>'
        + combined_exclusion_details("Show mediation exclusions", excluded)
        + '</details>'
    )


def correlation_model_html(rows: list[dict[str, Any]], x_var: str, y_vars: list[str]) -> str:
    body_rows: list[dict[str, Any]] = []
    exclusions: list[dict[str, str]] = []
    for y_var in y_vars:
        complete, excluded = make_complete_cases(rows, [x_var, y_var], f"correlation {x_var} with {y_var}")
        exclusions.extend(excluded)
        x_values = [float(parse_float(row.get(x_var))) for row in complete]
        y_values = [float(parse_float(row.get(y_var))) for row in complete]
        if len(complete) < 4 or scipy_stats is None:
            body_rows.append({"y_var": y_var, "n": len(complete), "r": None, "p": None, "low": None, "high": None, "r2": None, "rs": None, "ps": None, "available": False})
            continue
        r, p_value = scipy_stats.pearsonr(x_values, y_values)
        rs, ps = scipy_stats.spearmanr(x_values, y_values)
        z = math.atanh(max(-0.999999, min(0.999999, float(r))))
        se = 1 / math.sqrt(len(complete) - 3)
        low = math.tanh(z - 1.96 * se)
        high = math.tanh(z + 1.96 * se)
        body_rows.append({"y_var": y_var, "n": len(complete), "r": float(r), "p": float(p_value), "low": low, "high": high, "r2": float(r*r), "rs": float(rs), "ps": float(ps), "available": True})
    adjusted = holm_adjust([row["p"] for row in body_rows])
    body: list[str] = []
    for row, p_holm in zip(body_rows, adjusted):
        if not row.get("available"):
            body.append(f'<tr><td>{h(row["y_var"])}</td><td>{row["n"]}</td><td colspan="7">Not enough cases or SciPy unavailable.</td></tr>')
            continue
        cls = ' class="significant-row"' if p_holm is not None and p_holm < ALPHA else ""
        body.append(f'<tr{cls}><td>{h(row["y_var"])}</td><td>{row["n"]}</td><td>{fmt(row["r"])}</td><td>{p_text(row["p"])}</td><td>{p_text(p_holm)}</td><td>{ci_text(row["low"], row["high"])}</td><td>{fmt(row["r2"])}</td><td>{fmt(row["rs"])}</td><td>{p_text(row["ps"])}</td></tr>')
    return (
        '<details class="compact-details" open><summary>Final H4 correlations with Holm correction</summary>'
        '<p class="small">Holm correction is applied across the three Pearson H4 tests. Spearman correlations are sensitivity checks and are left as raw p-values.</p>'
        '<div class="table-wrap"><table class="model-table"><thead><tr><th>Variable paired with engagement</th><th>n</th><th>Pearson r</th><th>raw p</th><th>Holm p</th><th>95% CI</th><th>r²</th><th>Spearman ρ sensitivity</th><th>Spearman raw p</th></tr></thead><tbody>' + ''.join(body) + '</tbody></table></div>'
        + combined_exclusion_details("Show H4 correlation exclusions", exclusions)
        + '</details>'
    )


def chi_square_table_html(rows: list[dict[str, Any]], row_var: str, col_var: str, title: str) -> str:
    complete, excluded = make_complete_cases(rows, [row_var, col_var], f"chi-square {row_var} x {col_var}")
    row_levels = sorted({clean(row.get(row_var)) for row in complete if clean(row.get(row_var))})
    col_levels = sorted({clean(row.get(col_var)) for row in complete if clean(row.get(col_var))})
    if len(row_levels) < 2 or len(col_levels) < 2 or scipy_stats is None:
        return f'<details class="compact-details" open><summary>{h(title)}</summary>{model_status_note("Not enough categorical levels or SciPy unavailable.", "red")}{combined_exclusion_details("Show exclusions", excluded)}</details>'
    table = np.array([[sum(1 for row in complete if clean(row.get(row_var)) == rv and clean(row.get(col_var)) == cv) for cv in col_levels] for rv in row_levels], dtype=float)
    chi2, p_value, dof, expected = scipy_stats.chi2_contingency(table)
    n = table.sum()
    cramer_v = math.sqrt(float(chi2) / (n * max(1, min(table.shape[0] - 1, table.shape[1] - 1)))) if n > 0 else None
    rows_html: list[str] = []
    for i, rv in enumerate(row_levels):
        cells = [f'<th>{h(rv)}</th>']
        for j, cv in enumerate(col_levels):
            cells.append(f'<td>Observed: {int(table[i,j])}<br>Expected: {expected[i,j]:.2f}</td>')
        rows_html.append('<tr>' + ''.join(cells) + '</tr>')
    return (
        f'<details class="compact-details" open><summary>{h(title)} · n={int(n)}</summary>'
        f'<p>χ²({int(dof)}) = {fmt(float(chi2))}, p = {p_text(float(p_value))}, Cramér\'s V = {fmt(cramer_v)}.</p>'
        '<div class="table-wrap"><table class="model-table"><thead><tr><th>' + h(row_var) + '</th>' + ''.join(f'<th>{h(cv)}</th>' for cv in col_levels) + '</tr></thead><tbody>' + ''.join(rows_html) + '</tbody></table></div>'
        + combined_exclusion_details("Show chi-square exclusions", excluded)
        + '</details>'
    )


def dropout_ttest_html(rows: list[dict[str, Any]], variables: list[str]) -> str:
    body: list[str] = []
    exclusions: list[dict[str, str]] = []
    for variable in variables:
        complete, excluded = make_complete_cases(rows, ["completed_delayed", variable], f"dropout t-test {variable}")
        exclusions.extend(excluded)
        groups = {
            "Completed delayed retention": [float(parse_float(row.get(variable))) for row in complete if row.get("completed_delayed") == "Completed delayed retention" and parse_float(row.get(variable)) is not None],
            "Dropped before delayed retention": [float(parse_float(row.get(variable))) for row in complete if row.get("completed_delayed") == "Dropped before delayed retention" and parse_float(row.get(variable)) is not None],
        }
        if scipy_stats is None or any(len(values) < 2 for values in groups.values()):
            body.append(f'<tr><td>{h(variable)}</td><td colspan="10">Not enough data or SciPy unavailable. Complete pairwise n={len(complete)}.</td></tr>')
            continue
        comp = groups["Completed delayed retention"]
        drop = groups["Dropped before delayed retention"]
        lev_p = float(scipy_stats.levene(comp, drop).pvalue)
        t_result = scipy_stats.ttest_ind(comp, drop, equal_var=False)
        diff = statistics.fmean(comp) - statistics.fmean(drop)
        se = math.sqrt(statistics.variance(comp) / len(comp) + statistics.variance(drop) / len(drop))
        df_num = (statistics.variance(comp) / len(comp) + statistics.variance(drop) / len(drop)) ** 2
        df_den = ((statistics.variance(comp) / len(comp)) ** 2 / (len(comp) - 1)) + ((statistics.variance(drop) / len(drop)) ** 2 / (len(drop) - 1))
        df = df_num / df_den if df_den > 0 else float("nan")
        tcrit = scipy_stats.t.ppf(0.975, df) if math.isfinite(df) else 1.96
        cls = ' class="significant-row"' if float(t_result.pvalue) < ALPHA else ""
        body.append(f'<tr{cls}><td>{h(variable)}</td><td>{len(comp)}</td><td>{len(drop)}</td><td>{fmt(statistics.fmean(comp))}</td><td>{fmt(statistics.fmean(drop))}</td><td>{fmt(diff)}</td><td>{fmt(float(t_result.statistic))}</td><td>{fmt(float(df), 1)}</td><td>{p_text(float(t_result.pvalue))}</td><td>{ci_text(diff - tcrit * se, diff + tcrit * se)}</td><td>{p_text(lev_p)}</td></tr>')
    return (
        '<details class="compact-details" open><summary>Dropout follow-up Welch t-tests for continuous variables</summary>'
        '<p class="small">Difference is completed-delayed group minus dropped-before-delayed group. Welch t-tests are used as the default robust choice.</p>'
        '<div class="table-wrap"><table class="model-table"><thead><tr><th>Variable</th><th>n completed</th><th>n dropped</th><th>M completed</th><th>M dropped</th><th>Mean difference</th><th>t</th><th>df</th><th>p</th><th>95% CI</th><th>Levene p</th></tr></thead><tbody>' + ''.join(body) + '</tbody></table></div>'
        + combined_exclusion_details("Show dropout t-test exclusions", exclusions)
        + '</details>'
    )


# -----------------------------------------------------------------------------
# INTERVIEW SUBSAMPLE REPRESENTATIVENESS HELPERS
# -----------------------------------------------------------------------------

INTERVIEW_STATUS_INTERVIEWEE = "Interviewee"
INTERVIEW_STATUS_NOT_INTERVIEWED = "Full-sample non-interviewee"

INTERVIEW_COMPARISON_METRICS: list[dict[str, Any]] = [
    {"key": "age", "label": "Age", "kind": "number", "source_keys": ["age"]},
    {"key": "gender", "label": "Gender", "kind": "text", "source_keys": ["gender"]},
    {"key": "experiment_duration_seconds", "label": "Experiment duration", "kind": "duration", "source_keys": ["experiment_duration_seconds", "total_duration_seconds"]},
    {"key": "logs_creature_score_of_18", "label": "Creature score", "kind": "score18", "source_keys": ["logs_creature_score_of_18", "creature_score_of_18"]},
    {"key": "ret_immediate_score", "label": "Immediate retention", "kind": "proportion", "source_keys": ["ret_immediate_score"]},
    {"key": "ret_delayed_score", "label": "Delayed retention", "kind": "proportion", "source_keys": ["ret_delayed_score"]},
    {"key": "cl_intrinsic", "label": "IL", "kind": "number", "source_keys": ["cl_intrinsic"]},
    {"key": "cl_extraneous", "label": "EL", "kind": "number", "source_keys": ["cl_extraneous"]},
    {"key": "cl_germane", "label": "GL", "kind": "number", "source_keys": ["cl_germane"]},
    {"key": "engagement", "label": "Engagement", "kind": "number", "source_keys": ["eng_main", "engagement"]},
    {"key": "perceived_control", "label": "Control", "kind": "number", "source_keys": ["ctrl_perceived", "perceived_control"]},
    {"key": "ch0_duration_seconds", "label": "Ch0", "kind": "duration", "source_keys": ["ch0_duration_seconds"]},
    {"key": "time_to_sixth_creature_ch1_seconds", "label": "Ch1 sixth", "kind": "duration", "source_keys": ["time_to_sixth_creature_ch1_seconds"]},
    {"key": "time_to_sixth_creature_ch2_seconds", "label": "Ch2 sixth", "kind": "duration", "source_keys": ["time_to_sixth_creature_ch2_seconds"]},
    {"key": "time_to_sixth_creature_ch3_seconds", "label": "Ch3 sixth", "kind": "duration", "source_keys": ["time_to_sixth_creature_ch3_seconds"]},
]

INTERVIEW_COVARIATE_METRICS: list[dict[str, Any]] = [
    {"key": "completed_delayed_retention", "label": "Completed delayed retention", "kind": "text", "source_keys": ["completed_delayed_retention"]},
    {"key": "location", "label": "Location", "kind": "text", "source_keys": ["location"]},
    {"key": "co_present_participants", "label": "Co-present participants", "kind": "number", "source_keys": ["co_present_participants"]},
]


def first_metric_value(row: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        if key in row and clean(row.get(key)):
            return row.get(key)
    for key in keys:
        if key in row and row.get(key) is not None:
            return row.get(key)
    return None


def format_duration_seconds(value: float | int | None) -> str:
    if value is None or not math.isfinite(float(value)):
        return "—"
    raw_seconds = float(value)
    sign = "−" if raw_seconds < 0 else ""
    seconds = abs(raw_seconds)
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(round(seconds % 60))
    if secs == 60:
        secs = 0
        minutes += 1
    if minutes == 60:
        minutes = 0
        hours += 1
    if hours:
        return f"{sign}{hours}h {minutes:02d}m {secs:02d}s"
    if minutes:
        return f"{sign}{minutes}m {secs:02d}s"
    return f"{sign}{seconds:.1f}s"


def format_interview_metric_number(value: float | None, kind: str, *, digits: int = 2) -> str:
    if value is None or not math.isfinite(float(value)):
        return "—"
    if kind == "duration":
        return format_duration_seconds(value)
    if kind == "score18":
        return f"{float(value):.1f}/18"
    if kind == "proportion":
        return f"{float(value):.2f}"
    if kind == "count":
        return str(int(round(float(value))))
    return f"{float(value):.{digits}f}"


def mean_sd_metric_html(values: list[float], kind: str) -> str:
    clean_values = [float(value) for value in values if math.isfinite(float(value))]
    if not clean_values:
        return "—"
    mean = statistics.fmean(clean_values)
    sd = statistics.stdev(clean_values) if len(clean_values) >= 2 else None
    return (
        f"{format_interview_metric_number(mean, kind)} "
        f"({format_interview_metric_number(sd, kind) if sd is not None else '—'}), "
        f"n={len(clean_values)}"
    )


def text_distribution_html(values: list[Any], *, order: list[str] | None = None) -> str:
    counts: dict[str, int] = {}
    for value in values:
        label = clean(value) or "Unknown / missing"
        counts[label] = counts.get(label, 0) + 1
    if not counts:
        return "—"
    ordered: list[str] = []
    for label in order or []:
        if label in counts:
            ordered.append(label)
    ordered.extend(sorted(label for label in counts if label not in set(ordered)))
    return "; ".join(f"{h(label)}: {counts[label]}" for label in ordered)


def metric_values_for_rows(rows: list[dict[str, Any]], metric: dict[str, Any]) -> list[Any]:
    values: list[Any] = []
    for row in rows:
        value = first_metric_value(row, metric.get("source_keys") or [metric["key"]])
        if metric.get("kind") == "text":
            if clean(value):
                values.append(clean(value))
        else:
            parsed = parse_float(value)
            if parsed is not None:
                values.append(float(parsed))
    return values


def interview_metric_summary(rows: list[dict[str, Any]], metric: dict[str, Any]) -> str:
    values = metric_values_for_rows(rows, metric)
    if metric.get("kind") == "text":
        order = ["Yes", "No", "Male", "Female", "Other", "Unknown / missing", "CreaSp", "LivingR", "Remote"]
        return text_distribution_html(values, order=order)
    return mean_sd_metric_html([float(value) for value in values], clean(metric.get("kind")) or "number")


def compare_metric_cell(sample_rows: list[dict[str, Any]], full_rows: list[dict[str, Any]], metric: dict[str, Any]) -> str:
    return (
        f"<strong>Interview:</strong> {interview_metric_summary(sample_rows, metric)}"
        f"<br><span class=\"small\">Full: {interview_metric_summary(full_rows, metric)}</span>"
    )


def canonical_interview_location(row: dict[str, Any]) -> str:
    existing = clean(row.get("location"))
    if existing in {"CreaSp", "LivingR", "Remote"}:
        return existing
    room_type = clean(row.get("room_type"))
    if room_type == "Creative Space":
        return "CreaSp"
    if room_type == "Living Room":
        return "LivingR"
    if room_type in {"At home", "Remote"}:
        return "Remote"
    if clean(row.get("remote")).lower() in {"1", "true", "yes"}:
        return "Remote"
    return existing


def completed_delayed_label(row: dict[str, Any]) -> str:
    for key in ["completed_delayed_retention", "completed_delayed_retention_tick", "completed_delayed_retention_test"]:
        value = row.get(key)
        if isinstance(value, bool):
            return "Yes" if value else "No"
        text = clean(value).lower()
        if text in {"yes", "true", "1", "✓", "completed", "completed delayed retention"}:
            return "Yes"
        if text in {"no", "false", "0", "dropped", "dropped before delayed retention"}:
            return "No"
    return "Yes" if parse_float(row.get("ret_delayed_score")) is not None else "No"


def normalise_interview_comparison_rows(participants: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for participant in participants:
        row = dict(participant)
        participant_id = first_present(row, ["participant_id", "MCID", "mcid"])
        if not participant_id:
            continue
        condition = canonical_condition(first_present(row, ["condition", "Condition", "condition_raw"]))
        row["participant_id"] = participant_id
        row["MCID"] = participant_id
        row["condition"] = condition
        row["location"] = canonical_interview_location(row)
        row["co_present_participants"] = first_metric_value(row, ["co_present_participants", "same_room_n"])
        row["engagement"] = first_metric_value(row, ["engagement", "eng_main"])
        row["perceived_control"] = first_metric_value(row, ["perceived_control", "ctrl_perceived"])
        row["completed_delayed_retention"] = completed_delayed_label(row)
        rows.append(row)
    return rows


def attach_retention_scores_for_interview_rows(participants: list[dict[str, Any]], warnings: list[str]) -> None:
    """Attach participant-level retention scores from retention_scores_final.tsv."""
    participant_scores, errors, retention_warnings = participant_retention_scores_from_final()
    warnings.extend(errors)
    warnings.extend(retention_warnings)

    for participant in participants:
        participant_id = first_present(participant, ["participant_id", "MCID", "mcid"]).upper()
        scores = participant_scores.get(participant_id, {})
        if scores.get("Immediate") is not None:
            participant["ret_immediate_score"] = scores.get("Immediate")
        if scores.get("Delayed") is not None:
            participant["ret_delayed_score"] = scores.get("Delayed")


def build_participants_for_interview_representativeness() -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    participants: list[dict[str, Any]] = []

    if build_merged_dataset is not None:
        try:
            survey_rows = read_tsv(SURVEY_EXPORT_PATH)
            log_index = load_log_index(DATA_LOG_DIR) if load_log_index is not None else {}
            merged = build_merged_dataset(
                survey_rows,
                log_index,
                collection_locations_path=COLLECTION_LOCATIONS_PATH,
            )
            participants = [dict(row) for row in merged.get("participants", []) if clean(row.get("participant_id"))]
            attach_retention_scores_for_interview_rows(participants, warnings)
        except Exception as exc:
            warnings.append(f"Could not rebuild the sum_merged-style participant table for interview representativeness: {exc}")
            participants = []

    if not participants:
        inferential_rows, inferential_warnings = build_rows_for_inferential_models()
        warnings.append("Using the survey-only inferential participant rows as a fallback; log-duration and time-to-sixth-creature metrics may be unavailable.")
        warnings.extend(inferential_warnings)
        participants = [dict(row, participant_id=row.get("MCID")) for row in inferential_rows]

    return normalise_interview_comparison_rows(participants), warnings


def transcript_participant_ids_from_csv(path: Path) -> list[str]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
    except Exception:
        return []
    if not rows:
        return []
    header: list[str] | None = None
    header_index = 0
    for index, row in enumerate(rows):
        cleaned = [clean(cell) for cell in row]
        if any(cleaned):
            header = cleaned
            header_index = index
            break
    if not header:
        return []
    speaker_index = {value.lower(): idx for idx, value in enumerate(header)}.get("speaker")
    if speaker_index is None:
        return []
    ids: list[str] = []
    for row in rows[header_index + 1:]:
        speaker = clean(row[speaker_index]) if speaker_index < len(row) else ""
        if speaker and speaker.lower() != "researcher" and speaker not in ids:
            ids.append(speaker)
    return ids


def transcript_participant_ids_from_xlsx(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook as _load_workbook
        workbook = _load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header: list[str] | None = None
        for row in rows:
            cleaned = [clean(cell) for cell in row]
            if any(cleaned):
                header = cleaned
                break
        if not header:
            return []
        speaker_index = {value.lower(): idx for idx, value in enumerate(header)}.get("speaker")
        if speaker_index is None:
            return []
        ids: list[str] = []
        for row in rows:
            speaker = clean(row[speaker_index]) if speaker_index < len(row) else ""
            if speaker and speaker.lower() != "researcher" and speaker not in ids:
                ids.append(speaker)
        return ids
    except Exception:
        return []


def fallback_interview_overview_from_transcripts(transcripts_dir: Path) -> dict[str, Any]:
    if not transcripts_dir.exists():
        return {
            "available": False,
            "n_files": 0,
            "n_turns": 0,
            "unique_participant_ids": [],
            "transcripts": [],
            "notes": ["Interview transcript directory not found."],
        }
    paths = sorted(
        path for path in transcripts_dir.iterdir()
        if path.is_file() and not path.name.startswith("~$") and path.suffix.lower() in {".csv", ".xlsx"}
    )
    all_ids: set[str] = set()
    transcripts: list[dict[str, Any]] = []
    for index, path in enumerate(paths, start=1):
        ids = transcript_participant_ids_from_csv(path) if path.suffix.lower() == ".csv" else transcript_participant_ids_from_xlsx(path)
        all_ids.update(ids)
        transcripts.append({"transcript_id": f"interview-{index:02d}", "filename": path.name, "speaker_ids": ids, "n_turns": None})
    return {
        "available": bool(paths),
        "n_files": len(paths),
        "n_turns": 0,
        "unique_participant_ids": sorted(all_ids),
        "transcripts": transcripts,
        "notes": [] if paths else ["No .csv or .xlsx transcript files found in the interview transcript directory."],
    }


def load_interview_overview_for_statistics(participants: list[dict[str, Any]]) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []
    if load_interview_overview is not None:
        try:
            overview = load_interview_overview(
                INTERVIEW_TRANSCRIPTS_DIR,
                participants=participants,
                manifest_path=INTERVIEW_MANIFEST_PATH,
            )
        except Exception as exc:
            warnings.append(f"Could not load interview overview with sum_merged helper logic: {exc}")
            overview = fallback_interview_overview_from_transcripts(INTERVIEW_TRANSCRIPTS_DIR)
    else:
        warnings.append("sum_merged interview helper was unavailable; using a local transcript parser fallback.")
        overview = fallback_interview_overview_from_transcripts(INTERVIEW_TRANSCRIPTS_DIR)

    overview = dict(overview)
    ids = [clean(participant_id) for participant_id in overview.get("unique_participant_ids") or [] if clean(participant_id)]
    overview["unique_participant_ids"] = sorted(ids)
    overview["unique_participant_ids_upper"] = sorted({participant_id.upper() for participant_id in ids})
    for note in overview.get("notes") or []:
        if clean(note):
            warnings.append(clean(note))
    return overview, warnings


def interview_metadata_table(overview: dict[str, Any], rows: list[dict[str, Any]]) -> str:
    interview_ids = set(overview.get("unique_participant_ids_upper") or [])
    merged_ids = {clean(row.get("participant_id")).upper() for row in rows if clean(row.get("participant_id"))}
    matched_ids = sorted(interview_ids & merged_ids)
    missing_ids = sorted(interview_ids - merged_ids)
    body = (
        f"<tr><td>Transcript directory</td><td>{h(INTERVIEW_TRANSCRIPTS_DIR)}</td></tr>"
        f"<tr><td>Transcript files read</td><td>{h(overview.get('n_files', 0))}</td></tr>"
        f"<tr><td>Unique interview participant IDs in transcripts</td><td>{len(interview_ids)}</td></tr>"
        f"<tr><td>Interview participants found in full sample</td><td>{len(matched_ids)}</td></tr>"
        f"<tr><td>Interview participant IDs not found in full sample</td><td>{h(', '.join(missing_ids) if missing_ids else '—')}</td></tr>"
    )
    return '<h4>Interview transcript input check</h4><div class="table-wrap"><table><thead><tr><th>Check</th><th>Value</th></tr></thead><tbody>' + body + '</tbody></table></div>'


def group_label_for_interview_table(group: str) -> str:
    if group == "Overall":
        return "overall"
    return DISPLAY_CONDITION.get(group, group)


def interview_comparison_table(rows: list[dict[str, Any]], metrics: list[dict[str, Any]], title: str) -> str:
    full_rows = rows
    sample_rows = [row for row in rows if row.get("interview_status") == INTERVIEW_STATUS_INTERVIEWEE]
    header = '<tr><th>Group</th><th>n</th>' + ''.join(f'<th>{h(metric["label"])}</th>' for metric in metrics) + '</tr>'
    body: list[str] = []
    for group in CONDITION_ORDER + ["Overall"]:
        scoped_sample = sample_rows if group == "Overall" else [row for row in sample_rows if row.get("condition") == group]
        scoped_full = full_rows if group == "Overall" else [row for row in full_rows if row.get("condition") == group]
        cells = [
            f'<th>{h(group_label_for_interview_table(group))}</th>',
            f'<td><strong>Interview:</strong> {len(scoped_sample)}<br><span class="small">Full: {len(scoped_full)}</span></td>',
        ]
        for metric in metrics:
            cells.append('<td>' + compare_metric_cell(scoped_sample, scoped_full, metric) + '</td>')
        body.append('<tr>' + ''.join(cells) + '</tr>')
    return f'<h4>{h(title)}</h4><p class="small">Each cell shows the interview subsample first and the full included sample second, mirroring the merged-report interview panel.</p><div class="table-wrap"><table class="model-table"><thead>{header}</thead><tbody>' + ''.join(body) + '</tbody></table></div>'


def rows_with_interview_status() -> tuple[list[dict[str, Any]], dict[str, Any], list[str]]:
    rows, warnings = build_participants_for_interview_representativeness()
    overview, overview_warnings = load_interview_overview_for_statistics(rows)
    warnings.extend(overview_warnings)
    interview_ids = set(overview.get("unique_participant_ids_upper") or [])
    for row in rows:
        participant_id = clean(row.get("participant_id")).upper()
        row["interview_status"] = INTERVIEW_STATUS_INTERVIEWEE if participant_id in interview_ids else INTERVIEW_STATUS_NOT_INTERVIEWED
    not_found = sorted(interview_ids - {clean(row.get("participant_id")).upper() for row in rows if clean(row.get("participant_id"))})
    overview["interview_participant_ids_not_in_full_sample"] = not_found
    return rows, overview, warnings


def interview_status_observed_expected_table(rows: list[dict[str, Any]]) -> str:
    categories = [INTERVIEW_STATUS_INTERVIEWEE, INTERVIEW_STATUS_NOT_INTERVIEWED]
    counts = {(condition, category): 0 for condition in CONDITION_ORDER for category in categories}
    for row in rows:
        condition = row.get("condition")
        status = row.get("interview_status")
        if condition in CONDITION_ORDER and status in categories:
            counts[(condition, status)] += 1
    row_totals = {condition: sum(counts[(condition, category)] for category in categories) for condition in CONDITION_ORDER}
    col_totals = {category: sum(counts[(condition, category)] for condition in CONDITION_ORDER) for category in categories}
    grand_total = sum(row_totals.values())
    body: list[str] = []
    small_expected = 0
    for condition in CONDITION_ORDER:
        cells = [f'<th>{h(DISPLAY_CONDITION[condition])}</th>']
        for category in categories:
            expected = row_totals[condition] * col_totals[category] / grand_total if grand_total else 0
            if expected < 5:
                small_expected += 1
            cells.append(f'<td>Observed: {counts[(condition, category)]}<br>Expected: {expected:.2f}</td>')
        body.append('<tr>' + ''.join(cells) + '</tr>')
    read = "Expected counts look acceptable." if small_expected == 0 else f"{small_expected} expected cell count(s) below 5; inspect before interpreting the chi-square approximation."
    return '<h4>Interview-status contingency table for assumption checking only</h4><p class="small">This table displays observed and expected counts for design condition × interview status.</p><div class="table-wrap"><table><thead><tr><th>Design condition</th>' + ''.join(f'<th>{h(category)}</th>' for category in categories) + '</tr></thead><tbody>' + ''.join(body) + f'</tbody></table></div><p>{h(read)}</p>'


def interview_continuous_diagnostics(rows: list[dict[str, Any]], variables: list[dict[str, Any]]) -> str:
    diagnostics = ""
    for variable in variables:
        complete, _excluded = make_complete_cases(rows, ["interview_status", variable["key"]], f"interview representativeness {variable['key']}")
        groups = {
            label: [float(parse_float(row.get(variable["key"]))) for row in complete if row.get("interview_status") == label and parse_float(row.get(variable["key"])) is not None]
            for label in [INTERVIEW_STATUS_INTERVIEWEE, INTERVIEW_STATUS_NOT_INTERVIEWED]
        }
        levene_p = None
        shapiro_bits: list[str] = []
        if scipy_stats is not None and all(len(values) >= 2 for values in groups.values()):
            try:
                levene_p = float(scipy_stats.levene(*groups.values()).pvalue)
            except Exception:
                levene_p = None
        for label, values in groups.items():
            shapiro_p = None
            if scipy_stats is not None and 3 <= len(values) <= 5000:
                try:
                    shapiro_p = float(scipy_stats.shapiro(values).pvalue)
                except Exception:
                    shapiro_p = None
            shapiro_bits.append(f"{label}: Shapiro p={p_text(shapiro_p)} (n={len(values)})")
        tests = [
            {"Assumption": "Normality within interview-status groups", "How to read": "Shapiro p ≥ .05 is reassuring, but visual inspection matters more for small interview subsamples.", "Diagnostic": "; ".join(shapiro_bits), "Flag": '<span class="badge badge-neutral">inspect</span>'},
            {"Assumption": "Homogeneity of variance", "How to read": "Levene p ≥ .05 supports equal variances; Welch t-tests are reported regardless.", "Diagnostic": f"Levene p = {p_text(levene_p)}", "Flag": interpretation_badge(None if levene_p is None else levene_p >= .05, "equal variance plausible", "Welch/inspect")},
        ]
        diagnostics += f'<details class="compact-details"><summary>Interview continuous follow-up diagnostics: {h(variable["label"])} · n={len(complete)}</summary>' + diagnostic_tests_table(tests) + '</details>'
    return diagnostics


def interview_ttest_html(rows: list[dict[str, Any]], variables: list[dict[str, Any]]) -> str:
    body: list[str] = []
    exclusions: list[dict[str, str]] = []
    for variable in variables:
        key = variable["key"]
        kind = clean(variable.get("kind")) or "number"
        complete, excluded = make_complete_cases(rows, ["interview_status", key], f"interview representativeness t-test {key}")
        exclusions.extend(excluded)
        interview_values = [float(parse_float(row.get(key))) for row in complete if row.get("interview_status") == INTERVIEW_STATUS_INTERVIEWEE and parse_float(row.get(key)) is not None]
        non_values = [float(parse_float(row.get(key))) for row in complete if row.get("interview_status") == INTERVIEW_STATUS_NOT_INTERVIEWED and parse_float(row.get(key)) is not None]
        if scipy_stats is None or len(interview_values) < 2 or len(non_values) < 2:
            body.append(f'<tr><td>{h(variable["label"])}</td><td colspan="10">Not enough data or SciPy unavailable. Complete pairwise n={len(complete)}.</td></tr>')
            continue
        lev_p = None
        try:
            lev_p = float(scipy_stats.levene(interview_values, non_values).pvalue)
        except Exception:
            lev_p = None
        t_result = scipy_stats.ttest_ind(interview_values, non_values, equal_var=False)
        mean_interview = statistics.fmean(interview_values)
        mean_non = statistics.fmean(non_values)
        diff = mean_interview - mean_non
        var_interview = statistics.variance(interview_values)
        var_non = statistics.variance(non_values)
        se = math.sqrt(var_interview / len(interview_values) + var_non / len(non_values))
        df_num = (var_interview / len(interview_values) + var_non / len(non_values)) ** 2
        df_den = ((var_interview / len(interview_values)) ** 2 / (len(interview_values) - 1)) + ((var_non / len(non_values)) ** 2 / (len(non_values) - 1))
        df = df_num / df_den if df_den > 0 else float("nan")
        tcrit = scipy_stats.t.ppf(0.975, df) if math.isfinite(df) else 1.96
        ci_low = diff - tcrit * se if math.isfinite(se) else None
        ci_high = diff + tcrit * se if math.isfinite(se) else None
        cls = ' class="significant-row"' if math.isfinite(float(t_result.pvalue)) and float(t_result.pvalue) < ALPHA else ""
        body.append(
            f'<tr{cls}><td>{h(variable["label"])}</td>'
            f'<td>{len(interview_values)}</td><td>{len(non_values)}</td>'
            f'<td>{format_interview_metric_number(mean_interview, kind)}</td>'
            f'<td>{format_interview_metric_number(mean_non, kind)}</td>'
            f'<td>{format_interview_metric_number(diff, kind)}</td>'
            f'<td>{fmt(float(t_result.statistic))}</td><td>{fmt(float(df), 1)}</td><td>{p_text(float(t_result.pvalue))}</td>'
            f'<td>[{format_interview_metric_number(ci_low, kind)}, {format_interview_metric_number(ci_high, kind)}]</td>'
            f'<td>{p_text(lev_p)}</td></tr>'
        )
    return (
        '<details class="compact-details" open><summary>Interview representativeness Welch t-tests for continuous variables</summary>'
        '<p class="small">Difference is interviewees minus full-sample non-interviewees. The descriptive table above compares interviewees with the full included sample; these tests use non-interviewees as the independent reference group.</p>'
        '<div class="table-wrap"><table class="model-table"><thead><tr><th>Variable</th><th>n interview</th><th>n non-interview</th><th>M interview</th><th>M non-interview</th><th>Mean difference</th><th>t</th><th>df</th><th>p</th><th>95% CI</th><th>Levene p</th></tr></thead><tbody>' + ''.join(body) + '</tbody></table></div>'
        + combined_exclusion_details("Show interview t-test exclusions", exclusions)
        + '</details>'
    )


def inferential_interview_representativeness() -> str:
    rows, overview, warnings = rows_with_interview_status()
    interview_rows = [row for row in rows if row.get("interview_status") == INTERVIEW_STATUS_INTERVIEWEE]
    errors: list[str] = []
    if not rows:
        errors.append("No participant-level rows could be built for interview representativeness diagnostics.")
    if not interview_rows:
        errors.append("No interviewees from /data/transcripts were found in the full participant-level data.")

    continuous_variables = [metric for metric in INTERVIEW_COMPARISON_METRICS if metric.get("kind") != "text"] + [INTERVIEW_COVARIATE_METRICS[-1]]
    data_html = (
        interview_metadata_table(overview, rows)
        + interview_comparison_table(rows, INTERVIEW_COMPARISON_METRICS, "Interview subsample compared with full sample")
        + interview_comparison_table(rows, INTERVIEW_COVARIATE_METRICS, "Additional interview representativeness checks: completion and covariates")
        + interview_status_observed_expected_table(rows)
        + covariate_feasibility_table(rows)
    )
    diagnostics = interview_continuous_diagnostics(rows, continuous_variables)
    final_models_html = final_models_wrapper(
        chi_square_table_html(rows, "interview_status", "condition", "Interview representativeness: interview status × design condition"),
        chi_square_table_html(rows, "interview_status", "gender", "Interview representativeness: interview status × gender"),
        chi_square_table_html(rows, "interview_status", "location", "Interview representativeness: interview status × location"),
        chi_square_table_html(rows, "interview_status", "completed_delayed_retention", "Interview representativeness: interview status × delayed-retention completion"),
        interview_ttest_html(rows, continuous_variables),
    )
    return assumption_section_shell(
        "Interview subsample representativeness",
        "This section checks whether participants appearing in /data/transcripts look representative of the full included sample used in the manuscript statistics.",
        [
            "Interviewee status is derived from non-researcher Speaker IDs in /data/transcripts/*.csv and /data/transcripts/*.xlsx, using the same interview-overview logic as sum_merged when available.",
            "Descriptive comparisons mirror the merged-report panel: interview subsample first, full included sample second, split by design condition and overall.",
            "Inferential checks compare interviewees with full-sample non-interviewees so the groups are independent.",
            "Additional covariate checks include delayed-retention completion, collection location, and co-present participants.",
        ],
        [
            {"Assumption": "Correct interviewee identification", "How to test": "Read /data/transcripts and match non-researcher Speaker IDs to full-sample participant IDs.", "How to read": "IDs not found in the full sample should be explained before making representativeness claims."},
            {"Assumption": "Frequency data for categorical checks", "How to test": "Build interview-status × condition/gender/location/completion tables.", "How to read": "Expected counts below 5 make chi-square p-values approximate; treat them as screening diagnostics."},
            {"Assumption": "Independent groups for follow-up tests", "How to test": "Compare interviewees against non-interviewees rather than against the full sample in inferential tests.", "How to read": "This avoids testing overlapping groups."},
            {"Assumption": "Continuous follow-up t-test assumptions", "How to test": "Inspect normality within interview-status groups and Levene's test for variance equality.", "How to read": "Welch t-tests are reported as the default robust choice."},
        ],
        data_html,
        diagnostics,
        final_models_html,
        status_messages(errors, warnings, "Interview representativeness checks were generated."),
        combined_exclusion_details("Show interview representativeness unmatched transcript IDs", [
            {"MCID": participant_id, "condition": "", "reason": "Speaker ID appeared in /data/transcripts but was not found in the full participant-level data."}
            for participant_id in overview.get("interview_participant_ids_not_in_full_sample", [])
        ]),
    )


# =============================================================================
# ASSUMPTION CHECKS AND FINAL INFERENTIAL MODELS
# =============================================================================
# These functions prepare complete-case rows, diagnose assumptions, and then run
# the final planned models plus covariate-adjusted sensitivity models where feasible.


def build_rows_for_inferential_models(
    retention_scores_path: Path | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Build one transparent participant-level row used for assumption checks and final models."""
    warnings: list[str] = []
    survey_rows = read_tsv(SURVEY_EXPORT_PATH)
    immediate_by_mcid: dict[str, dict[str, str]] = {}
    for row in survey_rows:
        participant_id = mcid_from_row(row)
        if not participant_id:
            continue
        if not delayed_flag(row):
            immediate_by_mcid.setdefault(participant_id, row)

    participant_retention_scores, retention_errors, retention_warnings = participant_retention_scores_from_final(retention_scores_path)
    for error in retention_errors:
        warnings.append(error)
    warnings.extend(retention_warnings)
    scores_by_mcid_moment: dict[tuple[str, str], float] = {}
    for participant_id, scores in participant_retention_scores.items():
        immediate_score = scores.get("Immediate")
        delayed_score = scores.get("Delayed")
        if immediate_score is not None and 0 <= immediate_score <= 2:
            scores_by_mcid_moment[(participant_id, "Immediate")] = float(immediate_score)
        if delayed_score is not None and 0 <= delayed_score <= 2:
            scores_by_mcid_moment[(participant_id, "Delayed")] = float(delayed_score)

    raw_locations, location_warnings = load_collection_location_map()
    if location_warnings:
        warnings.extend(f"{message}; location/co-present covariates may be missing." for message in location_warnings)
    log_start_by_mcid, log_warnings = log_starts_by_mcid()
    warnings.extend(log_warnings)
    context_rows, context_excluded = build_collection_context_rows(
        immediate_by_mcid,
        raw_locations,
        log_start_by_mcid,
        require_valid_condition=False,
    )
    if context_excluded:
        warnings.append(f"Collection-context covariates missing for {len(context_excluded)} participant row(s); affected rows remain available for base models.")
    context_by_mcid = {clean(row.get("MCID")): row for row in context_rows}

    rows: list[dict[str, Any]] = []
    for participant_id, row in sorted(immediate_by_mcid.items()):
        condition = canonical_condition(first_present(row, ["condition", "Condition", "CONDITION", "experiment_condition", "condition_raw"]))
        if condition not in CONDITION_ORDER:
            continue
        codes = CONDITION_CODES[condition]
        age = parse_float(first_present(row, ["age", "Age"]))
        if age is not None and not (0 <= age <= 120):
            age = None
        gender = normalise_gender(first_present(row, ["gender", "Gender"]))
        context = context_by_mcid.get(participant_id, {})

        intrinsic_cols = [f"cl_ch{chapter}_scores_{index}" for chapter in (1, 2, 3) for index in (1, 2, 3)]
        intrinsic_values = [parse_float(row.get(column)) for column in intrinsic_cols]
        cl_intrinsic = statistics.fmean(intrinsic_values) if all(value is not None and 0 <= value <= 10 for value in intrinsic_values) else None

        env_cols = [f"cl_ch{chapter}_scores_{index}" for chapter in (1, 2, 3) for index in (4, 5, 6, 7)]
        instr_cols = [f"cl_overall_scores_{index}" for index in (1, 2, 3)]
        interact_cols = [f"cl_overall_scores_{index}" for index in (4, 5, 6, 7)]
        env_values = [parse_float(row.get(column)) for column in env_cols]
        instr_raw = [parse_float(row.get(column)) for column in instr_cols]
        interact_raw = [parse_float(row.get(column)) for column in interact_cols]
        if all(value is not None and 0 <= value <= 10 for value in env_values + instr_raw + interact_raw):
            cl_extraneous = statistics.fmean([statistics.fmean(env_values), statistics.fmean([10 - value for value in instr_raw]), statistics.fmean([10 - value for value in interact_raw])])
        else:
            cl_extraneous = None

        germane_cols = [f"cl_overall_scores_{index}" for index in (8, 9, 10, 11)]
        germane_values = [parse_float(row.get(column)) for column in germane_cols]
        cl_germane = statistics.fmean(germane_values) if all(value is not None and 0 <= value <= 10 for value in germane_values) else None

        eng_chapter_cols = [f"eng_ch{chapter}_scores_{index}" for chapter in (1, 2, 3) for index in (1, 2, 3, 4, 5)]
        eng_overall_cols = ["eng_overall_scores_1", "eng_overall_scores_2", "eng_overall_scores_3", "eng_overall_scores_4"]
        eng_chapter_values = [parse_float(row.get(column)) for column in eng_chapter_cols]
        eng_overall_raw = [parse_float(row.get(column)) for column in eng_overall_cols]
        if all(value is not None and 1 <= value <= 7 for value in eng_chapter_values + eng_overall_raw):
            eng_overall_values = [8 - eng_overall_raw[0], 8 - eng_overall_raw[1], eng_overall_raw[2], eng_overall_raw[3]]
            engagement = statistics.fmean([statistics.fmean(eng_chapter_values), statistics.fmean(eng_overall_values)])
        else:
            engagement = None

        ctrl1 = parse_float(row.get("ctrl_scores_1"))
        ctrl2 = parse_float(row.get("ctrl_scores_2"))
        perceived_control = (ctrl1 + ctrl2) / 2 if ctrl1 is not None and ctrl2 is not None and 1 <= ctrl1 <= 7 and 1 <= ctrl2 <= 7 else None

        retention_form_order = first_present(row, [
            "retention_form_order", "RetentionFormOrder", "retention_form", "form_order",
            "retention_order", "cue_order", "image_first", "name_first"
        ])

        ret_immediate_score = scores_by_mcid_moment.get((participant_id, "Immediate"))
        ret_delayed_score = scores_by_mcid_moment.get((participant_id, "Delayed"))
        retention_decay = ret_delayed_score - ret_immediate_score if ret_immediate_score is not None and ret_delayed_score is not None else None

        rows.append({
            "MCID": participant_id,
            "condition": condition,
            "required_pause_contrast": codes["required_pause_contrast"],
            "optional_pause_contrast": codes["optional_pause_contrast"],
            "ret_immediate_score": ret_immediate_score,
            "ret_delayed_score": ret_delayed_score,
            "retention_decay": retention_decay,
            "cl_intrinsic": cl_intrinsic,
            "cl_extraneous": cl_extraneous,
            "cl_germane": cl_germane,
            "engagement": engagement,
            "perceived_control": perceived_control,
            "retention_form_order": retention_form_order,
            "location": context.get("location"),
            "co_present_participants": context.get("co_present_participants"),
            "age": age,
            "gender": gender,
        })
    return rows, warnings


def covariate_feasibility_table(rows: list[dict[str, Any]]) -> str:
    location_counts: dict[str, int] = {}
    gender_counts: dict[str, int] = {}
    for row in rows:
        if clean(row.get("location")):
            location_counts[clean(row.get("location"))] = location_counts.get(clean(row.get("location")), 0) + 1
        if clean(row.get("gender")):
            gender_counts[clean(row.get("gender"))] = gender_counts.get(clean(row.get("gender")), 0) + 1
    location_note = "Feasible as a three-level categorical covariate." if all(location_counts.get(level, 0) > 0 for level in ["CreaSp", "LivingR", "Remote"]) else "Not all CreaSp/LivingR/Remote levels are present in these complete cases; keep the three-level coding where possible, but inspect sparse/absent levels."
    location_counts_text = "; ".join(f"{key}: {value}" for key, value in sorted(location_counts.items())) or "none"
    gender_counts_text = "; ".join(f"{key}: {value}" for key, value in sorted(gender_counts.items())) or "none"
    copresent_n = sum(parse_float(row.get("co_present_participants")) is not None for row in rows)
    age_n = sum(parse_float(row.get("age")) is not None for row in rows)
    rows_html = (
        f'<tr><td>location</td><td>CreaSp, LivingR, Remote</td><td>{h(location_counts_text)}</td><td>{h(location_note)}</td></tr>'
        f'<tr><td>co-present participants</td><td>numeric count; Remote coded 0</td><td>available n={copresent_n}</td><td>Included in covariate-adjusted sensitivity models where complete and feasible.</td></tr>'
        f'<tr><td>age</td><td>numeric years</td><td>available n={age_n}</td><td>Include if complete and plausible.</td></tr>'
        f'<tr><td>gender</td><td>categorical; Male/Female/Other when present</td><td>{h(gender_counts_text)}</td><td>Inspect sparse levels before inclusion.</td></tr>'
    )
    return '<h4>Covariate operationalisation / feasibility</h4><div class="table-wrap"><table><thead><tr><th>Covariate</th><th>Planned coding</th><th>Observed in complete-case pool</th><th>Reading</th></tr></thead><tbody>' + rows_html + '</tbody></table></div>'


def inferential_x_assumptions() -> str:
    rows, warnings = build_rows_for_inferential_models()
    errors: list[str] = [] if rows else ["No participant-level rows could be built for assumption checks."]
    base_complete, base_excluded = make_complete_cases(rows, ["perceived_control", "required_pause_contrast", "optional_pause_contrast"], "X base")
    cov_complete, cov_excluded = make_complete_cases(rows, ["perceived_control", "required_pause_contrast", "optional_pause_contrast", "location", "co_present_participants", "age", "gender"], "X covariate")
    diagnostics = diagnostics_block("Base diagnostic: perceived control ~ C1 + C2", fit_diagnostic_residuals(rows, "perceived_control", ["required_pause_contrast", "optional_pause_contrast"], []))
    diagnostics += diagnostics_block("Covariate diagnostic: perceived control ~ C1 + C2 + covariates", fit_diagnostic_residuals(rows, "perceived_control", ["required_pause_contrast", "optional_pause_contrast", "co_present_participants", "age"], ["location", "gender"]))
    data_html = condition_count_table(base_complete, "Base complete cases", ["perceived_control"]) + condition_count_table(cov_complete, "Covariate complete cases", ["perceived_control"]) + covariate_feasibility_table(cov_complete)
    final_models_html = final_models_wrapper(
        ols_model_html(
            "Final base model: perceived control",
            "perceived_control ~ C1 + C2",
            fit_ols_arrays(rows, "perceived_control", ["required_pause_contrast", "optional_pause_contrast"], []),
        ),
        ols_model_html(
            "Final covariate-adjusted sensitivity model: perceived control",
            "perceived_control ~ C1 + C2 + location + co_present_participants + age + gender",
            fit_ols_arrays(rows, "perceived_control", ["required_pause_contrast", "optional_pause_contrast", "co_present_participants", "age"], ["location", "gender"]),
        ),
    )
    return assumption_section_shell(
        "Hypothesis X — Design → perceived control",
        "Checkpoint Design is expected to affect perceived control; optional pauses are expected to produce higher perceived control than the system-controlled designs.",
        ["Outcome: perceived control, continuous 1–7.", "Predictor: three-level Checkpoint Design represented by C1 and C2 planned contrasts.", "Base diagnostic requires perceived control and both contrasts; covariate diagnostic also requires location, co-present participants, age, and gender."],
        [
            {"Assumption": "Independent observations", "How to test": "Check design: each MCID contributes one immediate perceived-control score.", "How to read": "No repeated rows or clusters should be analysed as independent cases."},
            {"Assumption": "Linearity/additivity", "How to test": "Inspect residual-vs-fitted plots from the diagnostic regression.", "How to read": "A random cloud supports the assumption; curves/patterns indicate problems."},
            {"Assumption": "Normality of residuals", "How to test": "Inspect Q-Q plot and Shapiro test of residuals.", "How to read": "Q-Q points near the diagonal and Shapiro p ≥ .05 are reassuring."},
            {"Assumption": "Homoscedasticity", "How to test": "Inspect residual-vs-fitted plot and Breusch-Pagan diagnostic.", "How to read": "No funnel and p ≥ .05 are reassuring."},
            {"Assumption": "Multicollinearity", "How to test": "Inspect VIF for C1, C2, and covariates.", "How to read": "VIF ≤ 5 is usually acceptable; >10 is serious."},
        ],
        data_html,
        diagnostics,
        final_models_html,
        status_messages(errors, warnings, "X assumption checks and final regression models were generated."),
        combined_exclusion_details("Show X complete-case exclusions", base_excluded + cov_excluded),
    )


def retention_long_rows(rows: list[dict[str, Any]], include_covariates: bool) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    long_rows: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    needed_cov = ["location", "co_present_participants", "age", "gender"] if include_covariates else []
    for row in rows:
        for time_label, time_value, column in [("Immediate", 0.0, "ret_immediate_score"), ("Delayed", 1.0, "ret_delayed_score")]:
            missing: list[str] = []
            if parse_float(row.get(column)) is None:
                missing.append(column)
            for cov in needed_cov:
                if cov in {"location", "gender"}:
                    if not clean(row.get(cov)):
                        missing.append(cov)
                elif parse_float(row.get(cov)) is None:
                    missing.append(cov)
            if missing:
                excluded.append({"MCID": clean(row.get("MCID")), "condition": clean(row.get("condition")), "reason": f"H1/Y {time_label}: missing/invalid " + ", ".join(missing)})
                continue
            out = dict(row)
            out["time"] = time_value
            out["time_label"] = time_label
            out["retention_score"] = parse_float(row.get(column))
            out["time_x_required_pause_contrast"] = time_value * float(row.get("required_pause_contrast"))
            out["time_x_optional_pause_contrast"] = time_value * float(row.get("optional_pause_contrast"))
            long_rows.append(out)
    return long_rows, excluded


def inferential_h1_y_assumptions() -> str:
    rows, warnings = build_rows_for_inferential_models()
    errors: list[str] = [] if rows else ["No participant-level rows could be built for retention diagnostics."]
    base_long, base_excluded = retention_long_rows(rows, include_covariates=False)
    cov_long, cov_excluded = retention_long_rows(rows, include_covariates=True)
    diagnostics = diagnostics_block("Base diagnostic proxy: retention ~ time*C1 + time*C2", fit_diagnostic_residuals(base_long, "retention_score", ["time", "required_pause_contrast", "optional_pause_contrast", "time_x_required_pause_contrast", "time_x_optional_pause_contrast"], []))
    diagnostics += diagnostics_block("Covariate diagnostic proxy: retention ~ time*C1 + time*C2 + covariates", fit_diagnostic_residuals(cov_long, "retention_score", ["time", "required_pause_contrast", "optional_pause_contrast", "time_x_required_pause_contrast", "time_x_optional_pause_contrast", "co_present_participants", "age"], ["location", "gender"]))
    participant_counts = f'<h4>Long-format retention rows</h4><p>Base diagnostic rows: {len(base_long)} observations from {len({row.get("MCID") for row in base_long})} participants. Covariate diagnostic rows: {len(cov_long)} observations from {len({row.get("MCID") for row in cov_long})} participants.</p>'
    delayed_missing = [row for row in rows if parse_float(row.get("ret_immediate_score")) is not None and parse_float(row.get("ret_delayed_score")) is None]
    participant_counts += f'<p class="small">Participants with immediate retention but no delayed retention score: {len(delayed_missing)}. These are retained for immediate rows in the eventual LMM, but they cannot contribute delayed/decay information.</p>'
    data_html = participant_counts + condition_count_table(base_long, "Base complete observation rows", ["retention_score", "time"]) + covariate_feasibility_table(cov_long)
    final_models_html = final_models_wrapper(
        h1_confirmatory_models_html(rows),
        h1_retention_score_variation_models_html(),
        '<h4>Added integrated retention/decay model</h4><p class="small">This LMM uses immediate and delayed rows together. It is useful, but should be labelled as a sensitivity/extension because the preregistration specified separate HC3 models.</p>',
        fit_mixedlm_html(
            "Sensitivity LMM: retention and decay",
            base_long,
            "retention_score ~ time * required_pause_contrast + time * optional_pause_contrast",
        ),
        fit_mixedlm_html(
            "Sensitivity covariate-adjusted LMM: retention and decay",
            cov_long,
            "retention_score ~ time * required_pause_contrast + time * optional_pause_contrast + C(location) + co_present_participants + age + C(gender)",
        ),
    )
    return assumption_section_shell(
        "Hypothesis H1 + Y — Design → retention and decay",
        "Checkpoint Design is expected to affect immediate retention and, secondarily, delayed retention one week later; the same model should also diagnose retention decay.",
        ["Confirmatory preregistered route: separate HC3 linear models for immediate retention and delayed retention.", "Predictors: C1 (required pauses vs required continue), C2 (optional pauses vs system-controlled average), and retention-form order where available.", "Covariate sensitivity: add location, co-present participants, age, and gender when complete and feasible.", "Added sensitivity route: LMM for retention/time/decay with participant random intercept; label this as an extension, not as a replacement."],
        [
            {"Assumption": "Correct nesting", "How to test": "Check long format: one row per MCID per available time point; participant ID groups repeated scores.", "How to read": "Repeated observations from the same MCID must not be treated as independent."},
            {"Assumption": "Sufficient repeated data", "How to test": "Compare immediate and delayed observation counts and list missing delayed cases.", "How to read": "Missing delayed rows reduce decay information; LMM can still use immediate-only rows under missing-at-random assumptions."},
            {"Assumption": "Residual normality", "How to test": "Inspect Q-Q plot from the diagnostic proxy residuals.", "How to read": "This is only a proxy until the final LMM is fitted."},
            {"Assumption": "Homoscedasticity", "How to test": "Inspect residual-vs-fitted plot from the diagnostic proxy.", "How to read": "Funnel shape suggests unequal residual variance."},
            {"Assumption": "Outliers / influence", "How to test": "Inspect standardized residuals and Cook's D from the diagnostic proxy.", "How to read": "Severe cases should be reviewed but not automatically removed."},
        ],
        data_html,
        diagnostics,
        final_models_html,
        status_messages(errors, warnings, "H1/Y assumption checks, preregistered HC3 models, and sensitivity LMM models were generated."),
        combined_exclusion_details("Show H1/Y complete-case observation exclusions", base_excluded + cov_excluded),
    )


def mediation_diagnostics(rows: list[dict[str, Any]], mediator_columns: list[str], outcome_columns: list[str], prefix: str) -> tuple[str, list[dict[str, str]]]:
    html_parts: list[str] = []
    all_exclusions: list[dict[str, str]] = []
    for mediator in mediator_columns:
        diagnostic = fit_diagnostic_residuals(rows, mediator, ["required_pause_contrast", "optional_pause_contrast"], [])
        all_exclusions.extend(diagnostic.get("excluded", []))
        html_parts.append(diagnostics_block(f"{prefix} a-path diagnostic: {mediator} ~ C1 + C2", diagnostic))
    for outcome in outcome_columns:
        diagnostic = fit_diagnostic_residuals(rows, outcome, ["required_pause_contrast", "optional_pause_contrast", *mediator_columns], [])
        all_exclusions.extend(diagnostic.get("excluded", []))
        html_parts.append(diagnostics_block(f"{prefix} b/direct-path diagnostic: {outcome} ~ C1 + C2 + mediator(s)", diagnostic))
    return ''.join(html_parts), all_exclusions


def inferential_h2_assumptions() -> str:
    rows, warnings = build_rows_for_inferential_models()
    errors: list[str] = [] if rows else ["No participant-level rows could be built for H2 diagnostics."]
    mediator_cols = ["cl_intrinsic", "cl_extraneous", "cl_germane"]
    outcome_cols = ["ret_immediate_score", "ret_delayed_score"]
    base_complete, base_excluded = make_complete_cases(rows, ["required_pause_contrast", "optional_pause_contrast", *mediator_cols], "H2 mediator complete-case pool")
    diagnostics, diag_excluded = mediation_diagnostics(rows, mediator_cols, outcome_cols, "H2")
    data_html = condition_count_table(base_complete, "Mediator complete-case pool", mediator_cols) + covariate_feasibility_table(rows)
    final_models_html = final_models_wrapper(
        h2a_holm_table_html(rows, include_covariates=False),
        h2a_holm_table_html(rows, include_covariates=True),
        mediation_model_html("Final H2 base mediation: immediate retention", rows, mediator_cols, "ret_immediate_score", include_covariates=False),
        mediation_model_html("Final H2 base mediation: delayed retention", rows, mediator_cols, "ret_delayed_score", include_covariates=False),
        mediation_model_html("Final H2 covariate-adjusted sensitivity: immediate retention", rows, mediator_cols, "ret_immediate_score", include_covariates=True),
        mediation_model_html("Final H2 covariate-adjusted sensitivity: delayed retention", rows, mediator_cols, "ret_delayed_score", include_covariates=True),
    )
    return assumption_section_shell(
        "Hypotheses H2/H2a/H2b — Cognitive-load mediation",
        "Checkpoint Design is expected to have indirect effects on immediate and delayed retention through intrinsic, extraneous, and germane cognitive load.",
        ["Predictor: Checkpoint Design represented by C1 and C2 planned contrasts.", "Mediators: intrinsic, extraneous, and germane cognitive load, each continuous 0–10.", "Outcomes: immediate and delayed retention, continuous/bounded 0–2.", "Final mediation should use bootstrapped indirect effects; delayed-retention paths require delayed-retention complete cases."],
        [
            {"Assumption": "Correct causal ordering", "How to test": "Confirm that design preceded cognitive-load reports and retention outcomes.", "How to read": "Mediation is not purely statistical; temporal and theoretical ordering must be defensible."},
            {"Assumption": "Regression assumptions for each path", "How to test": "Run residual diagnostics for each a-path and b/direct-path component equation.", "How to read": "Interpret Q-Q plots, residual-vs-fitted plots, outlier counts, and VIF before final mediation."},
            {"Assumption": "No severe multicollinearity among parallel mediators", "How to test": "Inspect VIF in the outcome equation containing intrinsic, extraneous, and germane load together.", "How to read": "High VIF suggests unstable unique mediator paths."},
            {"Assumption": "Indirect-effect uncertainty", "How to test": "Use bootstrap CIs in the final model, not normal-theory Sobel logic.", "How to read": "An indirect effect is supported when the bootstrapped CI excludes zero."},
        ],
        data_html,
        diagnostics,
        final_models_html,
        status_messages(errors, warnings, "H2 assumption checks and final mediation models were generated."),
        combined_exclusion_details("Show H2 diagnostic exclusions", base_excluded + diag_excluded),
    )


def inferential_h3_assumptions() -> str:
    rows, warnings = build_rows_for_inferential_models()
    errors: list[str] = [] if rows else ["No participant-level rows could be built for H3 diagnostics."]
    mediator_cols = ["engagement"]
    outcome_cols = ["ret_immediate_score", "ret_delayed_score"]
    base_complete, base_excluded = make_complete_cases(rows, ["required_pause_contrast", "optional_pause_contrast", "engagement"], "H3 mediator complete-case pool")
    diagnostics, diag_excluded = mediation_diagnostics(rows, mediator_cols, outcome_cols, "H3")
    data_html = condition_count_table(base_complete, "Engagement complete-case pool", ["engagement"]) + covariate_feasibility_table(rows)
    final_models_html = final_models_wrapper(
        h3a_path_table_html(rows, include_covariates=False),
        h3_component_model_fit_html(rows, include_covariates=False),
        mediation_model_html("Final H3 base mediation: immediate retention", rows, mediator_cols, "ret_immediate_score", include_covariates=False),
        mediation_model_html("Final H3 base mediation: delayed retention", rows, mediator_cols, "ret_delayed_score", include_covariates=False),
        h3a_path_table_html(rows, include_covariates=True),
        h3_component_model_fit_html(rows, include_covariates=True),
        mediation_model_html("Final H3 covariate-adjusted sensitivity: immediate retention", rows, mediator_cols, "ret_immediate_score", include_covariates=True),
        mediation_model_html("Final H3 covariate-adjusted sensitivity: delayed retention", rows, mediator_cols, "ret_delayed_score", include_covariates=True),
    )
    return assumption_section_shell(
        "Hypotheses H3/H3a/H3b — Engagement mediation",
        "Checkpoint Design is expected to have indirect effects on immediate and delayed retention through post-game reported engagement.",
        ["Predictor: Checkpoint Design represented by C1 and C2 planned contrasts.", "Mediator: engagement, continuous 1–7.", "Outcomes: immediate and delayed retention, continuous/bounded 0–2.", "Final mediation should use bootstrapped indirect effects; delayed-retention paths require delayed-retention complete cases."],
        [
            {"Assumption": "Correct causal ordering", "How to test": "Confirm that design preceded engagement reports and retention outcomes.", "How to read": "Mediation wording should remain cautious if ordering or omitted confounding is uncertain."},
            {"Assumption": "Regression assumptions for each path", "How to test": "Run residual diagnostics for a-path and b/direct-path component equations.", "How to read": "Interpret Q-Q plots, residual-vs-fitted plots, outlier counts, and VIF before final mediation."},
            {"Assumption": "Outcome scale/range", "How to test": "Check retention remains within 0–2 and engagement within 1–7.", "How to read": "Out-of-range values are data errors or coding issues."},
            {"Assumption": "Indirect-effect uncertainty", "How to test": "Use bootstrap CIs in the final model.", "How to read": "An indirect effect is supported when the bootstrapped CI excludes zero."},
        ],
        data_html,
        diagnostics,
        final_models_html,
        status_messages(errors, warnings, "H3 assumption checks and final mediation models were generated."),
        combined_exclusion_details("Show H3 diagnostic exclusions", base_excluded + diag_excluded),
    )


def h4_scatter_diagnostics(rows: list[dict[str, Any]]) -> tuple[str, list[dict[str, str]]]:
    outputs: list[str] = []
    all_exclusions: list[dict[str, str]] = []
    for outcome in ["cl_intrinsic", "cl_extraneous", "cl_germane"]:
        complete, excluded = make_complete_cases(rows, ["engagement", outcome], f"H4 engagement with {outcome}")
        all_exclusions.extend(excluded)
        x_values = [float(parse_float(row.get("engagement"))) for row in complete]
        y_values = [float(parse_float(row.get(outcome))) for row in complete]
        outlier_count = 0
        if len(x_values) > 2 and len(y_values) > 2:
            x_sd = np.std(x_values, ddof=1) or 1
            y_sd = np.std(y_values, ddof=1) or 1
            zx = (np.array(x_values) - np.mean(x_values)) / x_sd
            zy = (np.array(y_values) - np.mean(y_values)) / y_sd
            outlier_count = int(np.sum((np.abs(zx) > 3) | (np.abs(zy) > 3)))
        tests = [
            {"Assumption": "Linearity", "How to read": "Scatter should look approximately straight rather than curved.", "Diagnostic": "Inspect scatterplot below.", "Flag": '<span class="badge badge-neutral">visual check</span>'},
            {"Assumption": "Outliers", "How to read": "Points with |z| > 3 on either variable should be inspected.", "Diagnostic": f"Potential univariate outlier pairs: {outlier_count}", "Flag": interpretation_badge(outlier_count == 0, "no severe cases", "inspect cases")},
        ]
        outputs.append(f'<details class="compact-details" open><summary>H4 diagnostic pair: engagement with {h(outcome)} · n={len(complete)}</summary>' + diagnostic_tests_table(tests) + simple_scatter_svg("h4-" + outcome, x_values, y_values, f"Engagement with {outcome}", "Engagement", outcome) + '</details>')
    return ''.join(outputs), all_exclusions


def inferential_h4_assumptions() -> str:
    rows, warnings = build_rows_for_inferential_models()
    errors: list[str] = [] if rows else ["No participant-level rows could be built for H4 diagnostics."]
    complete, base_excluded = make_complete_cases(rows, ["engagement", "cl_intrinsic", "cl_extraneous", "cl_germane"], "H4 complete-case pool")
    diagnostics, diag_excluded = h4_scatter_diagnostics(rows)
    data_html = condition_count_table(complete, "H4 complete-case pool", ["engagement", "cl_intrinsic", "cl_extraneous", "cl_germane"]) + covariate_feasibility_table(rows)
    final_models_html = final_models_wrapper(
        correlation_model_html(rows, "engagement", ["cl_intrinsic", "cl_extraneous", "cl_germane"]),
    )
    return assumption_section_shell(
        "Hypothesis H4 — Engagement and cognitive load",
        "Engagement is expected to relate negatively to extraneous cognitive load and positively to germane cognitive load; its association with intrinsic load is examined without a directional prediction.",
        ["Variables: engagement, intrinsic load, extraneous load, and germane load are continuous floats.", "Primary final model is correlation, one pair at a time.", "Optional sensitivity could use multiple regression/partial-correlation logic with condition contrasts and covariates, but this is not the primary H4 test."],
        [
            {"Assumption": "Independent paired observations", "How to test": "Each MCID contributes one engagement score and one cognitive-load score per pair.", "How to read": "No participant should appear twice in the same pairwise correlation."},
            {"Assumption": "Linearity", "How to test": "Inspect scatterplots for engagement with each cognitive-load dimension.", "How to read": "Pearson correlation is appropriate when the cloud is approximately linear."},
            {"Assumption": "Outliers", "How to test": "Inspect scatterplots and z-score outlier counts.", "How to read": "Extreme points can dominate correlations and should be investigated."},
            {"Assumption": "Approximate normality / robust alternative", "How to test": "Inspect distributions; if strongly non-normal, consider bootstrapped Pearson CI or Spearman as sensitivity.", "How to read": "This affects p-values/CIs more than the descriptive scatter itself."},
        ],
        data_html,
        diagnostics,
        final_models_html,
        status_messages(errors, warnings, "H4 assumption checks and final correlations were generated."),
        combined_exclusion_details("Show H4 pairwise exclusions", base_excluded + diag_excluded),
    )


def z_dropout_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in rows:
        out = dict(row)
        out["completed_delayed"] = "Completed delayed retention" if parse_float(row.get("ret_delayed_score")) is not None else "Dropped before delayed retention"
        output.append(out)
    return output


def dropout_observed_expected_table(rows: list[dict[str, Any]]) -> str:
    categories = ["Completed delayed retention", "Dropped before delayed retention"]
    counts = {(condition, category): 0 for condition in CONDITION_ORDER for category in categories}
    for row in rows:
        if row.get("condition") in CONDITION_ORDER and row.get("completed_delayed") in categories:
            counts[(row.get("condition"), row.get("completed_delayed"))] += 1
    row_totals = {condition: sum(counts[(condition, category)] for category in categories) for condition in CONDITION_ORDER}
    col_totals = {category: sum(counts[(condition, category)] for condition in CONDITION_ORDER) for category in categories}
    grand_total = sum(row_totals.values())
    body: list[str] = []
    small_expected = 0
    for condition in CONDITION_ORDER:
        cells = [f'<th>{h(DISPLAY_CONDITION[condition])}</th>']
        for category in categories:
            expected = row_totals[condition] * col_totals[category] / grand_total if grand_total else 0
            if expected < 5:
                small_expected += 1
            cells.append(f'<td>Observed: {counts[(condition, category)]}<br>Expected: {expected:.2f}</td>')
        body.append('<tr>' + ''.join(cells) + '</tr>')
    read = "Expected counts look acceptable." if small_expected == 0 else f"{small_expected} expected cell count(s) below 5; inspect before running chi-square."
    return '<h4>Dropout contingency table for assumption checking only</h4><p class="small">This table displays observed and expected counts but does not run the chi-square test.</p><div class="table-wrap"><table><thead><tr><th>Design condition</th>' + ''.join(f'<th>{h(category)}</th>' for category in categories) + '</tr></thead><tbody>' + ''.join(body) + f'</tbody></table></div><p>{h(read)}</p>'


def inferential_z_assumptions() -> str:
    rows, warnings = build_rows_for_inferential_models()
    dropout_rows = z_dropout_rows(rows)
    errors: list[str] = [] if dropout_rows else ["No participant-level rows could be built for dropout diagnostics."]
    data_html = condition_count_table(dropout_rows, "Dropout status by condition", ["completed_delayed"]) + dropout_observed_expected_table(dropout_rows) + covariate_feasibility_table(dropout_rows)
    diagnostics = ""
    for variable in ["age", "ret_immediate_score", "perceived_control", "engagement", "cl_intrinsic", "cl_extraneous", "cl_germane"]:
        complete, _excluded = make_complete_cases(dropout_rows, ["completed_delayed", variable], f"Z {variable}")
        groups = {label: [float(parse_float(row.get(variable))) for row in complete if row.get("completed_delayed") == label and parse_float(row.get(variable)) is not None] for label in ["Completed delayed retention", "Dropped before delayed retention"]}
        levene_p = None
        shapiro_bits: list[str] = []
        if scipy_stats is not None and all(len(values) >= 2 for values in groups.values()):
            try:
                levene_p = float(scipy_stats.levene(*groups.values()).pvalue)
            except Exception:
                levene_p = None
        for label, values in groups.items():
            sp = None
            if scipy_stats is not None and 3 <= len(values) <= 5000:
                try:
                    sp = float(scipy_stats.shapiro(values).pvalue)
                except Exception:
                    sp = None
            shapiro_bits.append(f"{label}: Shapiro p={p_text(sp)} (n={len(values)})")
        tests = [
            {"Assumption": "Normality within dropout groups", "How to read": "Use Q-Q/histograms in final follow-up; Shapiro p ≥ .05 is reassuring but sensitive to sample size.", "Diagnostic": "; ".join(shapiro_bits), "Flag": '<span class="badge badge-neutral">inspect</span>'},
            {"Assumption": "Homogeneity of variance", "How to read": "Levene p ≥ .05 supports equal variances; otherwise use Welch t-test if a follow-up t-test is run.", "Diagnostic": f"Levene p = {p_text(levene_p)}", "Flag": interpretation_badge(None if levene_p is None else levene_p >= .05, "equal variance plausible", "use Welch/inspect")},
        ]
        diagnostics += f'<details class="compact-details"><summary>Z continuous follow-up diagnostics: {h(variable)} · n={len(complete)}</summary>' + diagnostic_tests_table(tests) + '</details>'
    final_models_html = final_models_wrapper(
        chi_square_table_html(dropout_rows, "condition", "completed_delayed", "Final primary dropout test: design condition × delayed-completion status"),
        chi_square_table_html(dropout_rows, "location", "completed_delayed", "Dropout sensitivity: location × delayed-completion status"),
        chi_square_table_html(dropout_rows, "gender", "completed_delayed", "Dropout sensitivity: gender × delayed-completion status"),
        dropout_ttest_html(dropout_rows, ["age", "co_present_participants", "ret_immediate_score", "perceived_control", "engagement", "cl_intrinsic", "cl_extraneous", "cl_germane"]),
    )
    return assumption_section_shell(
        "Hypothesis Z — Dropout analysis",
        "Dropout analysis checks whether delayed-retention completion appears random with respect to design condition and key participant characteristics.",
        ["Primary dropout variable: completed delayed retention vs. dropped before delayed retention.", "Primary predictor: three-level design condition.", "Follow-up variables: continuous baseline/study variables can be compared between completers and dropouts; categorical variables use association tables."],
        [
            {"Assumption": "Frequency data for chi-square", "How to test": "Build Design × delayed-completion table.", "How to read": "Each MCID should contribute to exactly one cell."},
            {"Assumption": "Expected cell counts", "How to test": "Inspect expected frequencies in the contingency table.", "How to read": "Expected counts should generally be ≥ 5 for chi-square approximation."},
            {"Assumption": "Independent observations", "How to test": "Check that each MCID appears once in dropout status.", "How to read": "Duplicated MCIDs would invalidate the table."},
            {"Assumption": "Follow-up t-test assumptions", "How to test": "Inspect normality within completer/dropout groups and Levene's test for variance equality.", "How to read": "If variances are unequal, use Welch version for any final t-test follow-up."},
        ],
        data_html,
        diagnostics,
        final_models_html,
        status_messages(errors, warnings, "Z assumption checks and final dropout models were generated."),
        combined_exclusion_details("Show Z exclusions", []),
    )



# -----------------------------------------------------------------------------
# HTML REPORT ASSEMBLY
# -----------------------------------------------------------------------------


def preregistration_alignment_section() -> str:
    """Human-facing stage-1 alignment table: planned, implemented, and label to use."""
    IN_BODY = True
    rows = [
        ("Manipulation and hypotheses", "Three checkpoint-design conditions; H1-H4/EQ1-EQ3.", "Kept.", "Confirmatory / planned."),
        ("H1 retention model", "Separate HC3 linear models for immediate and delayed retention.", "Now printed first; integrated LMM retained below as sensitivity/extension.", "Confirmatory first; sensitivity second."),
        ("Retention score source", "Retention scores derived from administered rubric rows.", "Read directly from retention_scores_final.tsv for participant-level analyses; retention_scores_merged.tsv is used only for scoring/agreement diagnostics.", "Revised implementation; document clearly."),
        ("Retention reliability", "Quadratic weighted Cohen's kappa for a predefined subset.", "Agreement table includes ordinal Krippendorff's alpha, exact agreement, and pairwise weighted kappa.", "Improved but deviates; note in manuscript."),
        ("H2a and H4 multiplicity", "Holm correction for relevant test families.", "Raw and Holm-adjusted p-values are now shown in the Results-template report.", "Confirmatory correction visible."),
        ("Covariate sensitivity", "Age/gender and contextual variables retained for sensitivity/robustness where feasible.", "Location is CreaSp/LivingR/Remote; Remote co-present participants = 0; age numeric; gender categorical.", "Sensitivity / robustness."),
    ]
    body = "".join(f"<tr><td>{h(a)}</td><td>{h(b)}</td><td>{h(c)}</td><td>{h(d)}</td></tr>" for a,b,c,d in rows)
    table_html = '<div class="table-wrap"><table><thead><tr><th>Decision area</th><th>Preregistered / planned</th><th>Implemented in this report</th><th>Manuscript label</th></tr></thead><tbody>' + body + '</tbody></table></div>'
    return table_shell(
        "Preregistration alignment and deviations",
        IN_BODY,
        "This table is intentionally placed before inferential tests so the reader can separate preregistered analyses from improved implementation choices and sensitivity analyses.",
        table_html,
        "",
        status_messages([], [], "Alignment table generated. Use it as a checklist for manuscript deviation notes; do not copy it wholesale into the Results unless needed."),
        "",
    )

def html_document(sections: list[str]) -> str:
    generated = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    css = """
    :root { --bg:#f6f7f8; --card:#ffffff; --text:#172026; --muted:#5f6c73; --line:#d9e0e4; --blue:#1d4ed8; --brown:#92400e; --green-bg:#ecfdf3; --green:#027a48; --orange-bg:#fff7ed; --orange:#b54708; --red-bg:#fef3f2; --red:#b42318; --neutral-bg:#f2f4f7; --neutral:#344054; }
    * { box-sizing:border-box; }
    html { scroll-behavior:smooth; }
    body { margin:0; background:var(--bg); color:var(--text); font-family:Arial, Helvetica, sans-serif; line-height:1.45; }
    .page-layout { display:grid; grid-template-columns:280px minmax(0, 1fr); gap:18px; max-width:1580px; margin:0 auto; padding:18px; }
    .side-toc { position:sticky; top:18px; align-self:start; max-height:calc(100vh - 36px); overflow:auto; background:var(--card); border:1px solid var(--line); border-radius:14px; padding:14px; }
    .side-toc h2 { font-size:16px; margin-bottom:8px; }
    .side-toc h3 { font-size:13px; margin:14px 0 6px; color:var(--muted); text-transform:uppercase; letter-spacing:.04em; }
    .side-toc ol, .side-toc ul { margin:0 0 0 18px; padding:0; }
    .side-toc li { margin:6px 0; font-size:13px; }
    .side-toc a { color:#1d4ed8; text-decoration:none; display:block; padding:3px 6px; margin-left:-6px; border-radius:8px; }
    .side-toc a:hover { text-decoration:underline; }
    .side-toc a.active { background:#dbeafe; color:#0f172a; font-weight:700; text-decoration:none; }
    main { min-width:0; }
    h1, h2, h3, h4 { margin:0 0 10px 0; }
    p { margin:0 0 10px 0; }
    .card { background:var(--card); border:1px solid var(--line); border-radius:14px; padding:16px; margin-bottom:16px; scroll-margin-top:18px; }
    .body-output h2, .appendix-output h2 { color:#fff; margin:-16px -16px 14px -16px; padding:12px 16px; border-radius:14px 14px 0 0; display:flex; justify-content:space-between; gap:12px; align-items:center; }
    .body-output h2 { background:var(--blue); }
    .appendix-output h2 { background:var(--brown); }
    h2 span { font-size:12px; font-weight:700; opacity:.9; }
    .small { color:var(--muted); font-size:12px; }
    .table-wrap { overflow:auto; margin:12px 0; }
    table { border-collapse:collapse; width:100%; background:#fff; font-size:13px; }
    th, td { border:1px solid var(--line); padding:8px 9px; vertical-align:top; text-align:left; }
    th { background:#eef2f7; }
    .micro-stat-table { width:auto; font-size:12px; background:transparent; }
    .micro-stat-table th, .micro-stat-table td { border:0; padding:1px 8px 1px 0; background:transparent; }
    .chart-box { border:1px solid var(--line); border-radius:12px; padding:12px; background:white; margin:14px 0; }
    .standalone-figure { width:100%; max-width:960px; display:block; }
    .clickable-box { cursor:pointer; }
    .clickable-box:hover { opacity:.75; }
    .boxplot-detail { border-left:3px solid var(--line); padding-left:9px; }
    .status { border-radius:10px; padding:9px 11px; margin:9px 0; border:1px solid transparent; }
    .status-green { background:var(--green-bg); color:var(--green); border-color:#abefc6; }
    .status-orange { background:var(--orange-bg); color:var(--orange); border-color:#fed7aa; }
    .status-red { background:var(--red-bg); color:var(--red); border-color:#fecdca; }
    .badge { display:inline-block; border-radius:999px; padding:2px 8px; font-size:12px; font-weight:700; }
    .badge-good { background:var(--green-bg); color:var(--green); border:1px solid #abefc6; }
    .badge-warning { background:var(--orange-bg); color:var(--orange); border:1px solid #fed7aa; }
    .badge-neutral { background:var(--neutral-bg); color:var(--neutral); border:1px solid #d0d5dd; }
    details { margin:10px 0; }
    summary { cursor:pointer; font-weight:700; }
    .compact-details { border:1px solid var(--line); border-radius:12px; background:#fff; padding:0; }
    .compact-details summary { padding:9px 12px; }
    .compact-details[open] summary { border-bottom:1px solid var(--line); }
    .compact-details .table-wrap { padding:0 12px 12px; }
    .model-table { font-size:12px; }
    .significant-row > th, .significant-row > td { background:var(--green-bg); }
    @media (max-width: 980px) { .page-layout { display:block; padding:12px; } .side-toc { position:relative; top:0; max-height:none; margin-bottom:14px; } }
    """
    js = """
    function showBoxplotDetails(id, text) {
      const element = document.getElementById(id + '-detail');
      if (element) element.textContent = text;
    }

    function updateActiveTocLink() {
      const links = Array.from(document.querySelectorAll('.side-toc a[href^="#"]'));
      const sections = links
        .map(link => ({ link, section: document.getElementById(decodeURIComponent(link.getAttribute('href').slice(1))) }))
        .filter(item => item.section);
      if (!sections.length) return;
      let current = sections[0];
      let bestTop = -Infinity;
      for (const item of sections) {
        const top = item.section.getBoundingClientRect().top;
        if (top <= 120 && top > bestTop) {
          bestTop = top;
          current = item;
        }
      }
      if (bestTop === -Infinity) {
        current = sections.reduce((best, item) => {
          const distance = Math.abs(item.section.getBoundingClientRect().top - 120);
          const bestDistance = Math.abs(best.section.getBoundingClientRect().top - 120);
          return distance < bestDistance ? item : best;
        }, sections[0]);
      }
      links.forEach(link => link.classList.remove('active'));
      current.link.classList.add('active');
      current.link.scrollIntoView({ block: 'nearest' });
    }

    document.addEventListener('DOMContentLoaded', updateActiveTocLink);
    window.addEventListener('scroll', updateActiveTocLink, { passive: true });
    window.addEventListener('resize', updateActiveTocLink);
    """
    toc = f"""
    <aside class="side-toc" aria-label="Table of contents">
      <h2>Table of contents</h2>
      <p class="small">Sticky navigation. The highlighted item tracks the current section.</p>
      <h3>Start</h3>
      <ul><li><a href="#report-intro">Report intro</a></li></ul>
      <h3>Scoring + descriptives</h3>
      <ol>
        <li><a href="#{slugify('Retention scoring agreement')}">Retention scoring</a></li>
        <li><a href="#{slugify('Descriptives: location')}">Location</a></li>
        <li><a href="#{slugify('Descriptives: co-present participants')}">Co-present participants</a></li>
        <li><a href="#{slugify('Descriptives: age')}">Age</a></li>
        <li><a href="#{slugify('Descriptives: gender')}">Gender</a></li>
        <li><a href="#{slugify('Descriptives: creatures seen')}">Creatures seen</a></li>
        <li><a href="#{slugify('Descriptives: perceived control')}">Perceived control</a></li>
        <li><a href="#{slugify('Descriptives: immediate retention')}">Immediate retention</a></li>
        <li><a href="#{slugify('Descriptives: delayed retention')}">Delayed retention</a></li>
        <li><a href="#{slugify('Descriptives: retention decay')}">Retention decay</a></li>
        <li><a href="#{slugify('Descriptives: intrinsic cognitive load')}">Intrinsic load</a></li>
        <li><a href="#{slugify('Descriptives: intrinsic cognitive load by chapter')}">Intrinsic load by chapter</a></li>
        <li><a href="#{slugify('Descriptives: extraneous cognitive load')}">Extraneous load</a></li>
        <li><a href="#{slugify('Descriptives: extraneous cognitive load by chapter')}">Extraneous load by chapter</a></li>
        <li><a href="#{slugify('Descriptives: germane cognitive load')}">Germane load</a></li>
        <li><a href="#{slugify('Descriptives: engagement')}">Engagement</a></li>
        <li><a href="#{slugify('Descriptives: engagement by chapter')}">Engagement by chapter</a></li>
        <li><a href="#{slugify('Internal consistency: cognitive load and engagement')}">Internal consistency</a></li>
      </ol>
      <h3>Alignment + assumptions + final models</h3>
      <ol>
        <li><a href="#{slugify('Preregistration alignment and deviations')}">Preregistration alignment</a></li>
        <li><a href="#{slugify('Hypothesis X — Design → perceived control')}">X: perceived control</a></li>
        <li><a href="#{slugify('Hypothesis H1 + Y — Design → retention and decay')}">H1 + Y: retention</a></li>
        <li><a href="#{slugify('Hypotheses H2/H2a/H2b — Cognitive-load mediation')}">H2: cognitive-load mediation</a></li>
        <li><a href="#{slugify('Hypotheses H3/H3a/H3b — Engagement mediation')}">H3: engagement mediation</a></li>
        <li><a href="#{slugify('Hypothesis H4 — Engagement and cognitive load')}">H4: engagement/load</a></li>
        <li><a href="#{slugify('Hypothesis Z — Dropout analysis')}">Z: dropout</a></li>
        <li><a href="#{slugify('Interview subsample representativeness')}">Interviews: representativeness</a></li>
      </ol>
    </aside>
    """
    intro = (
        '<section id="report-intro" class="card"><h1>Statistics manuscript report</h1>'
        f'<p class="small">Generated {h(generated)} from {h(ANALYSIS_DIR)}.</p>'
        '<p>This standalone report is generated by <code>statistics_manuscript.py</code>. It uses complete-case preparation for each individual table or diagnostic section.</p>'
        '<p><strong>Important:</strong> the inferential part now shows preregistration alignment, assumption checks, confirmatory preregistered models, and then sensitivity/exploratory extensions. The order is deliberate.</p>'
        '</section>'
    )
    return (
        '<!doctype html><html lang="en"><head><meta charset="utf-8">'
        '<meta name="viewport" content="width=device-width, initial-scale=1">'
        '<title>Statistics manuscript report</title>'
        f'<style>{css}</style></head><body><div class="page-layout">'
        + toc + '<main>' + intro
        + "".join(sections)
        + f'<script>{js}</script></main></div></body></html>'
    )


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    sections = [
        retention_scoring(),
        descriptive_location(),
        descriptive_co_present(),
        descriptive_age(),
        descriptive_gender(),
        descriptive_creatures_seen(),
        descriptive_perceived_control(),
        descriptive_retention("Immediate"),
        descriptive_retention("Delayed"),
        descriptive_retention_decay(),
        descriptive_intrinsic_load(),
        descriptive_intrinsic_load_by_chapter(),
        descriptive_extraneous_load(),
        descriptive_extraneous_load_by_chapter(),
        descriptive_germane_load(),
        descriptive_engagement(),
        descriptive_engagement_by_chapter(),
        internal_consistency_cognitive_load_engagement(),
        preregistration_alignment_section(),
        inferential_x_assumptions(),
        inferential_h1_y_assumptions(),
        inferential_h2_assumptions(),
        inferential_h3_assumptions(),
        inferential_h4_assumptions(),
        inferential_z_assumptions(),
        inferential_interview_representativeness(),
    ]
    STATISTICS_MANUSCRIPT_OUTPUT_PATH.write_text(html_document(sections), encoding="utf-8")
    print(f"Wrote {STATISTICS_MANUSCRIPT_OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
